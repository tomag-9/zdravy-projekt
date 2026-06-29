#!/usr/bin/env python3
"""
Generate Excel meal plan input template for Zdravý projekt.
Defaults to week 22/2026 (25.5.2026 - 29.5.2026).

Run: python3 scripts/generate_jedalnicky_template.py --week 22 --year 2026
Output: test/jedalnicky/jedalnicky_template_ukazka_W22_2026.xlsx

Clean template:
python3 scripts/generate_jedalnicky_template.py --week 22 --year 2026 --blank
Output: test/jedalnicky/jedalnicky_template_W22_2026.xlsx
"""

import argparse
import copy
import datetime as dt
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── colours ───────────────────────────────────────────────────────────────────
C_TITLE_BG = "1F4E79"  # navy
C_TITLE_FG = "FFFFFF"
C_DAY_BG = "2E75B6"  # mid-blue
C_DAY_FG = "FFFFFF"
C_MEAL_BG = "BDD7EE"  # light blue
C_MEAL_FG = "1F4E79"
C_INPUT_BG = "FFF2CC"  # yellow - manual input
C_ALT_BG = "F5F5F5"  # light grey alternating row
C_NOTE_BG = "EDEDED"  # grey
C_UNIT_BG = "E8F0FE"  # light blue-grey for unit column
C_RED_NOTE = "C00000"

PORTION_TYPES = [
    "Jasle",
    "Škôlka",
    "ZŠ 1.stupeň",
    "ZŠ 2.stupeň",
    "Dospelý (SŠ)",
]
BLANK_COMPONENT_ROWS_PER_MEAL = 4


# ── helpers ───────────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)


def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)


def border_thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def derive_unit(note):
    n = note.lower()
    if "ks" in n:
        return "ks"
    return "g"


def parse_args():
    parser = argparse.ArgumentParser(
        description="Generate jedalnicky XLSX input template."
    )
    parser.add_argument("--week", type=int, default=22, help="ISO week number.")
    parser.add_argument("--year", type=int, default=2026, help="ISO week year.")
    parser.add_argument(
        "--output",
        help="Output XLSX path. Defaults to test/jedalnicky/jedalnicky_template[_ukazka]_Wxx_yyyy.xlsx.",
    )
    parser.add_argument(
        "--blank",
        action="store_true",
        help="Generate a clean empty template without example meal rows or notes.",
    )
    return parser.parse_args()


def week_context(year, week):
    start = dt.date.fromisocalendar(year, week, 1)
    end = dt.date.fromisocalendar(year, week, 5)
    return {
        "week": week,
        "year": year,
        "start": start,
        "end": end,
        "label": f"Týždeň {week}/{year}",
        "date_range": f"{start.day}.{start.month}.{start.year} - {end.day}.{end.month}.{end.year}",
    }


# ── meal data ─────────────────────────────────────────────────────────────────
# Each entry: (meal_type, component_name, klasik_grams, alergeny, note)
# meal_type: 'Raňajky-desiata' | 'Polievka' | 'Obed' | 'Obed Menu B' | 'Obed Menu C' | 'Olovrant'
# klasik_grams: kept for data continuity but never rendered in the template

PON = "PONDELOK"
UTO = "UTOROK"
STR = "STREDA"
STV = "ŠTVRTOK"
PIA = "PIATOK"

DAYS = [PON, UTO, STR, STV, PIA]

R = "Raňajky-desiata"
P = "Polievka"
OA = "Obed"
OB = "Obed Menu B"
OC = "Obed Menu C"
OL = "Olovrant"

# ── KLASIK (Menu A) ───────────────────────────────────────────────────────────
KLASIK = {
    PON: [
        (R, "Čučoriedkový tvaroháčik", 110, "7", ""),
        (R, "Sladké krutóny", 10, "1, 3", ""),
        (R, "Zelené jablko", 50, "", ""),
        (P, "Paradajková polievka s bazalkou a ryžou", 200, "", "ml"),
        (OA, "Tekvicový prívarok", 90, "7", ""),
        (OA, "Varené zemiaky s pažítkou", 110, "", ""),
        (OA, "Vajíčko", 25, "3", ""),
        (OL, "Grahamové pečivo", 50, "1", ""),
        (OL, "Hydinová nátierka", 25, "7, 10", ""),
    ],
    UTO: [
        (R, "Pšenová kaša so slivkami a makom", 120, "7", ""),
        (R, "Pomaranč", 70, "", ""),
        (P, "Karfiolová krémová s cuketou a pohánkou", 200, "7", "ml"),
        (OA, "Boloňská omáčka s mletým hovädzím mäsom", 90, "9", ""),
        (OA, "Špagety", 110, "1", ""),
        (OA, "Syr", 10, "7", ""),
        (OL, "Kváskový chlieb", 50, "1", ""),
        (OL, "Nátierka z pečeného baklažánu", 25, "7", ""),
    ],
    STR: [
        (R, "Kváskový chlieb", 50, "1", ""),
        (R, "Paprikové maslo", 15, "7", ""),
        (R, "Vajce", 25, "3", ""),
        (R, "Mrkva", 50, "", ""),
        (P, "Zeleninový vývar s písmenkami", 200, "1, 9", "ml"),
        (OA, "Pečené kuracie stehno", None, "", "1 ks"),
        (OA, "Zemiaková kaša", 110, "", ""),
        (OA, "Broskyňový kompót", 25, "", ""),
        (OL, "Zebra koláč", 75, "1, 3, 7", ""),
    ],
    STV: [
        (R, "Ryžový puding s jahodami", 115, "7, 8", ""),
        (R, "Pečené mandle (topping)", 5, "8", ""),
        (R, "Banán", 70, "", ""),
        (P, "Vege detská kapustnica so zemiakmi", 200, "9", "ml"),
        (OA, "Morčacie na smotane s hráškom a paprikou", 120, "7", ""),
        (OA, "Ryža s bulgurom", 80, "1", ""),
        (OL, "Pšeničné pečivo", 50, "1", ""),
        (OL, "Cviklová nátierka s chrenom", 25, "7", ""),
        (OL, "Zelenina", 25, "", ""),
    ],
    PIA: [
        (R, "Kváskový chlieb", 50, "1", ""),
        (R, "Nátierka z pečenej cibule", 25, "7", ""),
        (R, "Paprika", 50, "", ""),
        (P, "Hydinový vývar s niťovkami", 200, "1, 9", "ml"),
        (OA, "Detské šošovicové ragú", 90, "9", ""),
        (OA, "Jasmínová ryža", 110, "", ""),
        (OA, "Zelený šalát s paradajkou a paprikou", 25, "", ""),
        (OL, "Špenátový koláč so syrom", 75, "1, 3, 7", ""),
    ],
}

# ── KLASIK Menu B - only rows that differ from Menu A ─────────────────────────
KLASIK_MENU_B = {
    PON: [
        (OB, "Granadírsky pochod", None, "1", "Menu B, 275 g spolu"),
        (OB, "Miešaný šalát", None, "", "Menu B"),
    ],
    UTO: [
        (OB, "Zapekané cestoviny s kuracím mäsom", None, "1, 7", "Menu B, 275 g spolu"),
        (OB, "Uhorkový šalát", None, "", "Menu B"),
    ],
    STR: [
        (OB, "Morčacie na bratislavský spôsob", None, "1", "Menu B, 275 g spolu"),
        (OB, "Ryža", None, "", "Menu B"),
        (OB, "Broskyňový kompót", None, "", "Menu B"),
    ],
    STV: [
        (OB, "Hovädzie v paradajkovej omáčke", None, "", "Menu B, 275 g spolu"),
        (OB, "Varené zemiaky", None, "", "Menu B"),
        (OB, "Miešaný šalát", None, "", "Menu B"),
    ],
    PIA: [
        (
            OB,
            "Šúľance s makom a ovocným rozvarom",
            None,
            "1, 3, 7",
            "Menu B, 275 g spolu",
        )
    ],
}

# ── VEGE ──────────────────────────────────────────────────────────────────────
VEGE = copy.deepcopy(KLASIK)
# Mon olovrant: syrová nátierka instead of hydinová
VEGE[PON][-1] = (OL, "Syrová nátierka", 25, "7, 10", "")
# Tue obed: nemäso
VEGE[UTO][3] = (OA, "Boloňská omáčka s nemäsom", 90, "9", "")
# Wed obed: zeleninové guličky instead of kuracie
VEGE[STR][5] = (OA, "Zeleninové guličky", None, "6", "1 ks / g?")
# Thu obed: sójové mäso
VEGE[STV][4] = (OA, "Sójové 'mäso' na smotane s hráškom a paprikou", 120, "6, 7", "")
# Fri polievka: zeleninová
VEGE[PIA][3] = (P, "Zeleninová polievka s cestovinami", 200, "1, 9", "ml")

# ── NO GLUTEN ────────────────────────────────────────────────────────────────
NOGLUTEN = copy.deepcopy(KLASIK)
# Mon R-D: bezlepkové krutóny (no alerg 1)
NOGLUTEN[PON][1] = (R, "Sladké bezlepkové krutóny", 10, "3", "")
# Mon olovrant: bezlepkové pečivo
NOGLUTEN[PON][-2] = (OL, "Bezlepkové pečivo", 50, "", "")
# Tue obed: bezlepkové cestoviny
NOGLUTEN[UTO][4] = (OA, "Bezlepkové cestoviny", 110, "", "")
# Tue olovrant: bezlepkový chlieb
NOGLUTEN[UTO][-2] = (OL, "Bezlepkový chlieb", 50, "", "")
# Wed R-D: bezlepkový chlieb
NOGLUTEN[STR][0] = (R, "Bezlepkový chlieb", 50, "", "")
# Wed polievka: bezlepkové cestoviny
NOGLUTEN[STR][4] = (P, "Zeleninový vývar s bezlepkovými cestovinami", 200, "9", "ml")
# Wed olovrant: bezlepkový zebra koláč
NOGLUTEN[STR][-1] = (OL, "Bezlepkový zebra koláč", 75, "3, 7", "")
# Thu obed: ryža (bez bulguru)
NOGLUTEN[STV][5] = (OA, "Jasmínová ryža", 80, "", "bez bulguru")
# Thu olovrant: bezlepkové pečivo
NOGLUTEN[STV][-3] = (OL, "Bezlepkové pečivo", 50, "", "")
# Fri R-D: bezlepkový chlieb
NOGLUTEN[PIA][0] = (R, "Bezlepkový chlieb", 50, "", "")
# Fri polievka: bezlepkové cestoviny
NOGLUTEN[PIA][3] = (P, "Hydinový vývar s bezlepkovými cestovinami", 200, "9", "ml")
# Fri olovrant: bezlepkový špenátový koláč
NOGLUTEN[PIA][-1] = (OL, "Bezlepkový špenátový koláč so syrom", 75, "3, 7", "")

# ── NO MILK ───────────────────────────────────────────────────────────────────
NOMILK = copy.deepcopy(KLASIK)
# Mon R-D: rastlinný tvaroháčik
NOMILK[PON][0] = (R, "Rastlinný čučoriedkový tvaroháčik", 110, "", "")
# Mon obed: rastlinná smotana
NOMILK[PON][4] = (OA, "Tekvicový prívarok s rastlinnou smotanou", 90, "", "")
# Mon olovrant: rastlinná nátierka
NOMILK[PON][-1] = (OL, "Rastlinná hydinová nátierka", 25, "10", "")
# Tue R-D: rastlinné mlieko
NOMILK[UTO][0] = (R, "Pšenová kaša s rastlinným mliekom, slivkami a makom", 120, "", "")
# Tue polievka: rastlinné mlieko
NOMILK[UTO][2] = (
    P,
    "Karfiolová krémová s rastlinným mliekom, cuketou a pohánkou",
    200,
    "",
    "ml",
)
# Tue obed: rastlinný syr
NOMILK[UTO][5] = (OA, "Rastlinný syr", 10, "", "")
# Tue olovrant: rastlinná nátierka
NOMILK[UTO][-1] = (OL, "Rastlinná nátierka z pečeného baklažánu", 25, "", "")
# Wed R-D: rastlinné paprikové maslo
NOMILK[STR][1] = (R, "Rastlinné paprikové maslo", 15, "", "")
# Wed olovrant: bez syra
NOMILK[STR][-1] = (OL, "Zebra koláč (rastlinný)", 75, "1, 3", "")
# Thu R-D: rastlinné mlieko
NOMILK[STV][0] = (R, "Ryžový puding s rastlinným mliekom a jahodami", 115, "8", "")
# Thu obed: rastlinná smotana
NOMILK[STV][4] = (
    OA,
    "Morčacie na rastlinnej smotane s hráškom a paprikou",
    120,
    "",
    "",
)
# Thu olovrant: rastlinná nátierka
NOMILK[STV][-2] = (OL, "Rastlinná cviklová nátierka s chrenom", 25, "", "")
# Fri R-D: rastlinná nátierka
NOMILK[PIA][1] = (R, "Rastlinná nátierka z pečenej cibule", 25, "", "")
# Fri olovrant: rastlinný syr
NOMILK[PIA][-1] = (OL, "Špenátový koláč s rastlinným syrom", 75, "1, 3", "")

# ── NO MILK + NO GLUTEN ───────────────────────────────────────────────────────
NOMILK_NOGLUTEN = copy.deepcopy(NOMILK)
# Combine NoGluten changes on top of NoMilk
NOMILK_NOGLUTEN[PON][1] = (R, "Sladké bezlepkové krutóny", 10, "3", "")
NOMILK_NOGLUTEN[PON][-2] = (OL, "Bezlepkové pečivo", 50, "", "")
NOMILK_NOGLUTEN[UTO][4] = (OA, "Bezlepkové cestoviny", 110, "", "")
NOMILK_NOGLUTEN[UTO][-2] = (OL, "Bezlepkový chlieb", 50, "", "")
NOMILK_NOGLUTEN[STR][0] = (R, "Bezlepkový chlieb", 50, "", "")
NOMILK_NOGLUTEN[STR][4] = (
    P,
    "Zeleninový vývar s bezlepkovými cestovinami",
    200,
    "9",
    "ml",
)
NOMILK_NOGLUTEN[STR][-1] = (OL, "Bezlepkový zebra koláč (rastlinný)", 75, "3", "")
NOMILK_NOGLUTEN[STV][5] = (OA, "Jasmínová ryža", 80, "", "bez bulguru")
NOMILK_NOGLUTEN[STV][-3] = (OL, "Bezlepkové pečivo", 50, "", "")
NOMILK_NOGLUTEN[PIA][0] = (R, "Bezlepkový chlieb", 50, "", "")
NOMILK_NOGLUTEN[PIA][3] = (
    P,
    "Hydinový vývar s bezlepkovými cestovinami",
    200,
    "9",
    "ml",
)
NOMILK_NOGLUTEN[PIA][-1] = (
    OL,
    "Bezlepkový špenátový koláč s rastlinným syrom",
    75,
    "3",
    "",
)

# ── NO NO NO (no milk, no gluten, no egg, no soy) ────────────────────────────
NONONO = copy.deepcopy(NOMILK_NOGLUTEN)
# Mon R-D: kukuričné lupienky instead of krutóny
NONONO[PON][1] = (R, "Kukuričné lupienky", 10, "", "bez lepku, mlieka, vajec")
# Mon obed: pečená zelenina instead of vajíčko
NONONO[PON][6] = (OA, "Pečená zelenina", 25, "", "bez vajca")
# Wed R-D: rastlinný syr instead of vajce
NONONO[STR][2] = (R, "Rastlinný syr", 25, "", "bez vajca")
# Wed olovrant: bez alergénov
NONONO[STR][-1] = (OL, "Bezlepkový zebra koláč", 75, "", "bez mlieka, vajca")

# ── UČITEĽ (teacher) – no gram amounts, has Menu A/B/C variants ──────────────
UCITEL = {
    PON: [
        (R, "Čučoriedkový tvaroháčik", None, "7", ""),
        (R, "Sladké krutóny", None, "1, 3", ""),
        (R, "Zelené jablko", None, "", ""),
        (P, "Paradajková polievka s bazalkou a ryžou", None, "", "ml"),
        (OA, "Tekvicový prívarok", None, "7", "Menu A"),
        (OA, "Varené zemiaky s pažítkou", None, "", "Menu A"),
        (OA, "Vajíčko", None, "3", "Menu A"),
        (OB, "Granadírsky pochod", None, "1", "Menu B"),
        (OB, "Miešaný šalát", None, "", "Menu B"),
        (OC, "Kuracie s ázijskou zeleninou", None, "1, 6, 11", "Menu C"),
        (OC, "Jasmínová ryža", None, "", "Menu C"),
        (OL, "Grahamové pečivo", None, "1", ""),
        (OL, "Hydinová nátierka", None, "7, 10", ""),
    ],
    UTO: [
        (R, "Pšenová kaša so slivkami a makom", None, "7", ""),
        (R, "Pomaranč", None, "", ""),
        (P, "Karfiolová krémová s cuketou a pohánkou", None, "7", "ml"),
        (OA, "Boloňská omáčka s mletým hovädzím mäsom", None, "9", "Menu A"),
        (OA, "Špagety", None, "1", "Menu A"),
        (OA, "Syr", None, "7", "Menu A"),
        (OB, "Zapekané cestoviny s kuracím mäsom", None, "1, 7", "Menu B"),
        (OB, "Uhorkový šalát", None, "", "Menu B"),
        (OC, "Pečený encián", None, "7", "Menu C"),
        (OC, "Grilovaná zelenina", None, "", "Menu C"),
        (OC, "Pučené zemiaky", None, "", "Menu C"),
        (OC, "Tatárska omáčka", None, "3, 10", "Menu C"),
        (OL, "Kváskový chlieb", None, "1", ""),
        (OL, "Nátierka z pečeného baklažánu", None, "7", ""),
    ],
    STR: [
        (R, "Kváskový chlieb", None, "1", ""),
        (R, "Paprikové maslo", None, "7", ""),
        (R, "Vajce", None, "3", ""),
        (R, "Mrkva", None, "", ""),
        (P, "Zeleninový vývar s písmenkami", None, "1, 9", "ml"),
        (OA, "Pečené kuracie stehno", None, "", "Menu A, 1 ks"),
        (OA, "Zemiaková kaša", None, "", "Menu A"),
        (OA, "Broskyňový kompót", None, "", "Menu A"),
        (OB, "Morčacie na bratislavský spôsob", None, "1", "Menu B"),
        (OB, "Ryža", None, "", "Menu B"),
        (OB, "Broskyňový kompót", None, "", "Menu B"),
        (OC, "Hovädzie rebro", None, "", "Menu C"),
        (OC, "Petržlenové rizoto", None, "7", "Menu C"),
        (OC, "Mrkva", None, "", "Menu C"),
        (OC, "Demi glace", None, "", "Menu C"),
        (OL, "Zebra koláč", None, "1, 3, 7", ""),
    ],
    STV: [
        (R, "Ryžový puding s jahodami a pečenými mandľami", None, "7, 8", ""),
        (R, "Banán", None, "", ""),
        (P, "Vege detská kapustnica so zemiakmi", None, "9", "ml"),
        (OA, "Morčacie na smotane s hráškom a paprikou", None, "7", "Menu A"),
        (OA, "Ryža s bulgurom", None, "1", "Menu A"),
        (OB, "Hovädzie v paradajkovej omáčke", None, "", "Menu B"),
        (OB, "Varené zemiaky", None, "", "Menu B"),
        (OB, "Miešaný šalát", None, "", "Menu B"),
        (OC, "Bravčová panenka", None, "", "Menu C"),
        (OC, "Šošovicové ragú", None, "", "Menu C"),
        (OC, "Dusený kel", None, "", "Menu C"),
        (OC, "Tymianová omáčka", None, "", "Menu C"),
        (OL, "Pšeničné pečivo", None, "1", ""),
        (OL, "Cviklová nátierka s chrenom", None, "7", ""),
        (OL, "Zelenina", None, "", ""),
    ],
    PIA: [
        (R, "Kváskový chlieb", None, "1", ""),
        (R, "Nátierka z pečenej cibule", None, "7", ""),
        (R, "Paprika", None, "", ""),
        (P, "Hydinový vývar s niťovkami", None, "1, 9", "ml"),
        (OA, "Detské šošovicové ragú", None, "9", "Menu A"),
        (OA, "Jasmínová ryža", None, "", "Menu A"),
        (OA, "Zelený šalát s paradajkou a paprikou", None, "", "Menu A"),
        (OB, "Šúľance s makom a ovocným rozvarom", None, "1, 3, 7", "Menu B"),
        (OC, "Poké s lososom", None, "4", "Menu C"),
        (OC, "Hnedá ryža", None, "", "Menu C"),
        (OC, "Kukurička, uhorka", None, "", "Menu C"),
        (OC, "Miso omáčka so sezamom", None, "6, 11", "Menu C"),
        (OL, "Špenátový koláč so syrom", None, "1, 3, 7", ""),
    ],
}

# ── KLASIK MONTE (no breakfast grams, just "Čerstvé ovocie a zelenina") ──────
MONTE = copy.deepcopy(KLASIK)
for day in DAYS:
    # Replace breakfast with single entry
    obed_and_after = [(mt, *rest) for (mt, *rest) in MONTE[day] if mt != R]
    MONTE[day] = [
        (R, "Čerstvé ovocie a zelenina", None, "", "bez presných gramáží"),
    ] + obed_and_after


def ucitel_menu_c_rows():
    """Adult Menu C is only a menu variant; render it inside Klasik."""
    menu_c = {}
    for day, rows in UCITEL.items():
        menu_c[day] = []
        for meal_type, comp_name, amount, alergeny, note in rows:
            if meal_type == OC:
                menu_c[day].append(
                    (meal_type, comp_name, amount, alergeny, note or "Menu C")
                )
    return menu_c


def merge_day_rows(*datasets):
    merged = {}
    for day in DAYS:
        merged[day] = []
        for dataset in datasets:
            merged[day].extend(dataset.get(day, []))
    return merged


KLASIK_WITH_MENUS = merge_day_rows(KLASIK, KLASIK_MENU_B, ucitel_menu_c_rows())


def diet_has_difference(base_data, other_data):
    return any(base_data.get(day, []) != other_data.get(day, []) for day in DAYS)


# ── All diet types to render ──────────────────────────────────────────────────
DIETS = [
    ("Klasik", KLASIK_WITH_MENUS, C_MEAL_BG, C_MEAL_FG),
    ("Vege", VEGE, "E2EFDA", "375623"),
    ("No Gluten", NOGLUTEN, "FCE4D6", "843C0C"),
    ("No Milk", NOMILK, "FFF2CC", "7D6608"),
    ("No Milk – No Gluten", NOMILK_NOGLUTEN, "FDE9D9", "843C0C"),
    ("No No No", NONONO, "F4CCFF", "4A0072"),
]

if diet_has_difference(KLASIK, MONTE):
    DIETS.append(("Klasik Monte", MONTE, "D9E1F2", "203864"))

# ── Excel builder ─────────────────────────────────────────────────────────────
MEAL_TYPE_ORDER = [R, P, OA, OB, OC, OL]
MAIN_MEAL_TYPE_ORDER = [R, P, OA, OL]
MEAL_TYPE_LABEL = {
    R: "Raňajky-desiata",
    P: "Polievka",
    OA: "Obed (Menu A)",
    OB: "Obed (Menu B)",
    OC: "Obed (Menu C)",
    OL: "Olovrant",
}
MEAL_TYPE_COLOR = {
    R: ("FFF9C4", "5D4037"),
    P: ("E3F2FD", "0D47A1"),
    OA: ("E8F5E9", "1B5E20"),
    OB: ("FCE4D6", "7D2D00"),
    OC: ("F3E5F5", "4A148C"),
    OL: ("FFF3E0", "E65100"),
}

# Column indices
COL_DEN = 1
COL_CHOD = 2
COL_KOMPONENT = 3
COL_ALERGENY = 4
COL_MNOZSTVO_START = 5
COL_MNOZSTVO_END = COL_MNOZSTVO_START + len(PORTION_TYPES) - 1
COL_JEDNOTKA = COL_MNOZSTVO_END + 1
COL_POZNAMKY = COL_JEDNOTKA + 1
TOTAL_COLS = COL_POZNAMKY


def total_cols(include_notes=True):
    return COL_POZNAMKY if include_notes else COL_JEDNOTKA


def meal_type_order_for_sheet(diet_name):
    return MEAL_TYPE_ORDER if diet_name == "Klasik" else MAIN_MEAL_TYPE_ORDER


def write_instructions(wb, ctx, include_notes=True):
    ws = wb.create_sheet("INŠTRUKCIE", 0)
    ws.sheet_properties.tabColor = "1F4E79"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 50

    def title_row(row, text):
        ws.merge_cells(f"A{row}:C{row}")
        c = ws.cell(row=row, column=1, value=text)
        c.fill = fill(C_TITLE_BG)
        c.font = Font(bold=True, color=C_TITLE_FG, size=13)
        c.alignment = center()
        c.border = border_thin()
        ws.row_dimensions[row].height = 28

    def section_row(row, text):
        ws.merge_cells(f"A{row}:C{row}")
        c = ws.cell(row=row, column=1, value=text)
        c.fill = fill(C_DAY_BG)
        c.font = Font(bold=True, color=C_DAY_FG, size=11)
        c.alignment = left()
        c.border = border_thin()
        ws.row_dimensions[row].height = 22

    def data_row(row, label, val_b, val_c=None, bold_b=False, bg_b=None):
        ws.row_dimensions[row].height = 20
        ca = ws.cell(row=row, column=1, value=label)
        ca.font = Font(bold=True, color="000000", size=10)
        ca.alignment = left()
        ca.border = border_thin()
        cb = ws.cell(row=row, column=2, value=val_b)
        cb.font = Font(bold=bold_b, color="000000", size=10)
        cb.alignment = center()
        cb.border = border_thin()
        if bg_b:
            cb.fill = fill(bg_b)
        cc = ws.cell(row=row, column=3, value=val_c or "")
        cc.font = Font(color="555555", size=9, italic=True)
        cc.alignment = left()
        cc.border = border_thin()

    title_row(1, "JEDÁLNIČEK – ŠABLÓNA NA VSTUP MNOŽSTIEV")

    # ── Week metadata in separate cells for parsing ───────────────────────────
    section_row(2, "Identifikácia týždňa")
    data_row(3, "Rok", ctx["year"], "")
    data_row(4, "Týždeň", ctx["week"], "")
    data_row(5, "Dátum od", ctx["start"].strftime("%d.%m.%Y"), "")
    data_row(6, "Dátum do", ctx["end"].strftime("%d.%m.%Y"), "")

    # ── Instructions ──────────────────────────────────────────────────────────
    section_row(8, "Ako používať túto šablónu")
    data_row(9, "Záložka", "Každá záložka = jeden typ jedálnička (Klasik, Vege, No Gluten ...)")
    data_row(10, "Deň", "Každý deň je farebne oddelená sekcia.")
    data_row(11, "Chod", "Raňajky-desiata / Polievka / Obed (Menu A, B, C) / Olovrant")
    data_row(12, "Komponent", "Každá zložka jedla = jeden riadok.")
    data_row(13, "Alergény", "Číslice alergénov podľa EÚ klasifikácie.")

    section_row(15, "Vypĺňanie množstiev")
    data_row(
        16,
        "Množstvá",
        "Ručný vstup – admin vypĺňa samostatne pre každú veľkosť porcie.",
        None,
        bold_b=True,
        bg_b=C_INPUT_BG,
    )
    data_row(17, "Veľkosti porcií", ", ".join(PORTION_TYPES), None, bg_b=C_INPUT_BG)
    data_row(
        18,
        "Jednotka",
        "g / ks – predvyplnené, manuálne prepísateľné.",
        None,
        bg_b=C_UNIT_BG,
    )
    if include_notes:
        data_row(
            19,
            "Poznámky",
            "Doplňujúce info (Menu A/B/C, špeciálne požiadavky ...).",
            None,
            bg_b="EDEDED",
        )


def write_table_header(ws, row, headers):
    header_bgs = {
        "Deň": C_DAY_BG,
        "Chod": C_DAY_BG,
        "Komponent / Jedlo": C_DAY_BG,
        "Alergény": C_DAY_BG,
        "Jednotka": C_UNIT_BG,
        "Poznámky": C_NOTE_BG,
    }
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        is_amount_header = h.startswith("Množstvo ")
        cell.fill = fill(
            C_INPUT_BG if is_amount_header else header_bgs.get(h, C_DAY_BG)
        )
        cell.font = Font(
            bold=True,
            color=(
                "000000"
                if is_amount_header or h in ("Jednotka", "Poznámky")
                else C_DAY_FG
            ),
            size=10,
        )
        cell.alignment = center()
        cell.border = border_thin()
    ws.row_dimensions[row].height = 30


def write_diet_rows(ws, start_row, data, include_notes=True):
    current_row = start_row
    alt = False
    sheet_total_cols = total_cols(include_notes)

    for day in DAYS:
        day_data = data.get(day, [])
        if not day_data:
            continue

        ws.merge_cells(
            f"A{current_row}:{get_column_letter(sheet_total_cols)}{current_row}"
        )
        c = ws.cell(row=current_row, column=1, value=f"  {day}")
        c.fill = fill(C_DAY_BG)
        c.font = Font(bold=True, color=C_DAY_FG, size=11)
        c.alignment = left()
        c.border = border_thin()
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        grouped = {}
        for entry in day_data:
            meal_type = entry[0]
            grouped.setdefault(meal_type, []).append(entry)

        for mt in MEAL_TYPE_ORDER:
            if mt not in grouped:
                continue
            mt_bg, mt_fg = MEAL_TYPE_COLOR[mt]
            mt_label = MEAL_TYPE_LABEL[mt]
            entries = grouped[mt]

            ws.merge_cells(
                f"B{current_row}:{get_column_letter(sheet_total_cols)}{current_row}"
            )
            c = ws.cell(row=current_row, column=COL_DEN)
            c.fill = fill(mt_bg)
            c.border = border_thin()
            c = ws.cell(row=current_row, column=COL_CHOD, value=mt_label)
            c.fill = fill(mt_bg)
            c.font = Font(bold=True, color=mt_fg, size=9)
            c.alignment = left()
            c.border = border_thin()
            ws.row_dimensions[current_row].height = 18
            current_row += 1

            for comp_row in entries:
                _, comp_name, _base_amount, alergeny, note = comp_row
                row_bg = C_ALT_BG if alt else "FFFFFF"
                alt = not alt

                # Deň + Chod columns (empty, styled)
                for col in (COL_DEN, COL_CHOD):
                    c = ws.cell(row=current_row, column=col)
                    c.fill = fill(row_bg)
                    c.border = border_thin()

                # Komponent
                c = ws.cell(row=current_row, column=COL_KOMPONENT, value=comp_name)
                c.fill = fill(row_bg)
                c.font = font(size=10)
                c.alignment = left()
                c.border = border_thin()

                # Alergény
                c = ws.cell(row=current_row, column=COL_ALERGENY, value=alergeny)
                c.fill = fill(row_bg)
                c.font = font(size=9, color="666666")
                c.alignment = center()
                c.border = border_thin()

                # Množstvá – empty, manual input per portion size
                for col in range(COL_MNOZSTVO_START, COL_MNOZSTVO_END + 1):
                    c = ws.cell(row=current_row, column=col, value=None)
                    c.fill = fill(C_INPUT_BG)
                    c.alignment = center()
                    c.border = border_thin()
                    c.number_format = "0.##"

                # Jednotka – derived, manually overridable
                unit = derive_unit(note)
                c = ws.cell(row=current_row, column=COL_JEDNOTKA, value=unit)
                c.fill = fill(C_UNIT_BG)
                c.font = font(size=10, color="1F4E79")
                c.alignment = center()
                c.border = border_thin()

                if include_notes:
                    # Poznámky
                    c = ws.cell(row=current_row, column=COL_POZNAMKY, value=note or "")
                    c.fill = fill(C_NOTE_BG)
                    c.font = font(
                        size=9,
                        color=C_RED_NOTE if note else "555555",
                        italic=True,
                    )
                    c.alignment = left()
                    c.border = border_thin()

                ws.row_dimensions[current_row].height = 18
                current_row += 1

        ws.row_dimensions[current_row].height = 8
        for ci in range(1, sheet_total_cols + 1):
            ws.cell(row=current_row, column=ci).fill = fill("DDDDDD")
        current_row += 1

    return current_row


def write_blank_diet_skeleton(
    ws,
    start_row,
    meal_type_order,
    include_notes=False,
):
    current_row = start_row
    sheet_total_cols = total_cols(include_notes)

    for day in DAYS:
        ws.merge_cells(
            f"A{current_row}:{get_column_letter(sheet_total_cols)}{current_row}"
        )
        c = ws.cell(row=current_row, column=1, value=f"  {day}")
        c.fill = fill(C_DAY_BG)
        c.font = Font(bold=True, color=C_DAY_FG, size=11)
        c.alignment = left()
        c.border = border_thin()
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        for mt in meal_type_order:
            mt_bg, mt_fg = MEAL_TYPE_COLOR[mt]
            mt_label = MEAL_TYPE_LABEL[mt]

            ws.merge_cells(
                f"B{current_row}:{get_column_letter(sheet_total_cols)}{current_row}"
            )
            c = ws.cell(row=current_row, column=COL_DEN)
            c.fill = fill(mt_bg)
            c.border = border_thin()
            c = ws.cell(row=current_row, column=COL_CHOD, value=mt_label)
            c.fill = fill(mt_bg)
            c.font = Font(bold=True, color=mt_fg, size=9)
            c.alignment = left()
            c.border = border_thin()
            ws.row_dimensions[current_row].height = 18
            current_row += 1

            for idx in range(BLANK_COMPONENT_ROWS_PER_MEAL):
                row_bg = C_ALT_BG if idx % 2 else "FFFFFF"
                for col in (COL_DEN, COL_CHOD):
                    c = ws.cell(row=current_row, column=col)
                    c.fill = fill(row_bg)
                    c.border = border_thin()

                c = ws.cell(row=current_row, column=COL_KOMPONENT)
                c.fill = fill(row_bg)
                c.alignment = left()
                c.border = border_thin()

                c = ws.cell(row=current_row, column=COL_ALERGENY)
                c.fill = fill(row_bg)
                c.alignment = center()
                c.border = border_thin()

                for col in range(COL_MNOZSTVO_START, COL_MNOZSTVO_END + 1):
                    c = ws.cell(row=current_row, column=col)
                    c.fill = fill(C_INPUT_BG)
                    c.alignment = center()
                    c.border = border_thin()
                    c.number_format = "0.##"

                c = ws.cell(row=current_row, column=COL_JEDNOTKA, value="g")
                c.fill = fill(C_UNIT_BG)
                c.font = font(size=10, color="1F4E79")
                c.alignment = center()
                c.border = border_thin()

                if include_notes:
                    c = ws.cell(row=current_row, column=COL_POZNAMKY)
                    c.fill = fill(C_NOTE_BG)
                    c.alignment = left()
                    c.border = border_thin()

                ws.row_dimensions[current_row].height = 18
                current_row += 1

        ws.row_dimensions[current_row].height = 8
        for ci in range(1, sheet_total_cols + 1):
            ws.cell(row=current_row, column=ci).fill = fill("DDDDDD")
        current_row += 1

    return current_row


def build_diet_sheet(
    wb,
    diet_name,
    data,
    header_bg,
    header_fg,
    include_rows=True,
    include_blank_skeleton=False,
    include_notes=True,
):
    sheet_name = diet_name[:31]
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = header_bg.replace("#", "")
    sheet_total_cols = total_cols(include_notes)

    ws.freeze_panes = "A3"

    ws.column_dimensions["A"].width = 12   # Deň
    ws.column_dimensions["B"].width = 20   # Chod
    ws.column_dimensions["C"].width = 42   # Komponent
    ws.column_dimensions["D"].width = 12   # Alergény
    for col in range(COL_MNOZSTVO_START, COL_MNOZSTVO_END + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.column_dimensions[get_column_letter(COL_JEDNOTKA)].width = 9
    if include_notes:
        ws.column_dimensions[get_column_letter(COL_POZNAMKY)].width = 28

    # Row 1: diet name only (no week info — week is only on INŠTRUKCIE)
    ws.merge_cells(f"A1:{get_column_letter(sheet_total_cols)}1")
    c = ws["A1"]
    c.value = f"JEDÁLNIČEK – {diet_name.upper()}"
    c.fill = fill(C_TITLE_BG)
    c.font = Font(bold=True, color=C_TITLE_FG, size=12)
    c.alignment = center()
    c.border = border_thin()
    ws.row_dimensions[1].height = 28

    headers = (
        ["Deň", "Chod", "Komponent / Jedlo", "Alergény"]
        + [f"Množstvo {portion}" for portion in PORTION_TYPES]
        + ["Jednotka"]
    )
    if include_notes:
        headers.append("Poznámky")
    write_table_header(ws, 2, headers)
    if include_blank_skeleton:
        write_blank_diet_skeleton(
            ws,
            3,
            meal_type_order_for_sheet(diet_name),
            include_notes=include_notes,
        )
    elif include_rows:
        write_diet_rows(ws, 3, data, include_notes=include_notes)


def main():
    args = parse_args()
    ctx = week_context(args.year, args.week)
    filename_prefix = "jedalnicky_template" if args.blank else "jedalnicky_template_ukazka"
    out_path = args.output or os.path.join(
        os.path.dirname(__file__),
        "..",
        "test",
        "jedalnicky",
        f"{filename_prefix}_W{args.week:02d}_{args.year}.xlsx",
    )
    out_path = os.path.normpath(out_path)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_instructions(wb, ctx, include_notes=not args.blank)

    for diet_name, data, h_bg, h_fg in DIETS:
        build_diet_sheet(
            wb,
            diet_name,
            data,
            h_bg,
            h_fg,
            include_rows=not args.blank,
            include_blank_skeleton=args.blank,
            include_notes=not args.blank,
        )

    wb.save(out_path)
    print(f"✓ Saved: {out_path}")


if __name__ == "__main__":
    main()
