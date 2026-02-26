import datetime
import io
import logging
import os

from django.contrib.auth.models import User
from django.http import FileResponse, HttpResponse
from django.utils import timezone
from rest_framework import permissions
from rest_framework import serializers as drf_serializers
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import AuthenticationFailed
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.views import TokenObtainPairView

from .models import DailyOrder, Diet
from .serializers import DailyOrderSerializer
from .serializers_user import AdminUserSerializer, DietSerializer, UserProfileSerializer
from .services import _is_order_empty, _last_non_empty_order

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# PDF Unicode font helpers
# ---------------------------------------------------------------------------
# reportlab's built-in fonts (Helvetica, Times…) are Type1 and limited to
# Latin-1, which drops many Slovak characters (č, š, ľ, ň, ť, ŕ, ĺ, ď…).
# We register DejaVu Sans (full Unicode/UTF-8 TrueType) and fall back to
# Helvetica only if the font file cannot be found (e.g. a minimal test env).

_DEJAVU_CANDIDATES = [
    # Debian / Ubuntu / Docker with fonts-dejavu-core
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    # macOS (via Homebrew font-dejavu)
    "/Library/Fonts/DejaVuSans.ttf",
    "/Library/Fonts/DejaVuSans-Bold.ttf",
]

_PDF_FONT_REGULAR = "Helvetica"  # fallback
_PDF_FONT_BOLD = "Helvetica-Bold"  # fallback
_pdf_fonts_registered = False


def _register_pdf_fonts():
    """Register DejaVu Sans TTFont once per process for full Unicode support."""
    global _PDF_FONT_REGULAR, _PDF_FONT_BOLD, _pdf_fonts_registered
    if _pdf_fonts_registered:
        return
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        regular = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
        bold = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"

        # Walk candidate pairs until we find a matching set
        candidates = [
            (
                "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
            ),
            ("/Library/Fonts/DejaVuSans.ttf", "/Library/Fonts/DejaVuSans-Bold.ttf"),
        ]
        for reg_path, bold_path in candidates:
            if os.path.isfile(reg_path) and os.path.isfile(bold_path):
                regular, bold = reg_path, bold_path
                break
        else:
            # No TTFont found — stay with Helvetica fallback
            _pdf_fonts_registered = True
            return

        pdfmetrics.registerFont(TTFont("DejaVuSans", regular))
        pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", bold))
        pdfmetrics.registerFontFamily(
            "DejaVuSans",
            normal="DejaVuSans",
            bold="DejaVuSans-Bold",
        )
        _PDF_FONT_REGULAR = "DejaVuSans"
        _PDF_FONT_BOLD = "DejaVuSans-Bold"
        logger.debug("Registered DejaVuSans TTFont for PDF generation.")
    except Exception:
        logger.warning(
            "Could not register DejaVuSans TTFont; falling back to Helvetica.",
            exc_info=True,
        )
    finally:
        _pdf_fonts_registered = True


class EmailTokenObtainPairSerializer(TokenObtainPairSerializer):
    """JWT token serializer that authenticates via email instead of username."""

    username_field = "email"

    def validate(self, attrs):
        email = attrs.get("email", "").strip().lower()
        password = attrs.get("password", "")

        if not email:
            raise AuthenticationFailed("Nesprávny email alebo heslo.")

        try:
            user = User.objects.get(email__iexact=email)
        except User.DoesNotExist:
            raise AuthenticationFailed("Nesprávny email alebo heslo.")
        except User.MultipleObjectsReturned:
            raise AuthenticationFailed("Nesprávny email alebo heslo.")

        if not user.check_password(password):
            raise AuthenticationFailed("Nesprávny email alebo heslo.")

        if not user.is_active:
            raise AuthenticationFailed("Tento účet je neaktívny.")

        refresh = RefreshToken.for_user(user)
        return {
            "refresh": str(refresh),
            "access": str(refresh.access_token),
        }


class EmailTokenObtainPairView(TokenObtainPairView):
    """Token endpoint that accepts email + password."""

    serializer_class = EmailTokenObtainPairSerializer


class DailyOrderViewSet(viewsets.ModelViewSet):
    serializer_class = DailyOrderSerializer
    # Authenticated users only
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = DailyOrder.objects.all()
        user = self.request.user

        if user.is_staff:
            user_id = self.request.query_params.get("user_id")
            if user_id:
                queryset = queryset.filter(user_id=user_id)
            else:
                # If no user_id is provided, return only the staff user's own orders
                # to prevent returning ALL orders by default (which breaks by_date logic).
                queryset = queryset.filter(user=user)
        else:
            queryset = queryset.filter(user=user)

        return queryset

    def perform_create(self, serializer):
        if self.request.user.is_staff:
            from rest_framework.exceptions import PermissionDenied

            raise PermissionDenied("Administrators cannot place orders.")
        # The serializer.save() will call create() which enables update_or_create logic
        serializer.save(user=self.request.user)

    @action(detail=False, methods=["get"], url_path="by-date/(?P<date>[^/.]+)")
    def by_date(self, request, date=None):
        try:
            order = self.get_queryset().get(date=date)
            serializer = self.get_serializer(order)
            return Response(serializer.data)
        except DailyOrder.DoesNotExist:
            return Response(
                {"data": {}}, status=status.HTTP_200_OK
            )  # Return empty struct if not found


class UserProfileViewSet(viewsets.ViewSet):
    """
    ViewSet for user profile management.
    """

    permission_classes = [permissions.IsAuthenticated]

    @action(detail=False, methods=["get", "patch"], url_path="profile")
    def profile(self, request):
        """Get or update current user's profile"""
        if request.method == "GET":
            serializer = UserProfileSerializer(request.user)
            return Response(serializer.data)

        elif request.method == "PATCH":
            serializer = UserProfileSerializer(
                request.user, data=request.data, partial=True
            )
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class DietViewSet(viewsets.ModelViewSet):
    queryset = Diet.objects.all()
    serializer_class = DietSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_permissions(self):
        if self.action in ["create", "update", "partial_update", "destroy"]:
            return [permissions.IsAdminUser()]
        return super().get_permissions()


class AdminUserViewSet(viewsets.ModelViewSet):
    """
    Admin ViewSet for managing users and their settings.
    """

    queryset = User.objects.all().order_by("email")
    serializer_class = AdminUserSerializer
    permission_classes = [permissions.IsAdminUser]


class AdminSummaryViewSet(viewsets.ViewSet):
    """
    Admin ViewSet for Dashboard Summaries.
    """

    permission_classes = [permissions.IsAdminUser]

    @action(detail=False, methods=["get"], url_path="daily-stats")
    def daily_stats(self, request):
        date_str = request.query_params.get("date")
        if not date_str:
            return Response(
                {"error": "Date parameter is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        orders = DailyOrder.objects.filter(date=date_str)

        # Structure:
        # meals: {
        #   "breakfast": { "CategoryName": { "menus": {"A": 10}, "diets": {"Gluten": 1}, "total": 10 } }
        # }
        stats = {
            "total_orders": 0,
            "status_breakdown": {"submitted": 0},
            "meals": {"breakfast": {}, "lunch": {}, "olovrant": {}},
        }

        for order in orders:
            stats["total_orders"] += 1
            # Status is now always submitted or ignored
            stats["status_breakdown"]["submitted"] += 1

            data = order.data or {}

            # Helper to aggregate a specific meal type (breakfast/lunch/olovrant)
            self._aggregate_meal(stats, data, "breakfast")
            self._aggregate_meal(stats, data, "lunch")
            self._aggregate_meal(stats, data, "olovrant")

        return Response(stats)

    def _aggregate_meal(self, stats, data, meal_key):
        if meal_key not in data or not data[meal_key]:
            return

        meal_data = data[meal_key]
        # iterate over categories in this meal
        # e.g. meal_data = { "Dospelý (SŠ)": { "menuCounts": {"A":1}, "diets": {} } }

        for category, details in meal_data.items():
            if not details:
                continue

            # Ensure category dict exists in stats
            if category not in stats["meals"][meal_key]:
                stats["meals"][meal_key][category] = {
                    "menus": {},
                    "diets": {},
                    "total": 0,
                }

            cat_stats = stats["meals"][meal_key][category]

            # Aggregate Menus
            menu_counts = details.get("menuCounts", {})
            for menu_type, count in menu_counts.items():
                count = int(count or 0)
                if count > 0:
                    cat_stats["menus"][menu_type] = (
                        cat_stats["menus"].get(menu_type, 0) + count
                    )
                    cat_stats["total"] += count

            # Aggregate Diets
            diets = details.get("diets", {})
            for diet_name, count in diets.items():
                count = int(count or 0)
                if count > 0:
                    cat_stats["diets"][diet_name] = (
                        cat_stats["diets"].get(diet_name, 0) + count
                    )

    # ------------------------------------------------------------------
    # Helpers shared by daily_report and daily_report_xlsx
    # ------------------------------------------------------------------

    @staticmethod
    def _safe_int(v) -> int:
        """Coerce a stored count value to int, returning 0 on any error."""
        try:
            return int(v or 0)
        except (TypeError, ValueError):
            return 0

    @classmethod
    def _build_user_meal_row(cls, order_data: dict, meal_key: str) -> dict:
        """Return {categories: [{name, menus, diets, total}], total: int}.

        Supports both data shapes:
        - Nested-by-category: {"Cat": {"menuCounts": {...}, "diets": {...}}, ...}
        - Flat (legacy): {"menuCounts": {...}, "diets": {...}}
        """
        meal = order_data.get(meal_key) or {}
        if not isinstance(meal, dict):
            return {"categories": [], "total": 0}
        categories = []
        meal_total = 0
        # Flat shape: meal itself contains menuCounts/diets keys
        iter_categories = (
            [(meal_key, meal)]
            if "menuCounts" in meal
            else [(k, v) for k, v in meal.items() if isinstance(v, dict)]
        )
        for cat_name, details in iter_categories:
            if not isinstance(details, dict):
                continue
            menu_counts = {
                k: cls._safe_int(v)
                for k, v in (details.get("menuCounts") or {}).items()
            }
            diets = {
                k: cls._safe_int(v)
                for k, v in (details.get("diets") or {}).items()
                if cls._safe_int(v) > 0
            }
            cat_total = sum(menu_counts.values())
            meal_total += cat_total
            categories.append(
                {
                    "name": cat_name,
                    "menus": menu_counts,
                    "diets": diets,
                    "total": cat_total,
                }
            )
        return {"categories": categories, "total": meal_total}

    @staticmethod
    def _merge_meal_totals(totals: dict, meal_row: dict) -> None:
        """Accumulate meal_row counts into totals dict (in-place)."""
        totals["total"] = totals.get("total", 0) + meal_row["total"]
        for cat in meal_row["categories"]:
            for menu, cnt in cat["menus"].items():
                totals.setdefault("menus", {})[menu] = (
                    totals["menus"].get(menu, 0) + cnt
                )
            for diet, cnt in cat["diets"].items():
                totals.setdefault("diets", {})[diet] = (
                    totals["diets"].get(diet, 0) + cnt
                )

    @action(detail=False, methods=["get"], url_path="daily-report")
    def daily_report(self, request):
        """
        Per-user order report for a given date.
        GET /api/admin/summary/daily-report/?date=YYYY-MM-DD
        """
        date_str = request.query_params.get("date")
        if not date_str:
            return Response(
                {"error": "date parameter required"}, status=status.HTTP_400_BAD_REQUEST
            )
        try:
            target_date = datetime.date.fromisoformat(date_str)
        except (TypeError, ValueError):
            return Response(
                {"error": "invalid date format, expected YYYY-MM-DD"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        orders = (
            DailyOrder.objects.filter(date=target_date)
            .select_related("user", "user__settings")
            .order_by("user__email")
        )

        totals = {
            "breakfast": {"menus": {}, "diets": {}, "total": 0},
            "lunch": {"menus": {}, "diets": {}, "total": 0},
            "olovrant": {"menus": {}, "diets": {}, "total": 0},
            "grand": 0,
        }
        rows = []
        for order in orders:
            user = order.user
            data = order.data or {}
            bf = self._build_user_meal_row(data, "breakfast")
            lu = self._build_user_meal_row(data, "lunch")
            ol = self._build_user_meal_row(data, "olovrant")
            row_total = bf["total"] + lu["total"] + ol["total"]
            _settings = getattr(user, "settings", None)
            visible_meals = getattr(_settings, "visible_meals", None) or [
                "breakfast",
                "lunch",
                "olovrant",
            ]
            rows.append(
                {
                    "user_id": user.id,
                    "name": f"{user.first_name} {user.last_name}".strip() or user.email,
                    "email": user.email,
                    "breakfast": bf,
                    "lunch": lu,
                    "olovrant": ol,
                    "visible_meals": visible_meals,
                    "total": row_total,
                }
            )
            self._merge_meal_totals(totals["breakfast"], bf)
            self._merge_meal_totals(totals["lunch"], lu)
            self._merge_meal_totals(totals["olovrant"], ol)
            totals["grand"] += row_total

        return Response(
            {"date": target_date.isoformat(), "rows": rows, "totals": totals}
        )

    # ------------------------------------------------------------------
    # XLSX helpers — extracted to keep daily_report_xlsx readable
    # ------------------------------------------------------------------

    @classmethod
    def _xlsx_collect_columns(cls, rows_data, meal_keys):
        """First-pass scan: gather every (category, menu, diet) combination per meal.

        Returns:
            {meal_key: {cat_name: {'menus': [sorted], 'diets': [sorted]}}}

        Handles both data shapes:
        - Nested-by-category: {"Cat": {"menuCounts": {...}, "diets": {...}}, ...}
        - Flat (legacy):       {"menuCounts": {...}, "diets": {...}}
        """
        # Canonical category order mirrors the client-side CATEGORIES constant.
        CAT_ORDER = ["Jasle", "Škôlka", "ZŠ 1.stupeň", "ZŠ 2.stupeň", "Dospelý (SŠ)"]

        raw = {m: {} for m in meal_keys}  # meal -> cat_name -> {menus: set, diets: set}

        def _scan(cat_name, details, mk):
            if not isinstance(details, dict):
                return
            for key, cnt in (details.get("menuCounts") or {}).items():
                if cls._safe_int(cnt) > 0:
                    raw[mk].setdefault(cat_name, {"menus": set(), "diets": set()})
                    raw[mk][cat_name]["menus"].add(key)
            for key, cnt in (details.get("diets") or {}).items():
                if cls._safe_int(cnt) > 0:
                    raw[mk].setdefault(cat_name, {"menus": set(), "diets": set()})
                    raw[mk][cat_name]["diets"].add(key)

        for row in rows_data:
            data = row["data"]
            for mk in meal_keys:
                meal = data.get(mk) or {}
                if "menuCounts" in meal:
                    # Flat shape — use meal_key as the category name
                    _scan(mk, meal, mk)
                else:
                    # Nested-by-category shape
                    for cat_name, details in meal.items():
                        _scan(cat_name, details, mk)

        # Sort categories by canonical order; menus/diets alphabetically.
        sorted_cats = {}
        for mk in meal_keys:
            sorted_cat_keys = sorted(
                raw[mk].keys(),
                key=lambda c: (CAT_ORDER.index(c) if c in CAT_ORDER else 99, c),
            )
            sorted_cats[mk] = {
                cat: {
                    "menus": sorted(raw[mk][cat]["menus"]),
                    "diets": sorted(raw[mk][cat]["diets"]),
                }
                for cat in sorted_cat_keys
            }
        return sorted_cats

    @staticmethod
    def _xlsx_build_column_meta(sorted_cats, meal_keys, meal_labels):
        """Build 3 header rows and column metadata list.

        Row 1 (meal level):     Klient | Raňajky...           | Obed...   | Celkovo
        Row 2 (category level):        | Jasle... | ZŠ... | Spolu | Jasle... | Spolu |
        Row 3 (menu/diet):             | Menu A | Diet | Menu A | B | Diet |   |   |
        """
        col_meta = [("fixed", None, "name", None)]
        header_row_1 = ["Klient"]
        header_row_2 = [""]
        header_row_3 = [""]

        for mk in meal_keys:
            cats = sorted_cats[mk]
            # meal span = sum of (menus + diets) per category + 1 Spolu column
            inner_cols = sum(len(v["menus"]) + len(v["diets"]) for v in cats.values())
            meal_span = inner_cols + 1
            header_row_1 += [meal_labels[mk]] + [""] * (meal_span - 1)

            for cat_name, cat_data in cats.items():
                cat_span = len(cat_data["menus"]) + len(cat_data["diets"])
                header_row_2 += [cat_name] + [""] * max(cat_span - 1, 0)
                for menu_key in cat_data["menus"]:
                    header_row_3.append(f"Menu {menu_key}")
                    col_meta.append((mk, cat_name, "menu", menu_key))
                for diet_name in cat_data["diets"]:
                    header_row_3.append(diet_name)
                    col_meta.append((mk, cat_name, "diet", diet_name))

            # Spolu column for this meal
            header_row_2.append("Spolu")
            header_row_3.append("")
            col_meta.append((mk, None, "total", None))

        header_row_1.append("Celkovo")
        header_row_2.append("")
        header_row_3.append("")
        col_meta.append(("fixed", None, "grand_total", None))
        return col_meta, header_row_1, header_row_2, header_row_3

    @staticmethod
    def _xlsx_style_headers(
        ws,
        col_meta,
        sorted_cats,
        meal_keys,
        header_fill_main,
        header_fill_meal,
        header_font,
        center,
    ):
        """Apply fills, fonts, merges to the three header rows (rows 3, 4, 5)."""
        from openpyxl.styles import Alignment  # already imported at call site

        total_cols = len(col_meta)

        # ── Row 3: meal-level spans ────────────────────────────────────────────
        current_col = 2  # col 1=Klient, col 2+ = meals
        for mk in meal_keys:
            cats = sorted_cats[mk]
            inner_cols = sum(len(v["menus"]) + len(v["diets"]) for v in cats.values())
            meal_span = inner_cols + 1
            if meal_span > 1:
                ws.merge_cells(
                    start_row=3,
                    start_column=current_col,
                    end_row=3,
                    end_column=current_col + meal_span - 1,
                )
            cell = ws.cell(row=3, column=current_col)
            cell.fill = header_fill_meal[mk]
            cell.font = header_font
            cell.alignment = center
            current_col += meal_span

        for c in [1, total_cols]:
            cell = ws.cell(row=3, column=c)
            cell.fill = header_fill_main
            cell.font = header_font
            cell.alignment = center

        # ── Row 4: category-level spans + Spolu per meal ───────────────────────
        row45_fill = {}  # col_index -> fill (shared for rows 4 and 5)
        current_col = 2
        for mk in meal_keys:
            for cat_name, cat_data in sorted_cats[mk].items():
                cat_span = len(cat_data["menus"]) + len(cat_data["diets"])
                if cat_span > 0:
                    if cat_span > 1:
                        ws.merge_cells(
                            start_row=4,
                            start_column=current_col,
                            end_row=4,
                            end_column=current_col + cat_span - 1,
                        )
                    for c in range(current_col, current_col + cat_span):
                        row45_fill[c] = header_fill_meal[mk]
                    cell = ws.cell(row=4, column=current_col)
                    cell.fill = header_fill_meal[mk]
                    cell.font = header_font
                    cell.alignment = center
                    current_col += cat_span
            # Spolu column for this meal
            row45_fill[current_col] = header_fill_meal[mk]
            cell = ws.cell(row=4, column=current_col)
            cell.fill = header_fill_meal[mk]
            cell.font = header_font
            cell.alignment = center
            current_col += 1

        for c in [1, total_cols]:
            cell = ws.cell(row=4, column=c)
            cell.fill = header_fill_main
            cell.font = header_font
            cell.alignment = center

        # ── Row 5: menu/diet labels ────────────────────────────────────────────
        for c_idx in range(1, total_cols + 1):
            cell = ws.cell(row=5, column=c_idx)
            cell.fill = row45_fill.get(c_idx, header_fill_main)
            cell.font = header_font
            cell.alignment = center

        # ── Merge Klient and Celkovo across all 3 header rows (3-5) ────────
        ws.merge_cells(start_row=3, start_column=1, end_row=5, end_column=1)
        ws.merge_cells(
            start_row=3, start_column=total_cols, end_row=5, end_column=total_cols
        )

    @classmethod
    def _xlsx_write_data(
        cls,
        ws,
        rows_data,
        meal_keys,
        sorted_cats,
        bold_font,
    ):
        """Append per-user data rows and a SPOLU totals row."""
        # Running totals per meal / category / menu+diet
        totals = {
            mk: {
                cat: {
                    "menus": {m: 0 for m in cat_data["menus"]},
                    "diets": {d: 0 for d in cat_data["diets"]},
                }
                for cat, cat_data in cats.items()
            }
            for mk, cats in sorted_cats.items()
        }
        meal_totals = {mk: 0 for mk in meal_keys}
        grand_total = 0

        for row_info in rows_data:
            user = row_info["user"]
            data = row_info["data"]
            visible_meals = row_info.get("visible_meals") or meal_keys
            display_name = f"{user.first_name} {user.last_name}".strip() or user.email
            row_vals = [display_name]
            row_grand = 0
            for mk in meal_keys:
                # Pre-compute column count for this meal (to fill blanks when disallowed)
                meal_col_count = (
                    sum(
                        len(cat_data["menus"]) + len(cat_data["diets"])
                        for cat_data in sorted_cats[mk].values()
                    )
                    + 1  # +1 for Spolu
                )
                if mk not in visible_meals:
                    row_vals.extend([""] * meal_col_count)
                    continue
                meal = data.get(mk) or {}
                meal_total = 0
                is_flat = "menuCounts" in meal
                for cat_name, cat_data in sorted_cats[mk].items():
                    if is_flat:
                        cat_details = meal if cat_name == mk else {}
                    else:
                        cat_details = meal.get(cat_name) or {}
                    for menu_key in cat_data["menus"]:
                        cnt = cls._safe_int(
                            (cat_details.get("menuCounts") or {}).get(menu_key, 0)
                        )
                        row_vals.append(cnt or "")
                        totals[mk][cat_name]["menus"][menu_key] += cnt
                        meal_total += cnt
                    for diet_name in cat_data["diets"]:
                        cnt = cls._safe_int(
                            (cat_details.get("diets") or {}).get(diet_name, 0)
                        )
                        row_vals.append(cnt or "")
                        totals[mk][cat_name]["diets"][diet_name] += cnt
                row_vals.append(meal_total or "")
                meal_totals[mk] += meal_total
                row_grand += meal_total
            row_vals.append(row_grand or "")
            grand_total += row_grand
            ws.append(row_vals)

        # SPOLU totals row
        total_row = ["SPOLU"]
        for mk in meal_keys:
            for cat_name, cat_data in sorted_cats[mk].items():
                for menu_key in cat_data["menus"]:
                    total_row.append(
                        totals[mk][cat_name]["menus"].get(menu_key, 0) or ""
                    )
                for diet_name in cat_data["diets"]:
                    total_row.append(
                        totals[mk][cat_name]["diets"].get(diet_name, 0) or ""
                    )
            total_row.append(meal_totals[mk] or "")
        total_row.append(grand_total or "")
        ws.append(total_row)

        totals_row_num = ws.max_row
        for c_idx in range(1, ws.max_column + 1):
            ws.cell(row=totals_row_num, column=c_idx).font = bold_font

    @action(detail=False, methods=["get"], url_path="daily-report-xlsx")
    def daily_report_xlsx(self, request):
        """
        Download per-user order report as XLSX.
        GET /api/admin/summary/daily-report-xlsx/?date=YYYY-MM-DD
        """
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        date_str = request.query_params.get("date")
        if not date_str:
            return Response(
                {"error": "date parameter required"}, status=status.HTTP_400_BAD_REQUEST
            )
        try:
            target_date = datetime.date.fromisoformat(date_str)
        except (TypeError, ValueError):
            return Response(
                {"error": "invalid date format, expected YYYY-MM-DD"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        safe_date = target_date.isoformat()  # guaranteed YYYY-MM-DD, safe for filenames

        orders = (
            DailyOrder.objects.filter(date=target_date)
            .select_related("user", "user__settings")
            .order_by("user__email")
        )

        meal_keys = ["breakfast", "lunch", "olovrant"]
        meal_labels = {"breakfast": "Raňajky", "lunch": "Obed", "olovrant": "Olovrant"}

        rows_data = [
            {
                "user": o.user,
                "data": o.data or {},
                "visible_meals": (
                    getattr(getattr(o.user, "settings", None), "visible_meals", None)
                    or ["breakfast", "lunch", "olovrant"]
                ),
            }
            for o in orders
        ]
        sorted_cats = self._xlsx_collect_columns(rows_data, meal_keys)
        col_meta, header_row_1, header_row_2, header_row_3 = (
            self._xlsx_build_column_meta(sorted_cats, meal_keys, meal_labels)
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Prehlad {safe_date}"

        header_fill_main = PatternFill("solid", fgColor="2563EB")
        header_fill_meal = {
            "breakfast": PatternFill("solid", fgColor="F97316"),
            "lunch": PatternFill("solid", fgColor="3B82F6"),
            "olovrant": PatternFill("solid", fgColor="22C55E"),
        }
        header_font = Font(bold=True, color="FFFFFF")
        bold_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")

        ws.append([f"Denný prehľad objednávok — {safe_date}"])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([])  # empty row
        ws.append(header_row_1)
        ws.append(header_row_2)
        ws.append(header_row_3)

        self._xlsx_style_headers(
            ws,
            col_meta,
            sorted_cats,
            meal_keys,
            header_fill_main,
            header_fill_meal,
            header_font,
            center,
        )
        self._xlsx_write_data(ws, rows_data, meal_keys, sorted_cats, bold_font)

        ws.column_dimensions["A"].width = 28
        for c_idx in range(2, len(col_meta) + 1):
            ws.column_dimensions[get_column_letter(c_idx)].width = 12
        # Freeze top 5 rows (title, empty, meal, category, menu/diet) + col A
        ws.freeze_panes = "B6"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"prehlad_{safe_date}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=["get"], url_path="daily-report-pdf")
    def daily_report_pdf(self, request):
        """
        Download per-user order report as portrait PDF.
        GET /api/admin/summary/daily-report-pdf/?date=YYYY-MM-DD
        """
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            HRFlowable,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        date_str = request.query_params.get("date")
        if not date_str:
            return Response(
                {"error": "date parameter required"}, status=status.HTTP_400_BAD_REQUEST
            )
        try:
            target_date = datetime.date.fromisoformat(date_str)
        except (TypeError, ValueError):
            return Response(
                {"error": "invalid date format, expected YYYY-MM-DD"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        safe_date = target_date.isoformat()

        orders = (
            DailyOrder.objects.filter(date=target_date)
            .select_related("user", "user__settings")
            .order_by("user__email")
        )

        CAT_ORDER = ["Jasle", "Škôlka", "ZŠ 1.stupeň", "ZŠ 2.stupeň", "Dospelý (SŠ)"]
        MEAL_LABELS = {
            "breakfast": "Raňajky",
            "lunch": "Obed",
            "olovrant": "Olovrant",
        }

        # ── Register Unicode fonts ──────────────────────────────────────────
        _register_pdf_fonts()
        _font = _PDF_FONT_REGULAR
        _font_bold = _PDF_FONT_BOLD

        # ── Styles ─────────────────────────────────────────────────────────────
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "title",
            parent=styles["Heading1"],
            fontSize=13,
            spaceAfter=4,
            fontName=_font_bold,
            textColor=colors.HexColor("#1e3a5f"),
        )
        user_style = ParagraphStyle(
            "user",
            parent=styles["Heading2"],
            fontSize=11,
            spaceBefore=10,
            spaceAfter=2,
            fontName=_font_bold,
            textColor=colors.HexColor("#111827"),
        )
        meal_style = ParagraphStyle(
            "meal",
            parent=styles["Normal"],
            fontSize=9,
            spaceBefore=4,
            spaceAfter=2,
            textColor=colors.HexColor("#374151"),
            fontName=_font_bold,
        )

        MEAL_COLORS = {
            "breakfast": colors.HexColor("#fff7ed"),
            "lunch": colors.HexColor("#eff6ff"),
            "olovrant": colors.HexColor("#f0fdf4"),
        }
        MEAL_HEADER_COLORS = {
            "breakfast": colors.HexColor("#f97316"),
            "lunch": colors.HexColor("#3b82f6"),
            "olovrant": colors.HexColor("#22c55e"),
        }

        # A4 portrait: usable width = 21cm - 2×1.5cm = 18cm
        PAGE_W = 18 * cm
        COL_WIDTHS = [4.5 * cm, 8 * cm, 5.5 * cm]

        def _build_meal_table(meal_data, meal_key):
            """Return a Table flowable for one meal's data, or None if empty."""
            meal = meal_data.get(meal_key)
            if not isinstance(meal, dict) or not meal:
                return None
            is_flat = "menuCounts" in meal
            if is_flat:
                cat_entries = [(meal_key, meal)]
            else:
                cat_entries = [(k, v) for k, v in meal.items() if isinstance(v, dict)]
            cat_dict = dict(cat_entries)
            ordered = [c for c in CAT_ORDER if c in cat_dict]
            ordered += [c for c in cat_dict if c not in ordered]

            rows = [["Kategória", "Menu", "Špeciálne diéty"]]
            for cat_name in ordered:
                details = cat_dict[cat_name]
                menus_str = ", ".join(
                    f"{k}×{v}"
                    for k, v in sorted((details.get("menuCounts") or {}).items())
                    if self._safe_int(v) > 0
                )
                diets_str = ", ".join(
                    (f"{k}×{v}" if self._safe_int(v) > 1 else k)
                    for k, v in sorted((details.get("diets") or {}).items())
                    if self._safe_int(v) > 0
                )
                if menus_str or diets_str:
                    rows.append([cat_name, menus_str or "–", diets_str or ""])

            if len(rows) == 1:
                return None  # header only → nothing ordered

            bg = MEAL_COLORS[meal_key]
            hdr_bg = MEAL_HEADER_COLORS[meal_key]
            t = Table(rows, colWidths=COL_WIDTHS)
            t.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), hdr_bg),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, -1), _font),
                        ("FONTNAME", (0, 0), (-1, 0), _font_bold),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [bg, colors.white]),
                        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#d1d5db")),
                        ("LEFTPADDING", (0, 0), (-1, -1), 5),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                        ("TOPPADDING", (0, 0), (-1, -1), 3),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ]
                )
            )
            return t

        # ── Build story ────────────────────────────────────────────────────────
        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=A4,
            topMargin=1.5 * cm,
            bottomMargin=1.5 * cm,
            leftMargin=1.5 * cm,
            rightMargin=1.5 * cm,
            title=f"Denný prehľad {safe_date}",
        )

        story = []
        story.append(Paragraph(f"Denný prehľad objednávok — {safe_date}", title_style))
        story.append(
            HRFlowable(width=PAGE_W, thickness=1, color=colors.HexColor("#2563eb"))
        )
        story.append(Spacer(1, 0.3 * cm))

        for order in orders:
            user = order.user
            data = order.data or {}
            _settings = getattr(user, "settings", None)
            visible_meals = getattr(_settings, "visible_meals", None) or [
                "breakfast",
                "lunch",
                "olovrant",
            ]

            display_name = f"{user.first_name} {user.last_name}".strip() or user.email
            story.append(Paragraph(display_name, user_style))

            any_meal = False
            for mk in ["breakfast", "lunch", "olovrant"]:
                if mk not in visible_meals:
                    continue
                tbl = _build_meal_table(data, mk)
                if tbl is None:
                    continue
                story.append(Paragraph(MEAL_LABELS[mk], meal_style))
                story.append(tbl)
                story.append(Spacer(1, 0.15 * cm))
                any_meal = True

            if not any_meal:
                story.append(
                    Paragraph(
                        "Žiadne objednávky",
                        ParagraphStyle(
                            "empty",
                            parent=styles["Normal"],
                            fontSize=8,
                            fontName=_font,
                            textColor=colors.grey,
                        ),
                    )
                )
            story.append(
                HRFlowable(
                    width=PAGE_W,
                    thickness=0.4,
                    color=colors.HexColor("#e5e7eb"),
                    spaceAfter=4,
                )
            )

        doc.build(story)
        buf.seek(0)

        filename = f"prehlad_{safe_date}.pdf"
        response = FileResponse(buf, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _next_workdays(start: datetime.date, count: int = 5) -> list[datetime.date]:
    """Return the next `count` Mon-Fri dates starting from (and including) start."""
    days = []
    d = start
    while len(days) < count:
        if d.weekday() < 5:  # Mon-Fri
            days.append(d)
        d += datetime.timedelta(days=1)
    return days


def _order_total(data: dict) -> tuple[int, dict]:
    """
    Return (total_portions, {meal: count}) for the order data dict.

    Supports both storage shapes:
    - Flat:            {"lunch": {"menuCounts": {"A": 5}}}
    - Category-nested: {"lunch": {"Dospelý": {"menuCounts": {"A": 5}}}}
    """
    meal_count = {"breakfast": 0, "lunch": 0, "olovrant": 0}
    total = 0
    for meal_key in ("breakfast", "lunch", "olovrant"):
        meal = data.get(meal_key, {}) or {}
        if "menuCounts" in meal:
            # Flat shape
            for count in (meal.get("menuCounts") or {}).values():
                c = int(count or 0)
                meal_count[meal_key] += c
                total += c
        else:
            # Category-nested shape
            for _cat, details in meal.items():
                if not isinstance(details, dict):
                    continue
                for count in (details.get("menuCounts") or {}).values():
                    c = int(count or 0)
                    meal_count[meal_key] += c
                    total += c
    return total, meal_count


# ---------------------------------------------------------------------------
# Planned orders endpoint
# ---------------------------------------------------------------------------


class PlannedOrdersViewSet(viewsets.ViewSet):
    """
    Returns the 5 upcoming workdays with order status for the logged-in client.
    GET /api/orders/planned/
    """

    permission_classes = [permissions.IsAuthenticated]

    def list(self, request):
        today = timezone.localdate()
        workdays = _next_workdays(today, 5)

        existing = {
            o.date: o
            for o in DailyOrder.objects.filter(user=request.user, date__in=workdays)
        }

        # Single historical query — avoids N+1 when multiple days have no order.
        # The planned-week cascade (existing dict) is checked first in-memory;
        # this serves as the fallback for every unset day.
        historical_template = _last_non_empty_order(request.user, workdays[0])

        def _template_for_day(day):
            """
            Find the most recent non-empty order before `day`.
            Checks orders already placed in the planned week first (cascade
            forward), then falls back to the pre-fetched historical template.
            No additional DB queries are made here.
            """
            for prev_day in reversed([d for d in workdays if d < day]):
                prev = existing.get(prev_day)
                if prev and not _is_order_empty(prev.data or {}):
                    return prev
            return historical_template

        result = []
        for day in workdays:
            order = existing.get(day)
            if order:
                total, meal_count = _order_total(order.data or {})
                result.append(
                    {
                        "date": str(day),
                        "exists": True,
                        "is_auto": order.is_auto,
                        "is_empty": total == 0,
                        "totalPortions": total,
                        "mealCount": meal_count,
                        "predictedTotal": 0,
                        "predictedMealCount": {
                            "breakfast": 0,
                            "lunch": 0,
                            "olovrant": 0,
                        },
                    }
                )
            else:
                tmpl = _template_for_day(day)
                if tmpl:
                    predicted_total, predicted_meal_count = _order_total(
                        tmpl.data or {}
                    )
                else:
                    predicted_total = 0
                    predicted_meal_count = {"breakfast": 0, "lunch": 0, "olovrant": 0}
                result.append(
                    {
                        "date": str(day),
                        "exists": False,
                        "is_auto": None,
                        "is_empty": None,
                        "totalPortions": 0,
                        "mealCount": {"breakfast": 0, "lunch": 0, "olovrant": 0},
                        "predictedTotal": predicted_total,
                        "predictedMealCount": predicted_meal_count,
                    }
                )

        return Response(result)


# ---------------------------------------------------------------------------
# Admin: manual trigger
# ---------------------------------------------------------------------------


class AdminAutoOrderViewSet(viewsets.ViewSet):
    """
    Admin endpoint to manually trigger auto-order for a given date.
    POST /api/admin/trigger-auto-orders/  { "date": "YYYY-MM-DD" }  (date optional)
    """

    permission_classes = [permissions.IsAdminUser]

    def create(self, request):
        date_str = request.data.get("date")
        target_date = None
        if date_str:
            try:
                target_date = datetime.date.fromisoformat(date_str)
            except ValueError:
                return Response(
                    {"error": "Invalid date format. Use YYYY-MM-DD."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        from .services import apply_auto_orders

        result = apply_auto_orders(target_date)
        return Response(result, status=status.HTTP_200_OK)


class GlobalSettingsViewSet(viewsets.ViewSet):
    """
    Manage system-wide settings.
    Singleton-like behavior: always returns the first instance.
    """

    def get_permissions(self):
        # Allow authenticated users to list (read) settings
        if self.action == "list":
            return [permissions.IsAuthenticated()]
        # Require admin for creation/updates
        return [permissions.IsAdminUser()]

    def list(self, request):
        from .models import GlobalSettings
        from .serializers import GlobalSettingsSerializer

        settings, _ = GlobalSettings.objects.get_or_create(pk=1)
        serializer = GlobalSettingsSerializer(settings)
        return Response(serializer.data)

    def create(self, request):
        """
        Using create/post to update settings for simplicity or
        standard REST conventions.
        """
        from .models import GlobalSettings
        from .serializers import GlobalSettingsSerializer

        settings, _ = GlobalSettings.objects.get_or_create(pk=1)
        serializer = GlobalSettingsSerializer(settings, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# ── Password Reset ─────────────────────────────────────────────────────────────


class PasswordResetRequestView(APIView):
    """
    POST /api/auth/password-reset/
    Body: {"email": "user@example.com"}

    Initiates a password-reset flow.
    Always returns HTTP 200 to avoid leaking whether an email is registered.
    Returns HTTP 429 with a retry_after field when rate-limited.
    """

    permission_classes = [permissions.AllowAny]

    def post(self, request):
        from .password_reset_service import (
            RateLimitExceeded,
            TooSoonError,
            request_password_reset,
        )

        email = request.data.get("email", "").strip()
        if not email:
            return Response(
                {"detail": "Pole e-mail je povinné."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            request_password_reset(email)
        except RateLimitExceeded as exc:
            minutes = round(exc.retry_after_seconds / 60)
            return Response(
                {
                    "detail": (
                        f"Príliš veľa pokusov. Skúste to znova za {minutes} minút."
                    ),
                    "retry_after_seconds": exc.retry_after_seconds,
                },
                status=status.HTTP_429_TOO_MANY_REQUESTS,
            )
        except TooSoonError as exc:
            return Response(
                {
                    "detail": (
                        f"E-mail bol práve odoslaný. "
                        f"Opätovné odoslanie bude možné za {exc.wait_seconds} sekúnd."
                    ),
                    "wait_seconds": exc.wait_seconds,
                },
                status=status.HTTP_429_TOO_MANY_REQUESTS,
            )
        except Exception:
            logger.exception(
                "Unexpected error during password reset request for %s", email
            )
            return Response(
                {
                    "detail": (
                        "Ak je táto e-mailová adresa registrovaná, "
                        "bol na ňu odoslaný odkaz na obnovu hesla."
                    )
                },
                status=status.HTTP_200_OK,
            )

        return Response(
            {
                "detail": (
                    "Ak je táto e-mailová adresa registrovaná, "
                    "bol na ňu odoslaný odkaz na obnovu hesla."
                )
            },
            status=status.HTTP_200_OK,
        )


class PasswordResetConfirmView(APIView):
    """
    POST /api/auth/password-reset/confirm/
    Body: {"token": "<reset_token>", "new_password": "<new_password>"}

    Validates the token and sets the new password.
    """

    permission_classes = [permissions.AllowAny]

    def post(self, request):
        from .password_reset_service import confirm_password_reset

        token = request.data.get("token", "").strip()
        new_password = request.data.get("new_password", "")

        if not token or not new_password:
            return Response(
                {"detail": "Polia token a new_password sú povinné."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            confirm_password_reset(token=token, new_password=new_password)
        except ValueError as exc:
            return Response(
                {"detail": str(exc)},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response(
            {"detail": "Heslo bolo úspešne zmenené. Môžete sa prihlásiť."},
            status=status.HTTP_200_OK,
        )
