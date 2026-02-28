"""XLSX Report Exporter - Generate XLSX reports from order data."""

from __future__ import annotations

import io
from typing import TYPE_CHECKING, Dict, List

from .report_helpers import safe_int

if TYPE_CHECKING:
    from openpyxl.styles import Alignment, Font, PatternFill


class XLSXReportExporter:
    """Generate XLSX reports from order data."""

    # Category order for consistent display
    CAT_ORDER = ["Jasle", "Škôlka", "ZŠ 1.stupeň", "ZŠ 2.stupeň", "Dospelý (SŠ)"]

    # Meal configuration
    MEAL_LABELS = {
        "breakfast": "Raňajky",
        "lunch": "Obed",
        "olovrant": "Olovrant",
    }

    def __init__(
        self, rows_data: List[dict], target_date: str, meal_keys: List[str] = None
    ):
        """
        Initialize exporter.

        Args:
            rows_data: List of dicts with user, data, visible_meals for each order
            target_date: ISO format date string for the report
            meal_keys: List of meal types to include (default: all)
        """
        self.rows_data = rows_data
        self.target_date = target_date
        self.meal_keys = meal_keys or ["breakfast", "lunch", "olovrant"]

    def generate(self) -> bytes:
        """
        Generate XLSX report.

        Returns:
            XLSX file content as bytes
        """
        # Lazy-import openpyxl to avoid startup overhead
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        # Collect unique menu/diet combinations
        sorted_cats = self._collect_columns()

        # Build header metadata
        col_meta, header_row_1, header_row_2, header_row_3 = self._build_column_meta(
            sorted_cats
        )

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Prehlad {self.target_date}"

        # Style definitions
        header_fill_main = PatternFill("solid", fgColor="2563EB")
        header_fill_meal = {
            "breakfast": PatternFill("solid", fgColor="F97316"),
            "lunch": PatternFill("solid", fgColor="3B82F6"),
            "olovrant": PatternFill("solid", fgColor="22C55E"),
        }
        header_font = Font(bold=True, color="FFFFFF")
        bold_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")

        # Add title
        ws.append([f"Denný prehľad objednávok — {self.target_date}"])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([])

        # Add headers
        ws.append(header_row_1)
        ws.append(header_row_2)
        ws.append(header_row_3)

        # Style headers
        self._style_headers(
            ws,
            col_meta,
            sorted_cats,
            header_fill_main,
            header_fill_meal,
            header_font,
            center,
        )

        # Write data
        self._write_data(ws, col_meta, sorted_cats, bold_font)

        # Format columns
        ws.column_dimensions["A"].width = 28
        for c_idx in range(2, len(col_meta) + 1):
            ws.column_dimensions[get_column_letter(c_idx)].width = 12
        ws.freeze_panes = "B6"

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _collect_columns(self) -> Dict:
        """Gather every (category, menu, diet) combination per meal."""
        raw = {m: {} for m in self.meal_keys}

        def _scan(cat_name, details, mk):
            if not isinstance(details, dict):
                return
            for key, cnt in (details.get("menuCounts") or {}).items():
                if safe_int(cnt) > 0:
                    raw[mk].setdefault(cat_name, {"menus": set(), "diets": set()})
                    raw[mk][cat_name]["menus"].add(key)
            for key, cnt in (details.get("diets") or {}).items():
                if safe_int(cnt) > 0:
                    raw[mk].setdefault(cat_name, {"menus": set(), "diets": set()})
                    raw[mk][cat_name]["diets"].add(key)

        for row in self.rows_data:
            data = row["data"]
            for mk in self.meal_keys:
                meal = data.get(mk) or {}
                if not isinstance(meal, dict):
                    continue
                if "menuCounts" in meal:
                    _scan(mk, meal, mk)
                else:
                    for cat_name, details in meal.items():
                        _scan(cat_name, details, mk)

        sorted_cats = {}
        for mk in self.meal_keys:
            sorted_cat_keys = sorted(
                raw[mk].keys(),
                key=lambda c: (
                    self.CAT_ORDER.index(c) if c in self.CAT_ORDER else 99,
                    c,
                ),
            )
            sorted_cats[mk] = {
                cat: {
                    "menus": sorted(raw[mk][cat]["menus"]),
                    "diets": sorted(raw[mk][cat]["diets"]),
                }
                for cat in sorted_cat_keys
            }
        return sorted_cats

    def _build_column_meta(self, sorted_cats: Dict) -> tuple:
        """Build 3 header rows and column metadata list."""
        col_meta = [("fixed", None, "name", None)]
        header_row_1 = ["Klient"]
        header_row_2 = [""]
        header_row_3 = [""]

        for mk in self.meal_keys:
            cats = sorted_cats[mk]
            inner_cols = sum(len(v["menus"]) + len(v["diets"]) for v in cats.values())
            meal_span = inner_cols + 1
            header_row_1 += [self.MEAL_LABELS[mk]] + [""] * (meal_span - 1)

            for cat_name, cat_data in cats.items():
                cat_span = len(cat_data["menus"]) + len(cat_data["diets"])
                header_row_2 += [cat_name] + [""] * max(cat_span - 1, 0)
                for menu_key in cat_data["menus"]:
                    header_row_3.append(f"Menu {menu_key}")
                    col_meta.append((mk, cat_name, "menu", menu_key))
                for diet_name in cat_data["diets"]:
                    header_row_3.append(diet_name)
                    col_meta.append((mk, cat_name, "diet", diet_name))

            header_row_2.append("Spolu")
            header_row_3.append("")
            col_meta.append((mk, None, "total", None))

        header_row_1.append("Celkovo")
        header_row_2.append("")
        header_row_3.append("")
        col_meta.append(("fixed", None, "grand_total", None))
        return col_meta, header_row_1, header_row_2, header_row_3

    def _style_headers(
        self,
        ws,
        col_meta: List,
        sorted_cats: Dict,
        header_fill_main: PatternFill,
        header_fill_meal: Dict[str, PatternFill],
        header_font: Font,
        center: Alignment,
    ) -> None:
        """Apply fills, fonts, merges to header rows."""
        total_cols = len(col_meta)
        current_col = 2

        for mk in self.meal_keys:
            cats = sorted_cats[mk]
            inner_cols = sum(len(v["menus"]) + len(v["diets"]) for v in cats.values())
            meal_span = inner_cols + 1
            if meal_span > 1:
                ws.merge_cells(
                    start_row=3,
                    start_column=current_col,
                    end_row=3,
                    end_column=current_col + meal_span - 1,
                )
            cell = ws.cell(row=3, column=current_col)
            cell.fill = header_fill_meal[mk]
            cell.font = header_font
            cell.alignment = center
            current_col += meal_span

        for c in [1, total_cols]:
            cell = ws.cell(row=3, column=c)
            cell.fill = header_fill_main
            cell.font = header_font
            cell.alignment = center

        row45_fill = {}
        current_col = 2
        for mk in self.meal_keys:
            for cat_name, cat_data in sorted_cats[mk].items():
                cat_span = len(cat_data["menus"]) + len(cat_data["diets"])
                if cat_span > 0:
                    if cat_span > 1:
                        ws.merge_cells(
                            start_row=4,
                            start_column=current_col,
                            end_row=4,
                            end_column=current_col + cat_span - 1,
                        )
                    for c in range(current_col, current_col + cat_span):
                        row45_fill[c] = header_fill_meal[mk]
                    cell = ws.cell(row=4, column=current_col)
                    cell.fill = header_fill_meal[mk]
                    cell.font = header_font
                    cell.alignment = center
                    current_col += cat_span
            row45_fill[current_col] = header_fill_meal[mk]
            cell = ws.cell(row=4, column=current_col)
            cell.fill = header_fill_meal[mk]
            cell.font = header_font
            cell.alignment = center
            current_col += 1

        for c in [1, total_cols]:
            cell = ws.cell(row=4, column=c)
            cell.fill = header_fill_main
            cell.font = header_font
            cell.alignment = center

        for c_idx in range(1, total_cols + 1):
            cell = ws.cell(row=5, column=c_idx)
            cell.fill = row45_fill.get(c_idx, header_fill_main)
            cell.font = header_font
            cell.alignment = center

        ws.merge_cells(start_row=3, start_column=1, end_row=5, end_column=1)
        ws.merge_cells(
            start_row=3, start_column=total_cols, end_row=5, end_column=total_cols
        )

    def _write_data(
        self,
        ws,
        col_meta: List,
        sorted_cats: Dict,
        bold_font: Font,
    ) -> None:
        """Append per-user data rows and SPOLU totals row."""
        totals = {
            mk: {
                cat: {
                    "menus": {m: 0 for m in cat_data["menus"]},
                    "diets": {d: 0 for d in cat_data["diets"]},
                }
                for cat, cat_data in cats.items()
            }
            for mk, cats in sorted_cats.items()
        }
        meal_totals = {mk: 0 for mk in self.meal_keys}
        grand_total = 0

        for row_info in self.rows_data:
            user = row_info["user"]
            data = row_info["data"]
            visible_meals = row_info.get("visible_meals") or self.meal_keys
            display_name = f"{user.first_name} {user.last_name}".strip() or user.email
            row_vals = [display_name]
            row_grand = 0

            for mk in self.meal_keys:
                meal_col_count = (
                    sum(
                        len(cat_data["menus"]) + len(cat_data["diets"])
                        for cat_data in sorted_cats[mk].values()
                    )
                    + 1
                )
                if mk not in visible_meals:
                    row_vals.extend([""] * meal_col_count)
                    continue

                meal = data.get(mk) or {}
                meal_total = 0
                is_flat = "menuCounts" in meal
                for cat_name, cat_data in sorted_cats[mk].items():
                    if is_flat:
                        cat_details = meal if cat_name == mk else {}
                    else:
                        cat_details = meal.get(cat_name) or {}
                    for menu_key in cat_data["menus"]:
                        cnt = safe_int(
                            (cat_details.get("menuCounts") or {}).get(menu_key, 0)
                        )
                        row_vals.append(cnt or "")
                        totals[mk][cat_name]["menus"][menu_key] += cnt
                        meal_total += cnt
                    for diet_name in cat_data["diets"]:
                        cnt = safe_int(
                            (cat_details.get("diets") or {}).get(diet_name, 0)
                        )
                        row_vals.append(cnt or "")
                        totals[mk][cat_name]["diets"][diet_name] += cnt
                row_vals.append(meal_total or "")
                meal_totals[mk] += meal_total
                row_grand += meal_total
            row_vals.append(row_grand or "")
            grand_total += row_grand
            ws.append(row_vals)

        total_row = ["SPOLU"]
        for mk in self.meal_keys:
            for cat_name, cat_data in sorted_cats[mk].items():
                for menu_key in cat_data["menus"]:
                    total_row.append(
                        totals[mk][cat_name]["menus"].get(menu_key, 0) or ""
                    )
                for diet_name in cat_data["diets"]:
                    total_row.append(
                        totals[mk][cat_name]["diets"].get(diet_name, 0) or ""
                    )
            total_row.append(meal_totals[mk] or "")
        total_row.append(grand_total or "")
        ws.append(total_row)

        totals_row_num = ws.max_row
        for c_idx in range(1, ws.max_column + 1):
            ws.cell(row=totals_row_num, column=c_idx).font = bold_font
