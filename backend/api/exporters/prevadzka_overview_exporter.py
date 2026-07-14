"""Exportéry pre admin prehľad dodania podkladov (EduPage vs app prevádzky)."""

from __future__ import annotations

import io
from typing import Any

from .pdf_exporter import PDFFontManager

MEAL_HEADERS = ["R", "Ob", "Ol", "Spolu"]
COUNT_KEYS = ["breakfast", "lunch", "olovrant", "total"]

CATEGORIES = [
    ("edupage", "EduPage prevádzky"),
    ("app", "App prevádzky"),
]


def _status_label(row: dict[str, Any]) -> str:
    if not row["delivered"]:
        return "NEDODANÉ"
    if row["has_warning"]:
        return "SKONTROLUJ"
    return "OK"


def _flags_text(row: dict[str, Any]) -> str:
    flags = row.get("flags", {})
    return "; ".join([*flags.get("config_notes", []), *flags.get("attention", [])])


class PrevadzkaOverviewXLSXExporter:
    """Prehľad dodania podkladov ako XLSX (jeden hárok, dve sekcie)."""

    def __init__(self, payload: dict[str, Any]):
        self.payload = payload

    def generate(self) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Dodanie podkladov"

        bold = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="E8EEF5")
        section_fill = PatternFill("solid", fgColor="D6E0EB")
        center = Alignment(horizontal="center")

        ws.append([f"Dodanie podkladov — {self.payload['date']}"])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([])

        headers = ["Stav", "Prevádzka", "Celok", *MEAL_HEADERS, "Poznámky"]

        for key, label in CATEGORIES:
            rows = self.payload.get(key, [])
            delivered = sum(1 for r in rows if r["delivered"])
            ws.append([f"{label} ({delivered}/{len(rows)} dodané)"])
            section_row = ws.max_row
            for col in range(1, len(headers) + 1):
                ws.cell(row=section_row, column=col).fill = section_fill
            ws.cell(row=section_row, column=1).font = bold

            ws.append(headers)
            head_row = ws.max_row
            for col in range(1, len(headers) + 1):
                c = ws.cell(row=head_row, column=col)
                c.font = bold
                c.fill = header_fill
                c.alignment = center

            for r in rows:
                ws.append(
                    [
                        _status_label(r),
                        r["nazov"],
                        r["celok"],
                        *[r["counts"][k] for k in COUNT_KEYS],
                        _flags_text(r),
                    ]
                )
            ws.append([])

        widths = [12, 28, 24, 6, 6, 6, 8, 40]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


class PrevadzkaOverviewPDFExporter:
    """Prehľad dodania podkladov ako PDF (portrait A4)."""

    def __init__(self, payload: dict[str, Any]):
        self.payload = payload

    def generate(self) -> bytes:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        font_regular, font_bold = PDFFontManager.get_fonts()
        base = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "ovr_title",
            parent=base["Title"],
            fontName=font_bold,
            fontSize=16,
        )
        section_style = ParagraphStyle(
            "ovr_section",
            parent=base["Heading2"],
            fontName=font_bold,
            fontSize=12,
            spaceBefore=10,
            spaceAfter=6,
        )

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=A4,
            leftMargin=1.5 * cm,
            rightMargin=1.5 * cm,
            topMargin=1.5 * cm,
            bottomMargin=1.5 * cm,
        )
        story: list[Any] = [
            Paragraph(f"Dodanie podkladov — {self.payload['date']}", title_style)
        ]

        for key, label in CATEGORIES:
            rows = self.payload.get(key, [])
            delivered = sum(1 for r in rows if r["delivered"])
            story.append(
                Paragraph(f"{label} ({delivered}/{len(rows)} dodané)", section_style)
            )

            table_rows = [["Stav", "Prevádzka", *MEAL_HEADERS, "Poznámky"]]
            for r in rows:
                table_rows.append(
                    [
                        _status_label(r),
                        r["nazov"],
                        *[str(r["counts"][k]) for k in COUNT_KEYS],
                        _flags_text(r),
                    ]
                )

            table = Table(
                table_rows,
                colWidths=[
                    2.2 * cm,
                    4.5 * cm,
                    1.0 * cm,
                    1.0 * cm,
                    1.0 * cm,
                    1.3 * cm,
                    None,
                ],
                repeatRows=1,
            )
            style = [
                ("FONTNAME", (0, 0), (-1, -1), font_regular),
                ("FONTNAME", (0, 0), (-1, 0), font_bold),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D6E0EB")),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#C0C8D0")),
                ("ALIGN", (2, 0), (5, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
            for i, r in enumerate(rows, start=1):
                if not r["delivered"]:
                    style.append(
                        ("TEXTCOLOR", (0, i), (0, i), colors.HexColor("#C0392B"))
                    )
                elif r["has_warning"]:
                    style.append(
                        ("TEXTCOLOR", (0, i), (0, i), colors.HexColor("#B9770E"))
                    )
                else:
                    style.append(
                        ("TEXTCOLOR", (0, i), (0, i), colors.HexColor("#1E8449"))
                    )
            table.setStyle(TableStyle(style))
            story.append(table)
            story.append(Spacer(1, 0.3 * cm))

        doc.build(story)
        return buf.getvalue()
