"""XLSX exporter for the gramage dashboard (orders × templates × coefficients).

Mirrors the AdminDashboard "Gramáž jedál" screen: same meal-hue header
colours, same diet colours, same delivery-block/route grouping and portion
summaries — see gramage_dashboard_export.py for the shared presentation
logic this and the PDF exporter both build on.
"""

from __future__ import annotations

import io

from ..services.meal_plan_service import _tidy_count
from .gramage_dashboard_export import (
    BRAND,
    MEAL_PALETTE,
    component_subtitle,
    contrast_text_color,
    diet_color,
    flat_client_rows,
    group_label,
    meal_hue,
)


class GramageDashboardXLSXExporter:
    def __init__(self, data: dict):
        self.data = data

    def generate(self) -> bytes:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.data["date"]

        col_groups = self.data["col_groups"]
        totals = self.data["totals"]
        all_rows = flat_client_rows(self.data)

        # ── Brand fonts & fills (mirrors admin.css BRAND palette / .mh-*) ────
        title_font = Font(bold=True, size=13, color=BRAND["green_900"])
        hdr_font = Font(bold=True, color="FFFFFF")
        cat_font = Font(bold=True)
        cat_fill = PatternFill("solid", fgColor=BRAND["cream_soft"])
        band_font = Font(bold=True, color="FFFFFF")
        block_fill = PatternFill("solid", fgColor=BRAND["green_900"])
        route_fill = PatternFill("solid", fgColor=BRAND["green_700"])
        unassigned_fill = PatternFill("solid", fgColor="C64545")
        standard_fill = PatternFill("solid", fgColor=BRAND["cream"])
        total_font = Font(bold=True, color="FFFFFF")
        total_fill = PatternFill("solid", fgColor=BRAND["green_900"])
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        right_align = Alignment(horizontal="right")
        thin_border = Border(
            left=Side(style="thin", color=BRAND["line"]),
            right=Side(style="thin", color=BRAND["line"]),
            top=Side(style="thin", color=BRAND["line"]),
            bottom=Side(style="thin", color=BRAND["line"]),
        )

        def meal_fills(meal, variant):
            dark_hex, mid_hex, _light_hex = MEAL_PALETTE[meal_hue(meal, variant)]
            return (
                PatternFill("solid", fgColor=dark_hex),
                PatternFill("solid", fgColor=mid_hex),
            )

        def diet_fill_for(row):
            hex_color = diet_color(self.data, row)
            return PatternFill("solid", fgColor=hex_color), Font(
                color=contrast_text_color(hex_color)
            )

        # ── Column layout ───────────────────────────────────────────────────
        # Columns: A=Prevádzka/Riadok, B=Počet, then components
        BASE_COLS = 2  # A, B
        col_start = []  # 1-based start column for each col_group
        cur = BASE_COLS + 1
        for cg in col_groups:
            col_start.append(cur)
            cur += len(cg["components"])
        total_cols = cur - 1

        # ── Title row ───────────────────────────────────────────────────────
        ws.cell(row=1, column=1, value=f"Gramáž jedál — {self.data['date']}")
        ws["A1"].font = title_font
        ws.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=max(total_cols, 2)
        )
        ws.row_dimensions[1].height = 22
        ws.append([])  # blank row 2

        # ── Header row 1: Prevádzka, Počet, meal group labels ────────────────
        HDR_ROW = 3
        ws.cell(row=HDR_ROW, column=1, value="Prevádzka / Riadok")
        ws.cell(row=HDR_ROW, column=2, value="Počet")
        ws.merge_cells(
            start_row=HDR_ROW, start_column=1, end_row=HDR_ROW + 1, end_column=1
        )
        ws.merge_cells(
            start_row=HDR_ROW, start_column=2, end_row=HDR_ROW + 1, end_column=2
        )

        for i, cg in enumerate(col_groups):
            c = col_start[i]
            ws.cell(row=HDR_ROW, column=c, value=group_label(cg))
            if len(cg["components"]) > 1:
                ws.merge_cells(
                    start_row=HDR_ROW,
                    start_column=c,
                    end_row=HDR_ROW,
                    end_column=c + len(cg["components"]) - 1,
                )

        # ── Header row 2: component labels with base gramage ────────────────
        for i, cg in enumerate(col_groups):
            for j, comp in enumerate(cg["components"]):
                ws.cell(
                    row=HDR_ROW + 1,
                    column=col_start[i] + j,
                    value=f"{comp['label']} ({component_subtitle(comp)})",
                )

        # Style both header rows in the base identity colours…
        for r in (HDR_ROW, HDR_ROW + 1):
            for c in range(1, total_cols + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = hdr_font
                cell.fill = block_fill
                cell.alignment = center

        # …then recolour each meal-group's two header cells to its hue: dark
        # for the merged label row, mid tone for the component subtitle row.
        for i, cg in enumerate(col_groups):
            dark_fill, mid_fill = meal_fills(cg["meal"], cg.get("variant"))
            for j in range(len(cg["components"])):
                ws.cell(row=HDR_ROW, column=col_start[i] + j).fill = dark_fill
                ws.cell(row=HDR_ROW + 1, column=col_start[i] + j).fill = mid_fill

        DATA_ROW = HDR_ROW + 2

        def write_row(label, count, col_grams, font=None, fill=None, indent=0):
            nonlocal DATA_ROW
            ws.cell(row=DATA_ROW, column=1, value=("  " * indent) + label)
            ws.cell(row=DATA_ROW, column=2, value=count)
            ws.cell(row=DATA_ROW, column=2).alignment = right_align
            for i, grams in enumerate(col_grams):
                for j, g in enumerate(grams):
                    c = col_start[i] + j
                    try:
                        ws.cell(row=DATA_ROW, column=c, value=float(g))
                    except (ValueError, TypeError):
                        ws.cell(row=DATA_ROW, column=c, value=g)
            if font or fill:
                for c in range(1, total_cols + 1):
                    cell = ws.cell(row=DATA_ROW, column=c)
                    if font:
                        cell.font = font
                    if fill:
                        cell.fill = fill
            DATA_ROW += 1

        def write_summary_row(label, count, col_grams, fill, font=None):
            write_row(label, count, col_grams, font=font or cat_font, fill=fill)

        def write_client(row, zebra: bool):
            nonlocal DATA_ROW
            ws.cell(
                row=DATA_ROW,
                column=1,
                value=f"{row['client']}  (spolu porcii {row.get('total_count', 0)})",
            )
            ws.merge_cells(
                start_row=DATA_ROW,
                start_column=1,
                end_row=DATA_ROW,
                end_column=total_cols,
            )
            client_fill = cat_fill if zebra else PatternFill("solid", fgColor="FFFFFF")
            for c in range(1, total_cols + 1):
                cell = ws.cell(row=DATA_ROW, column=c)
                cell.font = cat_font
                cell.fill = client_fill
            DATA_ROW += 1

            for sr in row["sub_rows"]:
                is_diet = sr["type"] in {"diet", "zvlast"}
                fill, font = diet_fill_for(sr) if is_diet else (None, None)
                write_row(
                    sr["label"],
                    sr["count"],
                    sr["col_grams"],
                    font=font,
                    fill=fill,
                    indent=1 if is_diet else 0,
                )

            write_summary_row(
                "Súčet bez diét",
                row.get("standard_total_count", 0),
                row.get("standard_col_grams", []),
                standard_fill,
            )
            for diet_row in row.get("diet_summary_rows", []):
                fill, font = diet_fill_for(diet_row)
                write_summary_row(
                    diet_row["name"],
                    diet_row["count"],
                    diet_row["col_grams"],
                    fill,
                    font=font,
                )

            if row.get("delivery_note"):
                ws.cell(
                    row=DATA_ROW, column=1, value=f"Poznámka: {row['delivery_note']}"
                )
                ws.merge_cells(
                    start_row=DATA_ROW,
                    start_column=1,
                    end_row=DATA_ROW,
                    end_column=total_cols,
                )
                for c in range(1, total_cols + 1):
                    ws.cell(row=DATA_ROW, column=c).fill = PatternFill(
                        "solid", fgColor="EEF2E3"
                    )
                DATA_ROW += 1

        def write_band(label, fill, font=None):
            nonlocal DATA_ROW
            ws.cell(row=DATA_ROW, column=1, value=label)
            ws.merge_cells(
                start_row=DATA_ROW,
                start_column=1,
                end_row=DATA_ROW,
                end_column=total_cols,
            )
            for c in range(1, total_cols + 1):
                cell = ws.cell(row=DATA_ROW, column=c)
                cell.fill = fill
                cell.font = font or cat_font
            DATA_ROW += 1

        # ── Data rows ────────────────────────────────────────────────────────
        blocks = self.data.get("blocks") or []
        unassigned_rows = self.data.get("unassigned_rows") or []
        if blocks:
            for block in blocks:
                write_band(block["name"], block_fill, band_font)
                for route in block.get("routes", []):
                    suffix = []
                    if route.get("departure_time"):
                        suffix.append(route["departure_time"][:5])
                    if route.get("driver"):
                        suffix.append(route["driver"])
                    route_label = route["name"]
                    if suffix:
                        route_label = f"{route_label} - {' / '.join(suffix)}"
                    write_band(route_label, route_fill, band_font)
                    for zebra, row in enumerate(route.get("rows", [])):
                        write_client(row, zebra % 2 == 1)
            if unassigned_rows:
                write_band("Nepriradené prevádzky", unassigned_fill, band_font)
                for zebra, row in enumerate(unassigned_rows):
                    write_client(row, zebra % 2 == 1)
        else:
            for zebra, row in enumerate(all_rows):
                write_client(row, zebra % 2 == 1)

        DATA_START_ROW = HDR_ROW + 2

        # ── Totals row ───────────────────────────────────────────────────────
        totals_count = _tidy_count(
            sum(
                sum(sr["count"] for sr in r["sub_rows"] if sr["type"] == "standard")
                for r in all_rows
            )
        )
        total_col_grams = [[g for g in grp] for grp in totals]
        ws.cell(row=DATA_ROW, column=1, value="CELKOM")
        ws.cell(row=DATA_ROW, column=2, value=totals_count)
        for i, grams in enumerate(total_col_grams):
            for j, g in enumerate(grams):
                try:
                    ws.cell(row=DATA_ROW, column=col_start[i] + j, value=float(g))
                except (ValueError, TypeError):
                    ws.cell(row=DATA_ROW, column=col_start[i] + j, value=g)
        for c in range(1, total_cols + 1):
            cell = ws.cell(row=DATA_ROW, column=c)
            cell.font = total_font
            cell.fill = total_fill

        # ── Column widths ────────────────────────────────────────────────────
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 9
        for i, cg in enumerate(col_groups):
            for j in range(len(cg["components"])):
                col_letter = openpyxl.utils.get_column_letter(col_start[i] + j)
                ws.column_dimensions[col_letter].width = 12

        # Freeze the row/column labels so long lists stay orientable while
        # scrolling — the header rows and the "Prevádzka / Riadok" column.
        ws.freeze_panes = ws.cell(row=DATA_START_ROW, column=3).coordinate

        # ── Count summary ────────────────────────────────────────────────────
        count_summary = self.data.get("count_summary", [])
        if count_summary:
            DATA_ROW += 2  # blank separator
            ws.cell(row=DATA_ROW, column=1, value="Súhrn objednávok")
            ws.cell(row=DATA_ROW, column=1).font = Font(
                bold=True, size=12, color=BRAND["green_900"]
            )
            DATA_ROW += 1

            for section in count_summary:
                if not section.get("standard") and not section.get("diets"):
                    continue
                dark_hex, _mid_hex, light_hex = MEAL_PALETTE[
                    meal_hue(section.get("meal", ""), section.get("variant"))
                ]
                ws.cell(row=DATA_ROW, column=1, value=section["label"])
                ws.merge_cells(
                    start_row=DATA_ROW,
                    start_column=1,
                    end_row=DATA_ROW,
                    end_column=2,
                )
                for c in range(1, 3):
                    cell = ws.cell(row=DATA_ROW, column=c)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor=dark_hex)
                    cell.alignment = center
                DATA_ROW += 1
                for row in section.get("standard", []):
                    ws.cell(row=DATA_ROW, column=1, value=f"  {row['name']}")
                    ws.cell(row=DATA_ROW, column=2, value=row["count"])
                    ws.cell(row=DATA_ROW, column=2).alignment = right_align
                    for c in range(1, 3):
                        ws.cell(row=DATA_ROW, column=c).fill = PatternFill(
                            "solid", fgColor=light_hex
                        )
                    DATA_ROW += 1
                for row in section.get("diets", []):
                    fill, font = diet_fill_for({"name": row["label"]})
                    ws.cell(row=DATA_ROW, column=1, value=f"  {row['label']}")
                    ws.cell(row=DATA_ROW, column=2, value=row["count"])
                    ws.cell(row=DATA_ROW, column=2).alignment = right_align
                    for c in range(1, 3):
                        cell = ws.cell(row=DATA_ROW, column=c)
                        cell.fill = fill
                        cell.font = font
                    DATA_ROW += 1

        for row in ws.iter_rows(
            min_row=HDR_ROW,
            max_row=ws.max_row,
            min_col=1,
            max_col=ws.max_column,
        ):
            for cell in row:
                cell.border = thin_border

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
