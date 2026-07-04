"""XLSX Exporter for the Jedálniček (Meal Plan) module."""

from __future__ import annotations

import io
from typing import List

from ..models import MealCategory


class MealPlanXLSXExporter:
    """
    Generate XLSX gramage reports from MealPlanService.calculate_gramage() data.

    Accepts a list of daily gramage dicts so the same class handles both
    single-day and multi-day exports.  Single-day → one sheet.
    Multi-day → one sheet per day + a "Súhrn" summary sheet.
    """

    # Single source of truth for category labels: MealCategory.choices.
    SECTION_LABELS = dict(MealCategory.choices)

    def __init__(self, gramage_list: List[dict]):
        self.gramage_list = gramage_list

    def generate(self) -> bytes:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default sheet

        # Styles
        title_font = Font(bold=True, size=13)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2563EB")
        section_fill = PatternFill("solid", fgColor="DBEAFE")
        section_font = Font(bold=True)
        total_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        right = Alignment(horizontal="right")

        for gramage in self.gramage_list:
            ws = wb.create_sheet(title=gramage["date"])
            self._write_day_sheet(
                ws,
                gramage,
                title_font,
                header_font,
                header_fill,
                section_fill,
                section_font,
                total_font,
                center,
                right,
            )

        if len(self.gramage_list) > 1:
            ws_sum = wb.create_sheet(title="Súhrn", index=0)
            self._write_summary_sheet(
                ws_sum, title_font, header_font, header_fill, total_font, center
            )

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _write_day_sheet(
        self,
        ws,
        gramage,
        title_font,
        header_font,
        header_fill,
        section_fill,
        section_font,
        total_font,
        center,
        right,
    ):
        from openpyxl.styles import Font

        date_str = gramage["date"]
        enrolled = gramage["enrolled_summary"]
        sections = gramage["sections"]
        unit_columns = self._unit_columns(sections)

        # Title
        ws.append([f"Jedálniček — {date_str}"])
        ws["A1"].font = title_font
        ws.append([])

        # Enrolled summary header
        ws.append(["Typ porcie", "Koeficient", "Počet osôb"])
        for cell in ws[ws.max_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        for e in enrolled:
            pct = int(float(e["coefficient"]) * 100)
            ws.append([e["portion_type"], f"{pct}%", e["count"]])
        ws.append([])

        # Column headers for meal items
        portion_names = [e["portion_type"] for e in enrolled]
        header_row = (
            ["Sekcia", "Šablóna", "Základ (g)"]
            + portion_names
            + [self._unit_column_label(col) for col in unit_columns]
            + ["Celkom (g)"]
        )
        ws.append(header_row)
        for cell in ws[ws.max_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        grand_total = 0.0

        def write_section(label, items_list, section_total):
            nonlocal grand_total
            if not items_list:
                return
            # Section label row
            section_row = [label] + [""] * (len(header_row) - 1)
            ws.append(section_row)
            for cell in ws[ws.max_row]:
                cell.fill = section_fill
                cell.font = section_font

            for item in items_list:
                counts_by_pt = {
                    b["portion_type"]: b["total_grams"] for b in item["breakdown"]
                }
                units_by_col = {
                    (
                        u.get("component_label", ""),
                        u.get("unit", "ks"),
                        u["portion_type"],
                    ): u["total_units"]
                    for u in item.get("unit_breakdown", [])
                }
                portion_cols = [
                    counts_by_pt.get(name, "0.00") for name in portion_names
                ]
                unit_cols = [
                    units_by_col.get(
                        (col["component_label"], col["unit"], col["portion_type"]),
                        "",
                    )
                    for col in unit_columns
                ]
                row = (
                    [
                        "",
                        item["template_name"],
                        item["base_weight_grams"],
                    ]
                    + portion_cols
                    + unit_cols
                    + [item["item_total_grams"]]
                )
                ws.append(row)

            # Section total row
            total_row = (
                ["", "SPOLU sekcia", ""]
                + [""] * len(portion_names)
                + [""] * len(unit_columns)
                + [section_total]
            )
            ws.append(total_row)
            for cell in ws[ws.max_row]:
                cell.font = total_font
            grand_total += float(section_total)

        for category, label in self.SECTION_LABELS.items():
            section = sections.get(category, {})
            write_section(
                label,
                section.get("items", []),
                section.get("section_total_grams", "0.00"),
            )

        # Grand total
        ws.append([])
        gt_row = (
            ["CELKOM (g)", "", ""]
            + [""] * len(portion_names)
            + [""] * len(unit_columns)
            + [gramage["grand_total_grams"]]
        )
        ws.append(gt_row)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True, size=12)

        # Auto-size columns
        for col_cells in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(
                max_len + 4, 40
            )

    @staticmethod
    def _unit_columns(sections: dict) -> list[dict]:
        columns: list[dict] = []
        seen: set[tuple[str, str, str]] = set()
        for section in sections.values():
            for item in section.get("items", []):
                for unit in item.get("unit_breakdown", []):
                    key = (
                        unit.get("component_label", ""),
                        unit.get("unit", "ks"),
                        unit["portion_type"],
                    )
                    if key in seen:
                        continue
                    seen.add(key)
                    columns.append(
                        {
                            "component_label": key[0],
                            "unit": key[1],
                            "portion_type": key[2],
                        }
                    )
        return columns

    @staticmethod
    def _unit_column_label(column: dict) -> str:
        label = column["component_label"] or "Kusy"
        return f"{column['portion_type']} - {label} ({column['unit']})"

    def _write_summary_sheet(
        self, ws, title_font, header_font, header_fill, total_font, center
    ):
        ws.append(["Súhrn jedálničkov"])
        ws["A1"].font = title_font
        ws.append([])

        ws.append(["Dátum", "Celkom gramáž (g)", "Poznámky"])
        for cell in ws[ws.max_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        grand = 0.0
        for g in self.gramage_list:
            ws.append([g["date"], g["grand_total_grams"], g.get("notes", "")])
            grand += float(g["grand_total_grams"])

        ws.append(["CELKOM", str(round(grand, 2)), ""])
        for cell in ws[ws.max_row]:
            cell.font = total_font

        for col_cells in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(
                max_len + 4, 40
            )
