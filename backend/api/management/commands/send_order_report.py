"""
Management command: send_order_report

Generates an XLSX daily order report and emails it to the configured
recipients stored in GlobalSettings.report_email_recipients.

Usage:
    python manage.py send_order_report           # yesterday (--days 1)
    python manage.py send_order_report --days 0  # today
    python manage.py send_order_report --date YYYY-MM-DD
"""

import datetime
import io
import logging

from django.core.management.base import BaseCommand

from api.email_utils import send_daily_report_email
from api.models import DailyOrder, GlobalSettings
from api.views import AdminSummaryViewSet

logger = logging.getLogger(__name__)

_MEAL_KEYS = ["breakfast", "lunch", "olovrant"]
_MEAL_LABELS = {"breakfast": "Raňajky", "lunch": "Obed", "olovrant": "Olovrant"}


def build_xlsx_bytes(
    target_date: datetime.date, meals: list[str] | None = None
) -> bytes:
    """Generate an XLSX report for *target_date* and optional meal filter.

    Args:
        target_date: The date to generate the report for
        meals: List of meals to include (e.g., ["breakfast"] or ["breakfast", "lunch", "olovrant"])
               If None, includes all meals.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    if meals is None:
        meals = _MEAL_KEYS
    else:
        # Ensure meals is a subset of valid meals
        meals = [m for m in meals if m in _MEAL_KEYS]
        if not meals:
            meals = _MEAL_KEYS

    safe_date = target_date.isoformat()

    orders = (
        DailyOrder.objects.filter(date=target_date)
        .select_related("user", "user__settings")
        .order_by("user__email")
    )

    rows_data = [
        {
            "user": o.user,
            "data": o.data or {},
            "visible_meals": (
                getattr(getattr(o.user, "settings", None), "visible_meals", None)
                or _MEAL_KEYS
            ),
        }
        for o in orders
    ]

    vs = AdminSummaryViewSet
    sorted_cats = vs._xlsx_collect_columns(rows_data, meals)
    col_meta, header_row_1, header_row_2, header_row_3 = vs._xlsx_build_column_meta(
        sorted_cats, meals, _MEAL_LABELS
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Prehlad {safe_date}"

    header_fill_main = PatternFill("solid", fgColor="2563EB")
    header_fill_meal = {
        "breakfast": PatternFill("solid", fgColor="F97316"),
        "lunch": PatternFill("solid", fgColor="3B82F6"),
        "olovrant": PatternFill("solid", fgColor="22C55E"),
    }
    header_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    ws.append([f"Denný prehľad objednávok — {safe_date}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(header_row_1)
    ws.append(header_row_2)
    ws.append(header_row_3)

    vs._xlsx_style_headers(
        ws,
        col_meta,
        sorted_cats,
        meals,
        header_fill_main,
        header_fill_meal,
        header_font,
        center,
    )
    vs._xlsx_write_data(ws, rows_data, meals, sorted_cats, bold_font)

    ws.column_dimensions["A"].width = 28
    for c_idx in range(2, len(col_meta) + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 12
    ws.freeze_panes = "B6"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


class Command(BaseCommand):
    help = "Generate the daily XLSX order report and email it to configured recipients."

    def add_arguments(self, parser):
        parser.add_argument(
            "--days",
            type=int,
            default=1,
            help="Days ago to report on (default: 1 = yesterday). Ignored when --date is set.",
        )
        parser.add_argument(
            "--date",
            type=str,
            default=None,
            help="Explicit target date in YYYY-MM-DD format.",
        )
        parser.add_argument(
            "--meals",
            type=str,
            default="breakfast,lunch,olovrant",
            help="Comma-separated list of meals to include (breakfast, lunch, olovrant). Default: all.",
        )

    def handle(self, *args, **options):
        # ── Resolve target date ───────────────────────────────────────────────
        if options.get("date"):
            try:
                target_date = datetime.date.fromisoformat(options["date"])
            except ValueError:
                self.stderr.write(
                    self.style.ERROR(
                        f"Invalid date: {options['date']}. Use YYYY-MM-DD."
                    )
                )
                return
        else:
            target_date = datetime.date.today() - datetime.timedelta(
                days=options["days"]
            )

        # ── Get recipients ────────────────────────────────────────────────────
        global_settings, _ = GlobalSettings.objects.get_or_create(pk=1)
        recipients: list = global_settings.report_email_recipients or []

        if not recipients:
            self.stdout.write(
                self.style.WARNING(
                    "No report email recipients configured in system settings. Skipping."
                )
            )
            return

        # ── Parse meals parameter ────────────────────────────────────────────
        meals_str = options.get("meals", "breakfast,lunch,olovrant")
        meals = [m.strip() for m in meals_str.split(",") if m.strip()]

        # ── Generate report ───────────────────────────────────────────────────
        try:
            report_bytes = build_xlsx_bytes(target_date, meals=meals)
        except Exception:
            logger.exception("Failed to generate XLSX report for %s", target_date)
            self.stderr.write(
                self.style.ERROR(f"Failed to generate report for {target_date}.")
            )
            raise

        filename = f"prehlad_{target_date.isoformat()}.xlsx"

        # ── Send email ────────────────────────────────────────────────────────
        try:
            send_daily_report_email(
                recipients=recipients,
                report_date=target_date.isoformat(),
                attachment_bytes=report_bytes,
                attachment_filename=filename,
                meals=meals,
            )
        except Exception:
            self.stderr.write(self.style.ERROR("Failed to send daily report email."))
            raise

        self.stdout.write(
            self.style.SUCCESS(
                f"Report for {target_date} sent to: {', '.join(recipients)}"
            )
        )
