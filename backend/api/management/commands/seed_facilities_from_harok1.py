"""Seed the facility roster (Celok + Prevadzka + login) from a real Hárok1 sheet.

Hárok1 is the only place the full customer list exists — it carries the name, the
delivery address, the delivery route and the operational notes for every facility the
kitchen cooks for. The app's DB only ever had the EduPage-scraped subset.

Facilities already known to the app keep their existing Celok (matched through the
``compare-real`` alias map, so the EduPage ones are recognised under their workbook
name) — the command only fills gaps, never renames or re-sources what is already there.
Everything else is created as an ``app`` celok, since a facility we do not scrape must
order in the app.

Logins are created with a **fictional** e-mail and password so the roster can be worked
with before real credentials exist. They are throwaway dev data: the password is random
per run and only printed once, via ``--credentials-out``.

Idempotent: re-running matches on the celok name and leaves existing rows alone.
"""

from __future__ import annotations

import glob
import json
import secrets
from pathlib import Path

from django.contrib.auth.models import User
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils.text import slugify
from openpyxl import load_workbook

from api.models import Celok, Prevadzka, UserProfile

from .reconcile_real import (
    REAL_DIR,
    _facility_header_rows,
    _is_address_row,
    _load_alias_map,
    _normalize,
)

# Fictional logins only — RFC 2606 reserves example.com precisely so seeded addresses
# can never reach a real inbox if this data ever leaks into a mailer.
_EMAIL_DOMAIN = "example.com"

ALIAS_MAP = (
    Path(__file__).resolve().parents[4]
    / ".claude"
    / "skills"
    / "compare-real"
    / "facility_aliases.json"
)


def _facility_rows(sheet) -> list[dict]:
    """The roster as it appears in Hárok1, in delivery order.

    A route header ("1.Trasa - Pezinská - Heňo/Ivan") is a bare text row between
    blocks; it applies to every facility below it until the next one.
    """
    rows = list(sheet.iter_rows(values_only=True))
    headers = set(_facility_header_rows(sheet))
    roster: list[dict] = []
    route = ""
    for index in range(2, sheet.max_row + 1):
        label = rows[index - 1][0]
        if isinstance(label, str) and "trasa" in label.lower() and index not in headers:
            route = label.strip()
        if index not in headers:
            continue
        address = rows[index][0] if _is_address_row(sheet, index + 1) else ""
        note = rows[index - 1][7] if len(rows[index - 1]) > 7 else None
        note = str(note or "").strip()
        roster.append(
            {
                "name": str(label).strip(),
                # Some blocks put a diet label where the address goes; the address is
                # optional metadata here, so a wrong guess is not worth a heuristic.
                "address": str(address or "").strip(),
                "route": route,
                # "*" is a bare marker in the notes column, not a note.
                "note": "" if note == "*" else note,
            }
        )
    return roster


def _duplicate_roster_names(roster: list[dict]) -> set[str]:
    counts: dict[str, int] = {}
    for entry in roster:
        key = _normalize(entry["name"])
        if key:
            counts[key] = counts.get(key, 0) + 1
    return {key for key, count in counts.items() if count > 1}


def _disambiguated_name(entry: dict, duplicate_names: set[str]) -> str:
    """Unique internal label for facilities that share the same workbook name."""
    name = entry["name"]
    if _normalize(name) not in duplicate_names:
        return name
    suffix = entry.get("address") or entry.get("route") or ""
    return f"{name} ({suffix})" if suffix else name


def _facility_identity_key(name: str, address: str) -> str:
    """Stable key for a workbook facility occurrence: name + delivery address."""
    if address:
        return f"{_normalize(name)}\n{_normalize(address)}"
    return _normalize(name)


def _index_existing_facilities():
    by_name: dict[str, list[Celok]] = {}
    by_identity: dict[str, Celok] = {}
    for celok in Celok.objects.prefetch_related("prevadzky").all():
        name_key = _normalize(celok.nazov)
        by_name.setdefault(name_key, []).append(celok)
        if celok.adresa:
            by_identity.setdefault(
                _facility_identity_key(celok.nazov, celok.adresa), celok
            )
        for prevadzka in celok.prevadzky.all():
            prev_name_key = _normalize(prevadzka.nazov)
            if prev_name_key != name_key:
                by_name.setdefault(prev_name_key, []).append(celok)
            if prevadzka.adresa:
                by_identity.setdefault(
                    _facility_identity_key(prevadzka.nazov, prevadzka.adresa), celok
                )
    return by_name, by_identity


def _first_unique(candidates: list[Celok] | None) -> Celok | None:
    if not candidates:
        return None
    unique = {c.pk: c for c in candidates}
    if len(unique) == 1:
        return next(iter(unique.values()))
    return None


def _addresses_for(celok: Celok) -> set[str]:
    addresses = {_normalize(celok.adresa)}
    addresses.update(_normalize(p.adresa) for p in celok.prevadzky.all())
    return {a for a in addresses if a}


def _rename_single_facility_if_safe(celok: Celok, target: str, address: str) -> None:
    """Upgrade an old duplicate-name seed to its unambiguous 1:1 label."""
    if not target or celok.nazov == target:
        return
    prevadzky = list(celok.prevadzky.all())
    if len(prevadzky) != 1:
        return
    if Celok.objects.filter(nazov=target).exclude(pk=celok.pk).exists():
        return

    old_name = celok.nazov
    prevadzka = prevadzky[0]
    celok.nazov = target
    if address and not celok.adresa:
        celok.adresa = address
    celok.save(update_fields=["nazov", "adresa"])

    if prevadzka.nazov == old_name or _normalize(prevadzka.nazov) == _normalize(
        old_name
    ):
        prevadzka.nazov = target
    if address and not prevadzka.adresa:
        prevadzka.adresa = address
    prevadzka.save(update_fields=["nazov", "adresa"])

    for profile in celok.profily.all():
        selected = list(profile.prevadzky.all())
        if selected and selected != [prevadzka]:
            continue
        if profile.company_name in ("", old_name):
            profile.company_name = target
            profile.save(update_fields=["company_name"])


def _match_existing_duplicate(
    entry: dict,
    by_name: dict[str, list[Celok]],
    by_identity: dict[str, Celok],
    used: set[int],
) -> Celok | None:
    identity = _facility_identity_key(entry["name"], entry["address"])
    match = by_identity.get(identity)
    if match is not None:
        return match

    candidates = [
        c for c in by_name.get(_normalize(entry["name"]), []) if c.pk not in used
    ]
    if len(candidates) != 1:
        return None
    candidate = candidates[0]
    entry_address = _normalize(entry["address"])
    candidate_addresses = _addresses_for(candidate)
    if not candidate_addresses or entry_address in candidate_addresses:
        return candidate
    return None


def _real_label_to_app(alias_map: dict[str, list[str]]) -> dict[str, str]:
    """{normalized Hárok1 label → app label} — the alias map, inverted."""
    inverted: dict[str, str] = {}
    for app_key, real_labels in alias_map.items():
        for real in real_labels:
            inverted[real] = app_key
    return inverted


def _split_app_label(app_label: str) -> tuple[str, str]:
    """``"jolly homeschool – jolly 1"`` → ``("jolly homeschool", "jolly 1")``.

    Multi-prevádzka celky are written ``Celok – Prevádzka`` (en-dash), the same shape
    `api.utils.order_row_label` emits. Without the split, Jolly 1/2/3 and Školička
    les/lúka look like unknown facilities and would be seeded as three sibling celky
    instead of prevádzky under the one that already exists.
    """
    celok, separator, prevadzka = app_label.partition("–")
    if not separator:
        return app_label.strip(), ""
    return celok.strip(), prevadzka.strip()


def _email_for(name: str, taken: set[str]) -> str:
    base = slugify(name) or "prevadzka"
    email = f"{base}@{_EMAIL_DOMAIN}"
    suffix = 2
    while email in taken:
        email = f"{base}-{suffix}@{_EMAIL_DOMAIN}"
        suffix += 1
    taken.add(email)
    return email


class Command(BaseCommand):
    help = "Seed Celok/Prevadzka/logins from a real workbook's Hárok1 roster."

    def add_arguments(self, parser):
        parser.add_argument(
            "--workbook",
            help="Workbook path (default: the newest one in test/data/real).",
        )
        parser.add_argument(
            "--credentials-out",
            help="Write the generated logins to this file (passwords appear ONLY here).",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Report what would change without writing.",
        )

    def _resolve_workbook(self, override: str | None) -> Path:
        if override:
            path = Path(override)
            if not path.exists():
                raise CommandError(f"{path}: file does not exist")
            return path
        matches = sorted(
            glob.glob(str(REAL_DIR / "*.xlsx")), key=lambda p: Path(p).stat().st_mtime
        )
        if not matches:
            raise CommandError(f"No .xlsx workbook in {REAL_DIR}.")
        return Path(matches[-1])

    def handle(self, *args, **options):
        workbook_path = self._resolve_workbook(options.get("workbook"))
        wb = load_workbook(workbook_path, data_only=True)
        if "Hárok1" not in wb.sheetnames:
            raise CommandError(f"{workbook_path.name} has no Hárok1 sheet.")

        roster = _facility_rows(wb["Hárok1"])
        duplicate_names = _duplicate_roster_names(roster)
        alias_map = _load_alias_map(str(ALIAS_MAP)) if ALIAS_MAP.exists() else {}
        real_to_app = _real_label_to_app(alias_map)

        existing_by_name, existing_by_identity = _index_existing_facilities()
        used_existing: set[int] = set()
        taken_emails = set(User.objects.values_list("username", flat=True))

        created, linked, added_prevadzky = [], [], []
        credentials = []

        with transaction.atomic():
            for entry in roster:
                key = _normalize(entry["name"])
                operation_name = _disambiguated_name(entry, duplicate_names)
                app_label = real_to_app.get(key, "")
                celok_key, prevadzka_name = _split_app_label(app_label)
                # Already in the DB — under its own name, or under the app name the
                # alias map ties this workbook label to.
                if key in duplicate_names and not app_label:
                    match = _first_unique(
                        existing_by_name.get(_normalize(operation_name))
                    ) or _match_existing_duplicate(
                        entry, existing_by_name, existing_by_identity, used_existing
                    )
                else:
                    match = _first_unique(existing_by_name.get(key)) or (
                        _first_unique(existing_by_name.get(celok_key))
                        if celok_key
                        else None
                    )
                if match:
                    used_existing.add(match.pk)
                    # An aliased sub-unit (Jolly 1) may still be missing its prevádzka
                    # even though its celok exists.
                    if prevadzka_name and not options["dry_run"]:
                        _, made = Prevadzka.objects.get_or_create(
                            celok=match,
                            nazov=entry["name"],
                            defaults={"adresa": entry["address"]},
                        )
                        if made:
                            added_prevadzky.append((entry["name"], match.nazov))
                    if (
                        key in duplicate_names
                        and not app_label
                        and not options["dry_run"]
                    ):
                        _rename_single_facility_if_safe(
                            match, operation_name, entry["address"]
                        )
                    linked.append((entry["name"], match.nazov))
                    continue
                if options["dry_run"]:
                    created.append(operation_name)
                    continue

                celok = Celok.objects.create(
                    nazov=operation_name,
                    adresa=entry["address"],
                    zdroj_objednavok=Celok.ZdrojObjednavok.APP,
                )
                prevadzka = Prevadzka.objects.create(
                    celok=celok, nazov=operation_name, adresa=entry["address"]
                )
                email = _email_for(operation_name, taken_emails)
                password = secrets.token_urlsafe(9)
                user = User.objects.create_user(
                    username=email, email=email, password=password
                )
                UserProfile.objects.create(
                    user=user,
                    company_name=operation_name,
                    celok=celok,
                    is_edupage=False,
                )
                user.profile.prevadzky.add(prevadzka)
                existing_by_name.setdefault(_normalize(operation_name), []).append(
                    celok
                )
                existing_by_identity[
                    _facility_identity_key(entry["name"], entry["address"])
                ] = celok
                used_existing.add(celok.pk)
                created.append(operation_name)
                credentials.append(
                    {
                        **entry,
                        "operation_name": operation_name,
                        "email": email,
                        "password": password,
                    }
                )

            if options["dry_run"]:
                transaction.set_rollback(True)

        if options.get("credentials_out") and credentials:
            Path(options["credentials_out"]).write_text(
                json.dumps(credentials, ensure_ascii=False, indent=2), encoding="utf-8"
            )

        prefix = "[dry-run] " if options["dry_run"] else ""
        self.stdout.write(
            f"{prefix}{workbook_path.name}: {len(roster)} v Hárok1 → "
            f"{len(created)} nových celkov, {len(linked)} už existuje "
            f"({len(added_prevadzky)} nových prevádzok)"
        )
        for name, celok_name in linked:
            self.stdout.write(f"  = {name} → {celok_name}")
        for name, celok_name in added_prevadzky:
            self.stdout.write(f"  + prevádzka {name} pod {celok_name}")
