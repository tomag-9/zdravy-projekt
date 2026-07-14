"""Two-tier reconciliation of app output against a real workbook.

Tier 1 (counts / sizes / diets): app per-client portion counts vs the
``vyúčtovanie`` sheet ``Počet pokrmov`` column (ground truth for counts).
Tier 2 (gramage): app gramage totals vs the ``Hárok1`` sheet grams.

The real workbook for a date is auto-resolved from ``test/data/real`` using the
``D.M.YYYY_...xlsx`` naming convention. Emits a machine-readable JSON report on
stdout (consumed by the ``compare-real`` skill) plus a human summary on stderr.
"""

from __future__ import annotations

import glob
import json
import unicodedata
from datetime import date as date_cls
from decimal import Decimal
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from api.services.meal_plan_service import MealPlanService

REAL_DIR = Path(__file__).resolve().parents[4] / "test" / "data" / "real"


def _normalize(value: object) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or "").casefold())
    ascii_value = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return " ".join(ascii_value.replace(".", " ").replace("-", " ").split())


def _decimal(value: object) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value).replace(",", "."))
    except Exception:
        return Decimal("0")


def _count_or_none(value: object) -> Decimal | None:
    """Parse a `Počet pokrmov` cell → Decimal, or None if it isn't a number.

    The workbook sometimes stores a count as TEXT (e.g. `'1'`), so an
    ``isinstance(int, float)`` guard silently drops those rows and undercounts a
    facility's diet lines. Accept numeric strings too; return None only for real
    non-numbers (blank cells, section headers).
    """
    if isinstance(value, bool):  # bool is an int subclass — never a count
        return None
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    if isinstance(value, str):
        text = value.strip().replace(",", ".")
        if text:
            try:
                return Decimal(text)
            except Exception:
                return None
    return None


def _load_alias_map(path: str | None) -> dict[str, list[str]]:
    """Return normalized-app-label → list of normalized-real-label(s).

    A value may be a single real label (``"deutsche schule"``) or a list
    (``["rozmanita skolka", "rozmanita skola"]``) meaning "sum these real rows for
    this one app facility" — used when the app scrapes a celok as one bucket but the
    real workbook bills its sub-units on separate rows. Single strings become
    1-element lists so the resolution code has a uniform shape. Keys starting with
    ``_`` (e.g. ``_comment``) are skipped.
    """
    if not path:
        return {}
    with Path(path).open(encoding="utf-8") as handle:
        raw = json.load(handle)
    if not isinstance(raw, dict):
        raise CommandError(
            "Alias map must be a JSON object of app label -> real label(s)."
        )
    result: dict[str, list[str]] = {}
    for app, real in raw.items():
        if app.startswith("_"):
            continue
        if isinstance(real, str):
            reals = [real]
        elif isinstance(real, list):
            reals = [str(r) for r in real]
        else:
            raise CommandError(
                f"Alias for {app!r} must be a string or list of strings."
            )
        result[_normalize(app)] = [_normalize(r) for r in reals]
    return result


def _rekey_by_alias(
    real_by_label: dict[str, Any],
    alias_map: dict[str, list[str]],
    combine,
) -> dict[str, Any]:
    """Re-key real-side data from raw workbook labels onto app-normalized keys.

    For each aliased app facility, pull the referenced real rows out of the map and
    fold them together with ``combine`` under the app-normalized key. Non-aliased
    labels are left untouched (they match app labels by identity). This keeps the
    raw aliased rows from lingering as spurious ``real_only`` entries.
    """
    resolved = dict(real_by_label)
    for app_key, real_labels in alias_map.items():
        pieces = [resolved.pop(label) for label in real_labels if label in resolved]
        if pieces:
            resolved[app_key] = combine(pieces)
    return resolved


def _combine_count_buckets(pieces: list[dict]) -> dict:
    merged: dict[str, Decimal] = {}
    for bucket in pieces:
        for meal_type, value in bucket.items():
            merged[meal_type] = merged.get(meal_type, Decimal("0")) + value
    return merged


def _resolve_workbook(date_str: str) -> Path:
    y, m, d = (int(p) for p in date_str.split("-"))
    date_cls(y, m, d)  # validate
    # Filenames use non-zero-padded D.M.YYYY, e.g. "9.7.2026_...".
    pattern = f"{d}.{m}.{y}_*.xlsx"
    matches = sorted(glob.glob(str(REAL_DIR / pattern)))
    if not matches:
        raise CommandError(
            f"No real workbook for {date_str} in {REAL_DIR} (looked for {pattern})."
        )
    return Path(matches[0])


# ── Tier 1: counts from the vyúčtovanie sheet ──────────────────────────────────
# Meal-type buckets are compared like-for-like: a facility that only bills OBED
# must not be faulted against the app's lunch+olovrant grand total. Order matters
# below — "ranajky + obed" must hit "obed" (lunch), not "ranajky" (breakfast).
MEAL_TYPES = ("lunch", "snack", "breakfast")


def _section_meal_type(section: str) -> str | None:
    if "olovrant" in section:
        return "snack"
    if "obed" in section:
        return "lunch"
    if "desiata" in section or "ranajky" in section:
        return "breakfast"
    return None


def _real_counts_by_facility(wb) -> dict[str, dict[str, Decimal]]:
    """{facility → {meal_type → Σ Počet pokrmov}} from the vyúčtovanie sheet.

    A row with a ``Druh pokrmu`` (col 2) but no ``Počet`` (col 4) is a section
    header (OBED / OLOVRANT / …); the rows beneath it carry the counts.
    """
    if "vyúčtovanie" not in wb.sheetnames:
        return {}
    sheet = wb["vyúčtovanie"]
    counts: dict[str, dict[str, Decimal]] = {}
    facility: str | None = None
    meal_type: str | None = None
    for row in range(2, sheet.max_row + 1):
        raw_facility = sheet.cell(row=row, column=1).value
        if raw_facility not in (None, ""):
            facility = _normalize(raw_facility)
            counts.setdefault(facility, {})
        druh = sheet.cell(row=row, column=2).value
        count = sheet.cell(row=row, column=4).value
        if druh not in (None, "") and count in (None, ""):
            meal_type = _section_meal_type(_normalize(druh))
        if facility is None or meal_type is None:
            continue
        parsed = _count_or_none(count)
        if parsed is not None:
            bucket = counts[facility]
            bucket[meal_type] = bucket.get(meal_type, Decimal("0")) + parsed
    # Drop facilities/meal-types that are entirely zero (not billed that day).
    return {
        fac: {mt: v for mt, v in buckets.items() if v > 0}
        for fac, buckets in counts.items()
        if any(v > 0 for v in buckets.values())
    }


# Maps app meal-plan categories to the vyúčtovanie meal-type buckets. Soup and
# main_course share one headcount (a "lunch"), so only main_course is counted.
_APP_MEAL_TO_TYPE = {
    "main_course": "lunch",
    "afternoon_snack": "snack",
    "breakfast_snack": "breakfast",
}


def _app_counts_by_meal_type(row: dict[str, Any]) -> dict[str, Decimal]:
    buckets: dict[str, Decimal] = {}
    for sub in row.get("sub_rows", []):
        meal_type = _APP_MEAL_TO_TYPE.get(sub.get("meal", ""))
        if meal_type is None:
            continue
        buckets[meal_type] = buckets.get(meal_type, Decimal("0")) + _decimal(
            sub.get("count", 0)
        )
    return buckets


# ── Tier 2: gramage from the Hárok1 sheet ──────────────────────────────────────
def _component_labels(col_groups: list[dict]) -> list[str]:
    labels: list[str] = []
    for group in col_groups:
        for component in group["components"]:
            labels.append(f"{group['label']} / {component['label']}")
    return labels


def _component_names(col_groups: list[dict]) -> list[str]:
    """Bare component (dish) names, parallel to `_component_labels`.

    These are the actual dish names (`Zemiakový prívarok`, `Vajce`, …) and match the
    Hárok1 header cells, which is how we align columns by NAME instead of position.
    """
    return [c["label"] for g in col_groups for c in g["components"]]


def _empty_totals(col_groups: list[dict]) -> list[list[Decimal]]:
    return [[Decimal("0")] * len(g["components"]) for g in col_groups]


def _merge_totals(target: list[list[Decimal]], source: list) -> None:
    for gi, group_values in enumerate(source or []):
        if gi >= len(target):
            continue
        for ci, value in enumerate(group_values or []):
            if ci >= len(target[gi]) or value in (None, ""):
                continue
            target[gi][ci] += _decimal(value)


def _app_gram_values(row: dict[str, Any], col_groups: list[dict]) -> list[Decimal]:
    totals = _empty_totals(col_groups)
    _merge_totals(totals, row.get("standard_col_grams", []))
    for diet_row in row.get("diet_summary_rows", []):
        _merge_totals(totals, diet_row.get("col_grams", []))
    return [v for group in totals for v in group]


def _real_rows_by_label(sheet) -> dict[str, list[int]]:
    rows: dict[str, list[int]] = {}
    for row_index in range(1, sheet.max_row + 1):
        normalized = _normalize(sheet.cell(row=row_index, column=1).value)
        if normalized:
            rows.setdefault(normalized, []).append(row_index)
    return rows


def _real_header_columns(sheet) -> dict[str, int]:
    """{normalized dish name → column index} from the Hárok1 header row (row 1).

    The header lists the day's actual dishes per column (they differ every day and
    are NOT in the same order as the app's components, plus there can be a blank
    spacer column). Mapping by name — not by position — is what makes Tier-2 align.
    """
    columns: dict[str, int] = {}
    for col in range(2, sheet.max_column + 1):
        name = _normalize(sheet.cell(row=1, column=col).value)
        if name:
            columns.setdefault(name, col)
    return columns


def _is_gram_row(sheet, row: int) -> bool:
    """A data row carrying grams: a text label in col A and a number in col B."""
    a = sheet.cell(row=row, column=1).value
    b = sheet.cell(row=row, column=2).value
    return isinstance(a, str) and a.strip() != "" and isinstance(b, (int, float))


def _is_address_row(sheet, row: int) -> bool:
    """Address line: non-numeric text label with an empty grams column.

    Distinct from a count line (`12` or the TEXT '1'), whose col-A value parses as a
    number — the ``_count_or_none`` guard keeps those from looking like addresses.
    """
    a = sheet.cell(row=row, column=1).value
    b = sheet.cell(row=row, column=2).value
    return (
        isinstance(a, str)
        and a.strip() != ""
        and b in (None, "")
        and _count_or_none(a) is None
    )


def _starts_new_facility(sheet, row: int) -> bool:
    """A facility header is any row immediately followed by an address line.

    This holds even when the header itself has empty grams (e.g. `Rozmanita Škola`),
    which a gram-row test would miss — that miss let one facility's block bleed into
    the next.
    """
    return _is_address_row(sheet, row + 1)


def _expand_block_rows(sheet, header_rows: list[int]) -> list[int]:
    """Expand each facility header row to its full block (klasik + diet sub-rows).

    In Hárok1 a facility is a block: the header row (KLASIK grams), an address line,
    then alternating diet gram-rows / count-lines, until the next facility header.
    The app aggregates klasik + diets, so Tier-2 must sum the whole block.
    """
    max_row = sheet.max_row
    result: list[int] = []
    for header in header_rows:
        result.append(header)
        row = header + 1
        while row <= max_row:
            if _starts_new_facility(sheet, row):
                break
            if _is_gram_row(sheet, row):
                result.append(row)
            row += 1
    return result


def _real_gram_values_by_name(
    sheet,
    row_numbers: list[int],
    component_names: list[str],
    header_columns: dict[str, int],
) -> list[Decimal | None]:
    """Real grams per app component, matched to the Hárok1 column by dish name.

    Returns ``None`` for a component whose dish isn't in this day's header (e.g. a
    diet-only facility whose menu differs from the standard) so the caller can skip
    it rather than report a bogus diff against 0.
    """
    values: list[Decimal | None] = []
    for name in component_names:
        col = header_columns.get(_normalize(name))
        if col is None:
            values.append(None)
            continue
        total = Decimal("0")
        for row_number in row_numbers:
            total += _decimal(sheet.cell(row=row_number, column=col).value)
        values.append(total)
    return values


class Command(BaseCommand):
    help = "Reconcile app output against a real workbook (counts + gramage tiers)."

    def add_arguments(self, parser):
        parser.add_argument("--date", required=True, help="Date YYYY-MM-DD.")
        parser.add_argument(
            "--workbook", help="Override workbook path (default: auto-resolve by date)."
        )
        parser.add_argument(
            "--alias-map",
            help="JSON {app label: real facility label} for name-mapping gaps.",
        )
        parser.add_argument(
            "--count-tolerance",
            default="0",
            help="Allowed absolute count difference before FAIL.",
        )
        parser.add_argument(
            "--gram-tolerance",
            default="0.01",
            help="Allowed absolute gram difference before FAIL.",
        )

    def handle(self, *args, **options):
        date_str = options["date"]
        workbook_path = (
            Path(options["workbook"])
            if options.get("workbook")
            else _resolve_workbook(date_str)
        )
        if not workbook_path.exists():
            raise CommandError(f"{workbook_path}: file does not exist")

        count_tol = _decimal(options["count_tolerance"])
        gram_tol = _decimal(options["gram_tolerance"])
        alias_map = _load_alias_map(options.get("alias_map"))

        # Alias resolution happens on the real side (rows are re-keyed onto the
        # app-normalized label, summing multi-row aliases), so the app key is just
        # its normalized label.
        def _key(client: object) -> str:
            return _normalize(client)

        dashboard = MealPlanService.gramage_dashboard(date_str)
        app_rows = dashboard.get("rows", [])
        col_groups = dashboard.get("col_groups", [])
        labels = _component_labels(col_groups)
        width = len(labels)

        # read_only=False: this command does random cell access per facility,
        # which is pathologically slow on a read_only worksheet.
        wb = load_workbook(workbook_path, data_only=True)

        # ── Tier 1: counts (per meal type) ────────────────────────────────────
        real_counts = _rekey_by_alias(
            _real_counts_by_facility(wb), alias_map, _combine_count_buckets
        )
        app_counts = {
            _key(r.get("client", "")): (r, _app_counts_by_meal_type(r))
            for r in app_rows
        }
        matched_keys = set(app_counts) & set(real_counts)

        count_findings = []
        for key in sorted(matched_keys):
            row, app_buckets = app_counts[key]
            real_buckets = real_counts[key]
            client = str(row.get("client", ""))
            # Only compare meal types that either side actually billed/ordered.
            for meal_type in MEAL_TYPES:
                app_val = app_buckets.get(meal_type)
                real_val = real_buckets.get(meal_type)
                if app_val is None and real_val is None:
                    continue
                app_val = app_val or Decimal("0")
                real_val = real_val or Decimal("0")
                diff = app_val - real_val
                count_findings.append(
                    {
                        "facility": client,
                        "meal_type": meal_type,
                        "app": str(app_val),
                        "real": str(real_val),
                        "diff": str(diff),
                        "status": "OK" if abs(diff) <= count_tol else "FAIL",
                    }
                )
        app_only = sorted(
            str(app_counts[k][0].get("client", ""))
            for k in set(app_counts) - matched_keys
        )
        real_only = sorted(set(real_counts) - matched_keys)

        # ── Tier 2: gramage ───────────────────────────────────────────────────
        gram_findings = []
        if width:
            har_sheet = wb["Hárok1"] if "Hárok1" in wb.sheetnames else wb.active
            header_columns = _real_header_columns(har_sheet)
            component_names = _component_names(col_groups)
            real_gram_rows = _rekey_by_alias(
                _real_rows_by_label(har_sheet),
                alias_map,
                lambda pieces: [row for rows in pieces for row in rows],
            )
            for row in app_rows:
                client = str(row.get("client", ""))
                key = _key(client)
                row_numbers = real_gram_rows.get(key, [])
                app_values = _app_gram_values(row, col_groups)
                if not row_numbers:
                    gram_findings.append(
                        {"facility": client, "status": "MISSING_REAL_ROW"}
                    )
                    continue
                block_rows = _expand_block_rows(har_sheet, row_numbers)
                real_values = _real_gram_values_by_name(
                    har_sheet, block_rows, component_names, header_columns
                )
                for label, real_v, app_v in zip(labels, real_values, app_values):
                    # Dish not in this day's header → can't compare, don't fake a diff.
                    if real_v is None:
                        continue
                    diff = app_v - real_v
                    if abs(diff) <= gram_tol:
                        continue
                    gram_findings.append(
                        {
                            "facility": client,
                            "component": label,
                            "app": str(app_v),
                            "real": str(real_v),
                            "diff": str(diff),
                            "status": "FAIL",
                        }
                    )

        report = {
            "date": date_str,
            "workbook": workbook_path.name,
            "meal_plan_id": dashboard.get("meal_plan_id"),
            "tier1_counts": {
                "matched": count_findings,
                "app_only": app_only,
                "real_only": real_only,
            },
            "tier2_gramage": {"diffs": gram_findings},
            "summary": {
                "count_fail": sum(1 for f in count_findings if f["status"] == "FAIL"),
                "count_ok": sum(1 for f in count_findings if f["status"] == "OK"),
                "unmatched_facilities": len(app_only) + len(real_only),
                "gram_fail": sum(1 for f in gram_findings if f.get("status") == "FAIL"),
                "gram_missing_rows": sum(
                    1 for f in gram_findings if f.get("status") == "MISSING_REAL_ROW"
                ),
            },
        }
        self.stdout.write(json.dumps(report, ensure_ascii=False, indent=2))

        s = report["summary"]
        self.stderr.write(
            f"[{date_str}] {workbook_path.name}: "
            f"counts OK={s['count_ok']} FAIL={s['count_fail']}  "
            f"gram FAIL={s['gram_fail']} missing={s['gram_missing_rows']}  "
            f"unmatched={s['unmatched_facilities']}"
        )
