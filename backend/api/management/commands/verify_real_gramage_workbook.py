"""Compare app gramage-dashboard totals with a real XLSX workbook."""

from __future__ import annotations

import csv
import json
import unicodedata
from decimal import Decimal
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from api.services.meal_plan_service import MealPlanService


def _normalize(value: object) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or "").casefold())
    ascii_value = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return " ".join(ascii_value.replace(".", " ").replace("-", " ").split())


def _decimal(value: object) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    return Decimal(str(value).replace(",", "."))


def _empty_totals(col_groups: list[dict]) -> list[list[Decimal]]:
    return [[Decimal("0")] * len(group["components"]) for group in col_groups]


def _merge_totals(target: list[list[Decimal]], source: list) -> None:
    for group_index, group_values in enumerate(source or []):
        if group_index >= len(target):
            continue
        for component_index, value in enumerate(group_values or []):
            if component_index >= len(target[group_index]) or value in (None, ""):
                continue
            target[group_index][component_index] += _decimal(value)


def _flatten(values: list[list[Decimal]]) -> list[Decimal]:
    return [value for group in values for value in group]


def _component_labels(col_groups: list[dict]) -> list[str]:
    labels: list[str] = []
    for group in col_groups:
        for component in group["components"]:
            labels.append(f"{group['label']} / {component['label']}")
    return labels


def _load_row_mapping(path: str | None) -> dict[str, list[int]]:
    if not path:
        return {}
    with Path(path).open(encoding="utf-8") as handle:
        raw = json.load(handle)
    if not isinstance(raw, dict):
        raise CommandError("Row map must be a JSON object of client -> row numbers.")
    mapping: dict[str, list[int]] = {}
    for client, rows in raw.items():
        if isinstance(rows, int):
            mapping[str(client)] = [rows]
        elif isinstance(rows, list):
            mapping[str(client)] = [int(row) for row in rows]
        else:
            raise CommandError(f"Invalid row map entry for {client!r}: {rows!r}")
    return mapping


def _real_rows_by_label(sheet) -> dict[str, list[int]]:
    rows: dict[str, list[int]] = {}
    for row_index in range(1, sheet.max_row + 1):
        label = sheet.cell(row=row_index, column=1).value
        normalized = _normalize(label)
        if not normalized:
            continue
        rows.setdefault(normalized, []).append(row_index)
    return rows


def _real_values(sheet, row_numbers: list[int], width: int) -> list[Decimal]:
    values = [Decimal("0")] * width
    for row_number in row_numbers:
        for offset in range(width):
            values[offset] += _decimal(
                sheet.cell(row=row_number, column=offset + 2).value
            )
    return values


def _app_values(row: dict[str, Any], col_groups: list[dict]) -> list[Decimal]:
    totals = _empty_totals(col_groups)
    _merge_totals(totals, row.get("standard_col_grams", []))
    for diet_row in row.get("diet_summary_rows", []):
        _merge_totals(totals, diet_row.get("col_grams", []))
    return _flatten(totals)


class Command(BaseCommand):
    help = "Compare app gramage-dashboard totals with a real XLSX workbook."

    def add_arguments(self, parser):
        parser.add_argument("--date", required=True, help="Date in YYYY-MM-DD format.")
        parser.add_argument(
            "--workbook", required=True, help="Real XLSX workbook path."
        )
        parser.add_argument(
            "--row-map",
            help="Optional JSON mapping of app client label to 1-based workbook row number(s).",
        )
        parser.add_argument(
            "--output-csv",
            help="Optional output CSV path. Defaults to stdout.",
        )
        parser.add_argument(
            "--tolerance",
            default="0.01",
            help="Allowed absolute difference before a cell is marked DIFF.",
        )

    def handle(self, *args, **options):
        workbook_path = Path(options["workbook"])
        if not workbook_path.exists():
            raise CommandError(f"{workbook_path}: file does not exist")

        dashboard = MealPlanService.gramage_dashboard(options["date"])
        col_groups = dashboard.get("col_groups", [])
        labels = _component_labels(col_groups)
        width = len(labels)
        if width == 0:
            raise CommandError(f"No app gramage columns for {options['date']}")

        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        sheet = workbook.active
        real_rows = _real_rows_by_label(sheet)
        explicit_mapping = _load_row_mapping(options.get("row_map"))
        tolerance = Decimal(str(options["tolerance"]))

        rows: list[dict[str, object]] = []
        has_diff = False
        for app_row in dashboard.get("rows", []):
            client = str(app_row.get("client", ""))
            row_numbers = explicit_mapping.get(client)
            if row_numbers is None:
                row_numbers = real_rows.get(_normalize(client), [])

            app_values = _app_values(app_row, col_groups)
            if not row_numbers:
                has_diff = True
                rows.append(
                    {
                        "date": options["date"],
                        "client": client,
                        "status": "MISSING_REAL_ROW",
                        "component": "",
                        "real": "",
                        "app": "",
                        "diff": "",
                    }
                )
                continue

            real_values = _real_values(sheet, row_numbers, width)
            for label, real_value, app_value in zip(labels, real_values, app_values):
                diff = app_value - real_value
                status = "OK" if abs(diff) <= tolerance else "DIFF"
                if status != "OK":
                    has_diff = True
                rows.append(
                    {
                        "date": options["date"],
                        "client": client,
                        "status": status,
                        "component": label,
                        "real": str(real_value),
                        "app": str(app_value),
                        "diff": str(diff),
                    }
                )

        fieldnames = ["date", "client", "status", "component", "real", "app", "diff"]
        if options.get("output_csv"):
            with Path(options["output_csv"]).open(
                "w", newline="", encoding="utf-8"
            ) as handle:
                writer = csv.DictWriter(handle, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(rows)
        else:
            writer = csv.DictWriter(self.stdout, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)

        if has_diff:
            raise CommandError("Real workbook verification found differences.")

        self.stdout.write(self.style.SUCCESS("Real workbook verification passed."))
