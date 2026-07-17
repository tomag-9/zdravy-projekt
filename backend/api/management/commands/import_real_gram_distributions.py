"""Import real daily gram distributions from ZP XLSX workbooks.

The real workbook is the accounting-facing source of truth. This command keeps
the import repeatable: it reads the date and the `KLASIK` row, then creates or
updates the soup, main-course, and afternoon-snack templates for that day.
"""

from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from api.models import DailyMealPlan, MealCategory, MealPlanItem, MealTemplate
from api.serializers_menu import (
    _base_weight_grams_from_components,
    _weight_label_from_components,
)

_DATE_RE = re.compile(r"(?P<day>\d{1,2})\.(?P<month>\d{1,2})\.(?P<year>\d{4})")
_BREAD_WORDS = ("pecivo", "pečivo", "chlieb", "zemla", "žemľa", "rozok", "rožok")
_PIECE_RATIOS = {
    "Jasle": Decimal("1"),
    "Škôlka": Decimal("1"),
    "Predškolák": Decimal("1.5"),
    "ZŠ 1.stupeň": Decimal("1.5"),
    "ZŠ 2.stupeň": Decimal("1.5"),
    "Dospelý (SŠ)": Decimal("2"),
}


def _decimal(value: object) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value).replace(",", "."))
    except (InvalidOperation, ValueError) as exc:
        raise CommandError(f"Invalid numeric value in workbook: {value!r}") from exc


def _decimal_str(value: Decimal) -> str:
    normalized = value.normalize()
    return format(normalized, "f")


def _component(label: object, grams: Decimal, unit: str) -> dict:
    return {
        "label": str(label or "Zložka").strip() or "Zložka",
        "grams": _decimal_str(grams),
        "unit": unit,
    }


def _date_from_filename(path: Path) -> date | None:
    match = _DATE_RE.search(path.name)
    if not match:
        return None
    return date(
        int(match.group("year")),
        int(match.group("month")),
        int(match.group("day")),
    )


def _date_from_value(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        match = _DATE_RE.search(value)
        if match:
            return date(
                int(match.group("year")),
                int(match.group("month")),
                int(match.group("day")),
            )
    return None


def _iter_input_files(paths: list[str], input_dir: str | None) -> list[Path]:
    files: list[Path] = []
    if input_dir:
        files.extend(sorted(Path(input_dir).glob("*.xlsx")))
    files.extend(Path(p) for p in paths)
    return [p for p in files if p.name and not p.name.startswith("~$")]


def _find_klasik_row(rows: Iterable[tuple]) -> tuple | None:
    for row in rows:
        label = str(row[0] or "").strip().casefold()
        if label == "klasik":
            return row
    return None


def _is_piece_component(label: object, value: Decimal) -> bool:
    if value <= 0 or value > 5:
        return False
    normalized_label = str(label or "").strip().casefold()
    return any(word in normalized_label for word in _BREAD_WORDS)


def _piece_exception(label: object, per_kindergarten: Decimal) -> dict:
    return {
        "component_label": str(label or "Pečivo").strip() or "Pečivo",
        "unit": "ks",
        "counts_by_portion_type": {
            name: _decimal_str(per_kindergarten * ratio)
            for name, ratio in _PIECE_RATIOS.items()
        },
    }


def _template_defaults(components: list[dict], unit_exception: dict | None) -> dict:
    return {
        "components": components,
        "unit_exception": unit_exception,
        "weight_label": _weight_label_from_components(components, unit_exception),
        "base_weight_grams": _base_weight_grams_from_components(components),
        "is_active": True,
    }


def _upsert_template(
    *,
    category: str,
    name: str,
    components: list[dict],
    unit_exception: dict | None = None,
) -> MealTemplate:
    template, _ = MealTemplate.objects.update_or_create(
        name=name,
        defaults={
            "category": category,
            "menu_variant": "",
            "diet": None,
            **_template_defaults(components, unit_exception),
        },
    )
    return template


def import_workbook(path: Path) -> date:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        raise CommandError(f"{path}: workbook does not contain enough rows")

    target_date = _date_from_value(rows[0][0]) or _date_from_filename(path)
    if not target_date:
        raise CommandError(f"{path}: could not determine date")

    headers = rows[0]
    klasik = _find_klasik_row(rows[1:])
    if not klasik:
        raise CommandError(f"{path}: KLASIK row not found")

    soup_value = _decimal(klasik[1] if len(klasik) > 1 else 0)
    main_values = [
        _decimal(klasik[index] if len(klasik) > index else 0) for index in (2, 3, 4)
    ]
    snack_values = [
        _decimal(klasik[index] if len(klasik) > index else 0) for index in (5, 6)
    ]

    templates: list[MealTemplate] = []
    if soup_value > 0:
        templates.append(
            _upsert_template(
                category=MealCategory.SOUP,
                name=f"Real {target_date.isoformat()} Polievka",
                components=[_component(headers[1], soup_value, "ml")],
            )
        )

    main_components = [
        _component(headers[index], value, "g")
        for index, value in zip((2, 3, 4), main_values)
        if value > 0
    ]
    if main_components:
        templates.append(
            _upsert_template(
                category=MealCategory.MAIN_COURSE,
                name=f"Real {target_date.isoformat()} Hlavný chod",
                components=main_components,
            )
        )

    snack_components: list[dict] = []
    snack_exception = None
    for index, value in zip((5, 6), snack_values):
        if value <= 0:
            continue
        label = headers[index] if len(headers) > index else "Olovrant"
        if _is_piece_component(label, value):
            snack_exception = _piece_exception(label, value)
        else:
            snack_components.append(_component(label, value, "g"))
    if snack_components or snack_exception:
        templates.append(
            _upsert_template(
                category=MealCategory.AFTERNOON_SNACK,
                name=f"Real {target_date.isoformat()} Olovrant",
                components=snack_components,
                unit_exception=snack_exception,
            )
        )

    plan, _ = DailyMealPlan.objects.get_or_create(date=target_date)
    categories = [template.category for template in templates]
    MealPlanItem.objects.filter(meal_plan=plan, category__in=categories).delete()
    for template in templates:
        MealPlanItem.objects.create(
            meal_plan=plan,
            template=template,
            category=template.category,
            menu_variant="",
            diet=None,
        )

    return target_date


class Command(BaseCommand):
    help = "Import real daily gram distributions from ZP XLSX workbooks."

    def add_arguments(self, parser):
        parser.add_argument("paths", nargs="*", help="Workbook path(s) to import.")
        parser.add_argument(
            "--input-dir",
            help="Directory containing real XLSX workbooks. All *.xlsx files are imported.",
        )

    def handle(self, *args, **options):
        files = _iter_input_files(options["paths"], options.get("input_dir"))
        if not files:
            raise CommandError("Provide at least one XLSX path or --input-dir.")

        imported: list[date] = []
        for path in files:
            if not path.exists():
                raise CommandError(f"{path}: file does not exist")
            imported.append(import_workbook(path))

        self.stdout.write(
            self.style.SUCCESS(
                "Imported real gram distributions for "
                + ", ".join(d.isoformat() for d in imported)
            )
        )
