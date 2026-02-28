"""XLSX Report Generation."""

import datetime
import io

from django.http import HttpResponse
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework import status
from rest_framework.response import Response

from ..models import DailyOrder
from .report_xlsx_helpers import (
    xlsx_build_column_meta,
    xlsx_collect_columns,
    xlsx_style_headers,
    xlsx_write_data,
)


def generate_xlsx_report(request):
    """Generate XLSX report for daily orders."""
    import openpyxl

    date_str = request.query_params.get("date")
    if not date_str:
        return Response(
            {"error": "date parameter required"}, status=status.HTTP_400_BAD_REQUEST
        )
    try:
        target_date = datetime.date.fromisoformat(date_str)
    except (TypeError, ValueError):
        return Response(
            {"error": "invalid date format, expected YYYY-MM-DD"},
            status=status.HTTP_400_BAD_REQUEST,
        )
    safe_date = target_date.isoformat()

    orders = (
        DailyOrder.objects.filter(date=target_date)
        .select_related("user", "user__settings")
        .order_by("user__email")
    )

    meal_keys = ["breakfast", "lunch", "olovrant"]
    meal_labels = {"breakfast": "Raňajky", "lunch": "Obed", "olovrant": "Olovrant"}

    rows_data = [
        {
            "user": o.user,
            "data": o.data or {},
            "visible_meals": (
                getattr(getattr(o.user, "settings", None), "visible_meals", None)
                or ["breakfast", "lunch", "olovrant"]
            ),
        }
        for o in orders
    ]
    sorted_cats = xlsx_collect_columns(rows_data, meal_keys)
    col_meta, header_row_1, header_row_2, header_row_3 = xlsx_build_column_meta(
        sorted_cats, meal_keys, meal_labels
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Prehlad {safe_date}"

    header_fill_main = PatternFill("solid", fgColor="2563EB")
    header_fill_meal = {
        "breakfast": PatternFill("solid", fgColor="F97316"),
        "lunch": PatternFill("solid", fgColor="3B82F6"),
        "olovrant": PatternFill("solid", fgColor="22C55E"),
    }
    header_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    ws.append([f"Denný prehľad objednávok — {safe_date}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(header_row_1)
    ws.append(header_row_2)
    ws.append(header_row_3)

    xlsx_style_headers(
        ws,
        col_meta,
        sorted_cats,
        meal_keys,
        header_fill_main,
        header_fill_meal,
        header_font,
        center,
    )
    xlsx_write_data(ws, rows_data, meal_keys, sorted_cats, bold_font)

    ws.column_dimensions["A"].width = 28
    for c_idx in range(2, len(col_meta) + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 12
    ws.freeze_panes = "B6"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"prehlad_{safe_date}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
