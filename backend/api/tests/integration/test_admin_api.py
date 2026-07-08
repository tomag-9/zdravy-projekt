import io
from datetime import date

import openpyxl
import pytest
from django.contrib.auth.models import User
from rest_framework import status
from rest_framework.test import APITestCase

from api.models import (
    ClientSettings,
    DailyMealPlan,
    DailyOrder,
    Diet,
    EnrolledCount,
    MealPlanItem,
    MealTemplate,
    PortionType,
)

pytestmark = pytest.mark.integration


class AdminSummaryTest(APITestCase):
    def setUp(self):
        # Create Admin
        self.admin = User.objects.create_user(
            username="admin@example.com",
            password="password",
            email="admin@example.com",
            is_staff=True,
        )
        # Create Client
        self.client_user = User.objects.create_user(
            username="client@example.com",
            password="password",
            email="client@example.com",
            is_staff=False,
        )

        self.today = date.today().isoformat()

        # Create complex Detailed Order 1
        DailyOrder.objects.create(
            user=self.client_user,
            date=self.today,
            status="submitted",
            data={
                "breakfast": {
                    "Dospelý (SŠ)": {"menuCounts": {"A": 2}, "diets": {"Bez lepku": 1}}
                },
                "lunch": {
                    "ZŠ 1.stupeň": {"menuCounts": {"A": 5, "B": 2}, "diets": {}},
                    "Dospelý (SŠ)": {
                        "menuCounts": {"A": 1},
                        "diets": {"Bez laktózy": 1},
                    },
                },
                "olovrant": {"Jasle": {"menuCounts": {"A": 3}, "diets": {}}},
            },
        )

        # Create second client/order
        self.client2 = User.objects.create_user(
            username="client2@example.com",
            password="password",
            email="client2@example.com",
        )
        DailyOrder.objects.create(
            user=self.client2,
            date=self.today,
            status="draft",
            data={"lunch": {"ZŠ 1.stupeň": {"menuCounts": {"A": 3}, "diets": {}}}},
        )

    def test_admin_can_access_summary(self):
        self.client.force_authenticate(user=self.admin)
        response = self.client.get(f"/api/admin/summary/daily-stats/?date={self.today}")
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        data = response.json()
        self.assertEqual(data["total_orders"], 2)

        # Check Breakfast - Dospelý
        # Expecting: 2xA, 1x Gluten
        bf_stats = data["meals"]["breakfast"]["Dospelý (SŠ)"]
        self.assertEqual(bf_stats["menus"]["A"], 2)
        self.assertEqual(bf_stats["diets"]["Bez lepku"], 1)
        self.assertEqual(bf_stats["total"], 2)

        # Check Lunch - ZŠ 1.stupeň
        # Order 1: 5xA, 2xB. Order 2: 3xA. Total: 8xA, 2xB
        lunch_zs = data["meals"]["lunch"]["ZŠ 1.stupeň"]
        self.assertEqual(lunch_zs["menus"]["A"], 8)
        self.assertEqual(lunch_zs["menus"]["B"], 2)
        self.assertEqual(lunch_zs["total"], 10)

        # Check Lunch - Dospelý
        lunch_adult = data["meals"]["lunch"]["Dospelý (SŠ)"]
        self.assertEqual(lunch_adult["menus"]["A"], 1)
        self.assertEqual(lunch_adult["diets"]["Bez laktózy"], 1)

        # Check Olovrant
        olovrant_jasle = data["meals"]["olovrant"]["Jasle"]
        self.assertEqual(olovrant_jasle["menus"]["A"], 3)

    def test_client_cannot_access_summary(self):
        self.client.force_authenticate(user=self.client_user)
        response = self.client.get(f"/api/admin/summary/daily-stats/?date={self.today}")
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)


class ClientMealPlanAccessTest(APITestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            username="meal-admin@example.com",
            password="password",
            email="meal-admin@example.com",
            is_staff=True,
        )
        self.client_user = User.objects.create_user(
            username="meal-client@example.com",
            password="password",
            email="meal-client@example.com",
            is_staff=False,
        )
        self.plan_date = date(2026, 3, 16)
        self.template = MealTemplate.objects.create(
            category="main_course",
            name="Kuracie soté",
            weight_label="200g + 50g",
            base_weight_grams="250.00",
            is_active=True,
        )
        self.plan = DailyMealPlan.objects.create(
            date=self.plan_date,
            created_by=self.admin,
        )
        MealPlanItem.objects.create(
            meal_plan=self.plan,
            template=self.template,
            category="main_course",
        )

    def test_client_can_read_meal_plan_by_date(self):
        self.client.force_authenticate(user=self.client_user)

        response = self.client.get(
            f"/api/meal-plans/by-date/?date={self.plan_date.isoformat()}"
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = response.json()
        self.assertTrue(payload["exists"])
        self.assertEqual(payload["items"][0]["template_detail"]["name"], "Kuracie soté")

    def test_client_cannot_write_meal_plan(self):
        self.client.force_authenticate(user=self.client_user)

        response = self.client.post(
            "/api/meal-plans/",
            {"date": "2026-03-17"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)


class MealTemplateCatalogApiTest(APITestCase):
    def setUp(self):
        from django.core.management import call_command

        call_command("seed_meal_weight_catalog")
        self.admin = User.objects.create_user(
            username="catalog-admin@example.com",
            password="password",
            email="catalog-admin@example.com",
            is_staff=True,
        )
        self.client_user = User.objects.create_user(
            username="catalog-client@example.com",
            password="password",
            email="catalog-client@example.com",
            is_staff=False,
        )

    def test_admin_can_filter_meal_templates_by_category(self):
        self.client.force_authenticate(user=self.admin)

        response = self.client.get("/api/admin/meal-templates/?category=soup")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = response.json()
        results = payload if isinstance(payload, list) else payload["results"]
        self.assertEqual(len(results), 4)
        self.assertTrue(all(r["category"] == "soup" for r in results))

    def test_admin_day_editor_saves_soup_and_main_course_as_separate_items(self):
        self.client.force_authenticate(user=self.admin)
        soup = MealTemplate.objects.get(name="Polievka 2")
        main_course = MealTemplate.objects.get(name="Hlavný chod 3")

        response = self.client.post(
            "/api/admin/meal-plans/",
            {
                "date": "2026-04-01",
                "items_write": [
                    {"template_id": soup.id, "menu_variant": ""},
                    {"template_id": main_course.id, "menu_variant": ""},
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        plan = DailyMealPlan.objects.get(date="2026-04-01")
        self.assertEqual(plan.items.count(), 2)
        self.assertEqual(
            set(plan.items.values_list("category", flat=True)),
            {"soup", "main_course"},
        )

        # Backend keeps soup/main_course as separate items; the frontend merges
        # them into a single "Obed" section for the client.
        self.client.force_authenticate(user=self.client_user)
        by_date = self.client.get("/api/meal-plans/by-date/?date=2026-04-01")
        self.assertEqual(by_date.status_code, status.HTTP_200_OK)
        payload = by_date.json()
        categories = {item["category"] for item in payload["items"]}
        self.assertEqual(categories, {"soup", "main_course"})

    def test_admin_can_add_a_new_catalog_template_with_structured_components(self):
        self.client.force_authenticate(user=self.admin)

        response = self.client.post(
            "/api/admin/meal-templates/",
            {
                "category": "soup",
                "name": "Polievka 5",
                "components": [
                    {"label": "Hlavná zložka", "grams": "180", "unit": "ml"},
                    {"label": "Extra zložka 1", "grams": "20", "unit": "g"},
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        payload = response.json()
        self.assertEqual(payload["weight_label"], "180ml + 20g")
        self.assertEqual(payload["base_weight_grams"], "200.00")
        template = MealTemplate.objects.get(name="Polievka 5")
        self.assertEqual(template.category, "soup")
        self.assertTrue(template.is_active)

    def test_admin_can_add_a_new_template_with_a_unit_exception(self):
        self.client.force_authenticate(user=self.admin)

        response = self.client.post(
            "/api/admin/meal-templates/",
            {
                "category": "main_course",
                "name": "Hlavný chod 8",
                "components": [
                    {"label": "Príloha", "grams": "150", "unit": "g"},
                ],
                "unit_exception": {
                    "component_label": "Klobása",
                    "unit": "ks",
                    "counts_by_portion_type": {"Škôlka": "1", "Dospelý (SŠ)": "2"},
                },
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        template = MealTemplate.objects.get(name="Hlavný chod 8")
        self.assertEqual(template.unit_exception["component_label"], "Klobása")
        # The piece-count exception must be baked into the composition text
        # itself, not just stored as a separate field admins might miss.
        self.assertEqual(
            template.weight_label, "150g + Klobása (ks podľa vekovej skupiny)"
        )

    def test_admin_can_edit_an_existing_template(self):
        self.client.force_authenticate(user=self.admin)
        template = MealTemplate.objects.get(name="Olovrant 1")

        response = self.client.patch(
            f"/api/admin/meal-templates/{template.id}/",
            {"is_active": False},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        template.refresh_from_db()
        self.assertFalse(template.is_active)

    def test_creating_a_template_without_weight_data_is_rejected(self):
        self.client.force_authenticate(user=self.admin)

        response = self.client.post(
            "/api/admin/meal-templates/",
            {"category": "soup", "name": "Polievka bez váhy"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_client_cannot_write_meal_templates(self):
        self.client.force_authenticate(user=self.client_user)

        response = self.client.post(
            "/api/admin/meal-templates/",
            {
                "category": "soup",
                "name": "Polievka X",
                "components": [
                    {"label": "Hlavná zložka", "grams": "100", "unit": "ml"}
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)


class AdminPortionTypeApiTest(APITestCase):
    def setUp(self):
        from django.core.management import call_command

        call_command("init_reference_data")
        self.admin = User.objects.create_user(
            username="pt-admin@example.com",
            password="password",
            email="pt-admin@example.com",
            is_staff=True,
        )
        self.client_user = User.objects.create_user(
            username="pt-client@example.com",
            password="password",
            email="pt-client@example.com",
            is_staff=False,
        )

    def test_admin_can_update_a_portion_type_coefficient(self):
        self.client.force_authenticate(user=self.admin)
        skolka = PortionType.objects.get(name="Škôlka")

        response = self.client.patch(
            f"/api/admin/portion-types/{skolka.id}/",
            {"coefficient": "1.1000"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        skolka.refresh_from_db()
        self.assertEqual(str(skolka.coefficient), "1.1000")

    def test_client_cannot_update_a_portion_type_coefficient(self):
        self.client.force_authenticate(user=self.client_user)
        skolka = PortionType.objects.get(name="Škôlka")

        response = self.client.patch(
            f"/api/admin/portion-types/{skolka.id}/",
            {"coefficient": "1.1000"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        skolka.refresh_from_db()
        self.assertNotEqual(str(skolka.coefficient), "1.1000")


class AdminDailyReportTest(APITestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            username="admin2@example.com",
            password="password",
            email="admin2@example.com",
            is_staff=True,
        )
        self.client_user = User.objects.create_user(
            username="anna@test.sk",
            password="password",
            first_name="Anna",
            last_name="Novák",
            email="anna@test.sk",
            is_staff=False,
        )
        self.today = date.today().isoformat()
        DailyOrder.objects.create(
            user=self.client_user,
            date=self.today,
            status="submitted",
            data={
                "breakfast": {
                    "Dospelý": {"menuCounts": {"A": 2}, "diets": {"No Milk": 1}}
                },
                "lunch": {"ZŠ": {"menuCounts": {"A": 5, "B": 1}, "diets": {}}},
                "olovrant": {},
            },
        )

    def test_daily_report_returns_rows(self):
        self.client.force_authenticate(user=self.admin)
        response = self.client.get(
            f"/api/admin/summary/daily-report/?date={self.today}"
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        data = response.json()

        self.assertEqual(data["date"], self.today)
        self.assertEqual(len(data["rows"]), 1)

        row = data["rows"][0]
        self.assertEqual(row["email"], "anna@test.sk")
        self.assertEqual(row["breakfast"]["total"], 2)
        self.assertEqual(row["lunch"]["total"], 6)
        self.assertEqual(row["olovrant"]["total"], 0)
        self.assertEqual(row["total"], 8)

        self.assertEqual(data["totals"]["grand"], 8)
        self.assertEqual(data["totals"]["breakfast"]["total"], 2)
        self.assertEqual(data["totals"]["breakfast"]["diets"]["No Milk"], 1)

    def test_daily_report_requires_date(self):
        self.client.force_authenticate(user=self.admin)
        response = self.client.get("/api/admin/summary/daily-report/")
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_daily_report_invalid_date_format(self):
        self.client.force_authenticate(user=self.admin)
        for bad_date in ["not-a-date", "2024/01/01", "01-01-2024", "<script>"]:
            response = self.client.get(
                f"/api/admin/summary/daily-report/?date={bad_date}"
            )
            self.assertEqual(
                response.status_code,
                status.HTTP_400_BAD_REQUEST,
                msg=f"Expected 400 for bad date: {bad_date!r}",
            )

    def test_daily_report_client_forbidden(self):
        self.client.force_authenticate(user=self.client_user)
        response = self.client.get(
            f"/api/admin/summary/daily-report/?date={self.today}"
        )
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_daily_report_xlsx_returns_file(self):
        self.client.force_authenticate(user=self.admin)
        response = self.client.get(
            f"/api/admin/summary/daily-report-xlsx/?date={self.today}"
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("prehlad_", response["Content-Disposition"])
        self.assertGreater(len(response.content), 100)

    def test_daily_report_xlsx_client_forbidden(self):
        self.client.force_authenticate(user=self.client_user)
        response = self.client.get(
            f"/api/admin/summary/daily-report-xlsx/?date={self.today}"
        )
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_daily_report_xlsx_invalid_date_format(self):
        self.client.force_authenticate(user=self.admin)
        for bad_date in ["not-a-date", "2024/01/01", "01-01-2024"]:
            response = self.client.get(
                f"/api/admin/summary/daily-report-xlsx/?date={bad_date}"
            )
            self.assertEqual(
                response.status_code,
                status.HTTP_400_BAD_REQUEST,
                msg=f"Expected 400 for bad date: {bad_date!r}",
            )

    def test_daily_report_flat_shape(self):
        """Flat meal shape {menuCounts, diets} must be aggregated correctly."""
        flat_user = User.objects.create_user(
            username="flatclient@example.com",
            password="password",
            email="flatclient@example.com",
            is_staff=False,
        )
        DailyOrder.objects.create(
            user=flat_user,
            date=self.today,
            status="submitted",
            data={
                "breakfast": {"menuCounts": {"A": 3}, "diets": {"Bezlepkový": 1}},
                "lunch": {"menuCounts": {"B": 2}, "diets": {}},
                "olovrant": {},
            },
        )
        self.client.force_authenticate(user=self.admin)
        response = self.client.get(
            f"/api/admin/summary/daily-report/?date={self.today}"
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        flat_row = next(
            r for r in response.json()["rows"] if r["email"] == "flatclient@example.com"
        )
        self.assertEqual(flat_row["breakfast"]["total"], 3)
        self.assertEqual(flat_row["lunch"]["total"], 2)
        self.assertEqual(flat_row["total"], 5)

    def test_daily_report_xlsx_flat_shape(self):
        """XLSX export must include columns for flat-shape meal data."""
        flat_user = User.objects.create_user(
            username="flatclient2@example.com",
            password="password",
            email="flatclient2@example.com",
            is_staff=False,
        )
        DailyOrder.objects.create(
            user=flat_user,
            date=self.today,
            status="submitted",
            data={
                "breakfast": {"menuCounts": {"A": 1}, "diets": {}},
                "lunch": {},
                "olovrant": {},
            },
        )
        self.client.force_authenticate(user=self.admin)
        response = self.client.get(
            f"/api/admin/summary/daily-report-xlsx/?date={self.today}"
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertGreater(len(response.content), 100)

    def test_daily_report_xlsx_has_category_columns(self):
        """XLSX must have 3-level headers: meal → category/size → menu/diet."""
        import io

        import openpyxl

        self.client.force_authenticate(user=self.admin)
        response = self.client.get(
            f"/api/admin/summary/daily-report-xlsx/?date={self.today}"
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Row 3: meal-level labels
        row3 = [ws.cell(row=3, column=c).value for c in range(1, ws.max_column + 1)]
        self.assertIn("Raňajky", row3)
        self.assertIn("Obed", row3)

        # Row 4: category-level labels (from setUp order data)
        row4 = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
        self.assertIn("Dospelý", row4)  # breakfast category in setUp
        self.assertIn("ZŠ", row4)  # lunch category in setUp

        # Row 5: menu/diet labels
        row5 = [ws.cell(row=5, column=c).value for c in range(1, ws.max_column + 1)]
        self.assertIn("Menu A", row5)
        self.assertIn("No Milk", row5)  # diet from setUp breakfast

        # Data rows start at row 6
        row6 = [ws.cell(row=6, column=c).value for c in range(1, ws.max_column + 1)]
        self.assertIn("anna@test.sk", row6)

        # SPOLU row is the last row
        last_row = ws.max_row
        spolu_vals = [
            ws.cell(row=last_row, column=c).value for c in range(1, ws.max_column + 1)
        ]
        self.assertIn("SPOLU", spolu_vals)

    def test_daily_report_pdf_returns_file(self):
        self.client.force_authenticate(user=self.admin)
        response = self.client.get(
            f"/api/admin/summary/daily-report-pdf/?date={self.today}"
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertIn("prehlad_", response["Content-Disposition"])
        # PDF magic bytes — consume streaming response
        content = b"".join(response.streaming_content)
        self.assertTrue(content.startswith(b"%PDF"))

    def test_daily_report_pdf_client_forbidden(self):
        self.client.force_authenticate(user=self.client_user)
        response = self.client.get(
            f"/api/admin/summary/daily-report-pdf/?date={self.today}"
        )
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_daily_report_pdf_requires_date(self):
        self.client.force_authenticate(user=self.admin)
        response = self.client.get("/api/admin/summary/daily-report-pdf/")
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_daily_report_pdf_invalid_date_format(self):
        self.client.force_authenticate(user=self.admin)
        for bad_date in ["not-a-date", "2024/01/01", "01-01-2024"]:
            response = self.client.get(
                f"/api/admin/summary/daily-report-pdf/?date={bad_date}"
            )
            self.assertEqual(
                response.status_code,
                status.HTTP_400_BAD_REQUEST,
                msg=f"Expected 400 for bad date: {bad_date!r}",
            )


class AdminMealPlanByDateTest(APITestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            username="mealplan-admin@example.com",
            password="password",
            email="mealplan-admin@example.com",
            is_staff=True,
        )

    def test_by_date_returns_empty_payload_when_plan_missing(self):
        self.client.force_authenticate(user=self.admin)
        response = self.client.get("/api/admin/meal-plans/by-date/?date=2026-03-16")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response.json(),
            {
                "exists": False,
                "date": "2026-03-16",
                "notes": "",
                "items": [],
            },
        )

    def test_by_date_returns_existing_plan_payload(self):
        DailyMealPlan.objects.create(date="2026-03-16", created_by=self.admin)
        self.client.force_authenticate(user=self.admin)

        response = self.client.get("/api/admin/meal-plans/by-date/?date=2026-03-16")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.json()["exists"])


class AdminMealPlanApiTest(APITestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            username="mealplan-admin2@example.com",
            password="password",
            email="mealplan-admin2@example.com",
            is_staff=True,
        )
        self.client_user = User.objects.create_user(
            username="mealplan-client@example.com",
            password="password",
            email="mealplan-client@example.com",
        )
        self.kindergarten = PortionType.objects.create(
            name="Škôlka", coefficient="0.5000", sort_order=1
        )
        self.adult = PortionType.objects.create(
            name="Dospelý (SŠ)", coefficient="1.0000", sort_order=2
        )
        self.breakfast_template = MealTemplate.objects.create(
            category="breakfast_snack",
            name="Chlieb + maslo",
            weight_label="80g + 20g",
            base_weight_grams="100.00",
        )
        self.lunch_template = MealTemplate.objects.create(
            category="main_course",
            name="Kuracie prsia + ryža",
            weight_label="120g + 80g",
            base_weight_grams="200.00",
            menu_variant="A",
        )
        self.snack_template = MealTemplate.objects.create(
            category="afternoon_snack",
            name="Jablko",
            weight_label="50g",
            base_weight_grams="50.00",
        )
        self.client.force_authenticate(user=self.admin)

    def _create_plan(self, target_date="2026-03-16", notes="Plan notes"):
        plan = DailyMealPlan.objects.create(
            date=target_date,
            notes=notes,
            created_by=self.admin,
        )
        MealPlanItem.objects.create(
            meal_plan=plan,
            template=self.breakfast_template,
            category="breakfast_snack",
            menu_variant="",
        )
        MealPlanItem.objects.create(
            meal_plan=plan,
            template=self.lunch_template,
            category="main_course",
            menu_variant="A",
        )
        MealPlanItem.objects.create(
            meal_plan=plan,
            template=self.snack_template,
            category="afternoon_snack",
            menu_variant="",
        )
        EnrolledCount.objects.create(
            meal_plan=plan,
            portion_type=self.kindergarten,
            count=10,
        )
        EnrolledCount.objects.create(
            meal_plan=plan,
            portion_type=self.adult,
            count=2,
        )
        return plan

    def test_create_meal_plan_via_api_persists_nested_data(self):
        response = self.client.post(
            "/api/admin/meal-plans/",
            {
                "date": "2026-03-16",
                "notes": "Novy plan",
                "items_write": [
                    {"template_id": self.breakfast_template.id, "menu_variant": ""},
                    {"template_id": self.lunch_template.id, "menu_variant": "A"},
                ],
                "enrolled_counts_write": [
                    {"portion_type_id": self.kindergarten.id, "count": 12},
                    {"portion_type_id": self.adult.id, "count": 3},
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        plan = DailyMealPlan.objects.get(date="2026-03-16")
        self.assertEqual(plan.created_by, self.admin)
        self.assertEqual(plan.notes, "Novy plan")
        self.assertEqual(plan.items.count(), 2)
        self.assertEqual(plan.enrolled_counts.count(), 2)
        self.assertEqual(
            plan.enrolled_counts.get(portion_type=self.kindergarten).count,
            12,
        )

    def test_posting_the_same_date_twice_updates_the_existing_plan_instead_of_400ing(
        self,
    ):
        # Admin day editor always POSTs the full day state, even when a plan
        # for that date already exists (create_or_replace_plan is meant to
        # upsert by date) — this must not be rejected by a uniqueness check.
        first = self.client.post(
            "/api/admin/meal-plans/",
            {
                "date": "2026-03-18",
                "items_write": [
                    {"template_id": self.breakfast_template.id, "menu_variant": ""},
                ],
            },
            format="json",
        )
        self.assertEqual(first.status_code, status.HTTP_201_CREATED)

        second = self.client.post(
            "/api/admin/meal-plans/",
            {
                "date": "2026-03-18",
                "items_write": [
                    {"template_id": self.lunch_template.id, "menu_variant": "A"},
                ],
            },
            format="json",
        )

        self.assertEqual(second.status_code, status.HTTP_201_CREATED)
        self.assertEqual(DailyMealPlan.objects.filter(date="2026-03-18").count(), 1)
        plan = DailyMealPlan.objects.get(date="2026-03-18")
        self.assertEqual(plan.items.count(), 1)
        self.assertEqual(plan.items.first().template_id, self.lunch_template.id)

    def test_gramage_report_returns_computed_totals(self):
        plan = self._create_plan()

        response = self.client.get(f"/api/admin/meal-plans/{plan.id}/gramage-report/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = response.json()
        self.assertEqual(payload["date"], "2026-03-16")
        self.assertEqual(
            payload["sections"]["breakfast_snack"]["section_total_grams"], "700.00"
        )
        self.assertEqual(
            payload["sections"]["main_course"]["section_total_grams"], "1400.00"
        )
        self.assertEqual(
            payload["sections"]["afternoon_snack"]["section_total_grams"], "350.00"
        )
        self.assertEqual(payload["grand_total_grams"], "2450.00")

    def test_meal_plan_exports_return_downloadable_files(self):
        plan = self._create_plan()

        xlsx_response = self.client.get(f"/api/admin/meal-plans/{plan.id}/export-xlsx/")
        self.assertEqual(xlsx_response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            xlsx_response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = openpyxl.load_workbook(io.BytesIO(xlsx_response.content))
        self.assertIn("2026-03-16", workbook.sheetnames)
        self.assertEqual(workbook.active["A1"].value, "Jedálniček — 2026-03-16")

        pdf_response = self.client.get(f"/api/admin/meal-plans/{plan.id}/export-pdf/")
        self.assertEqual(pdf_response.status_code, status.HTTP_200_OK)
        self.assertEqual(pdf_response["Content-Type"], "application/pdf")
        self.assertTrue(pdf_response.content.startswith(b"%PDF"))

    def test_range_report_and_range_export_xlsx_cover_multiple_days(self):
        self._create_plan("2026-03-16", "Prvy plan")
        self._create_plan("2026-03-17", "Druhy plan")

        range_response = self.client.get(
            "/api/admin/meal-plans/range-report/?from=2026-03-16&to=2026-03-17"
        )
        self.assertEqual(range_response.status_code, status.HTTP_200_OK)
        payload = range_response.json()
        self.assertEqual(len(payload), 2)
        self.assertEqual(
            {item["date"] for item in payload}, {"2026-03-16", "2026-03-17"}
        )

        xlsx_response = self.client.get(
            "/api/admin/meal-plans/range-export-xlsx/?from=2026-03-16&to=2026-03-17"
        )
        self.assertEqual(xlsx_response.status_code, status.HTTP_200_OK)
        workbook = openpyxl.load_workbook(io.BytesIO(xlsx_response.content))
        self.assertEqual(workbook.sheetnames[0], "Súhrn")
        self.assertIn("2026-03-16", workbook.sheetnames)
        self.assertIn("2026-03-17", workbook.sheetnames)

    def test_diet_summary_and_gramage_dashboard_return_expected_contract(self):
        self._create_plan()
        settings, _ = ClientSettings.objects.get_or_create(user=self.client_user)
        settings.admin_order_note = "Bez cibule v pondelok"
        settings.save(update_fields=["admin_order_note"])
        DailyOrder.objects.create(
            user=self.client_user,
            date="2026-03-16",
            status="submitted",
            data={
                "breakfast": {
                    "Škôlka": {"menuCounts": {"A": 4}, "diets": {"Vegan": 1}}
                },
                "lunch": {
                    "Škôlka": {
                        "menuCounts": {"A": 6},
                        "diets": {"Bezlepková": 2},
                    }
                },
                "olovrant": {"Škôlka": {"menuCounts": {"A": 3}, "diets": {}}},
            },
        )

        diet_summary = self.client.get(
            "/api/admin/meal-plans/diet-summary/?date=2026-03-16"
        )
        self.assertEqual(diet_summary.status_code, status.HTTP_200_OK)
        diet_payload = diet_summary.json()
        self.assertEqual(diet_payload["diet_by_meal"]["breakfast"]["Vegan"], 1)
        self.assertEqual(diet_payload["diet_by_meal"]["lunch"]["Bezlepková"], 2)
        self.assertEqual(diet_payload["menu_totals"]["lunch"]["A"], 6)

        dashboard = self.client.get(
            "/api/admin/meal-plans/gramage-dashboard/?date=2026-03-16"
        )
        self.assertEqual(dashboard.status_code, status.HTTP_200_OK)
        dashboard_payload = dashboard.json()
        self.assertEqual(
            dashboard_payload["meal_plan_id"],
            DailyMealPlan.objects.get(date="2026-03-16").id,
        )
        self.assertEqual(len(dashboard_payload["col_groups"]), 3)
        self.assertEqual(len(dashboard_payload["rows"]), 1)
        self.assertEqual(
            dashboard_payload["rows"][0]["diet_summary_rows"][0]["name"], "Bezlepková"
        )
        self.assertEqual(
            dashboard_payload["rows"][0]["admin_order_note"],
            "Bez cibule v pondelok",
        )

    def test_gramage_dashboard_diet_count_exceeding_first_variant_is_not_double_counted(
        self,
    ):
        """Regression test: when a portion's diet count is larger than the
        first (lowest-VARIANT_ORDER) menu variant's own count, the excess
        must carry over and be subtracted from the next variant too -
        otherwise those diet students are never removed from the raw
        variant total and get counted twice (once in the variant column,
        once in the diet sub-row)."""
        plan = DailyMealPlan.objects.create(
            date="2026-03-18",
            notes="Two lunch variants",
            created_by=self.admin,
        )
        variant_b_template = MealTemplate.objects.create(
            category="main_course",
            name="Zeleninové rizoto",
            weight_label="200g",
            base_weight_grams="200.00",
            menu_variant="B",
        )
        MealPlanItem.objects.create(
            meal_plan=plan,
            template=self.lunch_template,
            category="main_course",
            menu_variant="A",
        )
        MealPlanItem.objects.create(
            meal_plan=plan,
            template=variant_b_template,
            category="main_course",
            menu_variant="B",
        )
        EnrolledCount.objects.create(
            meal_plan=plan, portion_type=self.kindergarten, count=22
        )

        DailyOrder.objects.create(
            user=self.client_user,
            date="2026-03-18",
            status="submitted",
            data={
                "lunch": {
                    "Škôlka": {
                        "menuCounts": {"A": 2, "B": 20},
                        "diets": {"Bezlepková": 8},
                    }
                },
            },
        )

        dashboard = self.client.get(
            "/api/admin/meal-plans/gramage-dashboard/?date=2026-03-18"
        )
        self.assertEqual(dashboard.status_code, status.HTTP_200_OK)
        row = dashboard.json()["rows"][0]
        standard_rows = {
            r["variant"]: r["count"] for r in row["sub_rows"] if r["type"] == "standard"
        }
        diet_rows = [r for r in row["sub_rows"] if r["type"] == "diet"]

        # 8 diet students exceed variant A's raw count of 2, so the
        # remaining 6 must be subtracted from variant B, not left in it.
        self.assertEqual(standard_rows.get("A", 0), 0)
        self.assertEqual(standard_rows["B"], 14)
        self.assertEqual(diet_rows[0]["count"], 8)
        # Total headcount (2 + 20) must be preserved: 0 + 14 + 8 == 22.
        self.assertEqual(
            standard_rows.get("A", 0) + standard_rows["B"] + diet_rows[0]["count"], 22
        )

    def test_gramage_dashboard_splits_main_course_menu_variants(self):
        plan = DailyMealPlan.objects.create(
            date="2026-03-19",
            notes="Four lunch variants",
            created_by=self.admin,
        )
        templates_by_variant = {
            "A": self.lunch_template,
            "B": MealTemplate.objects.create(
                category="main_course",
                name="Menu B",
                weight_label="300g",
                base_weight_grams="300.00",
                menu_variant="B",
            ),
            "C": MealTemplate.objects.create(
                category="main_course",
                name="Menu C",
                weight_label="400g",
                base_weight_grams="400.00",
                menu_variant="C",
            ),
            "V": MealTemplate.objects.create(
                category="main_course",
                name="Menu V",
                weight_label="500g",
                base_weight_grams="500.00",
                menu_variant="V",
            ),
        }
        for variant, template in templates_by_variant.items():
            MealPlanItem.objects.create(
                meal_plan=plan,
                template=template,
                category="main_course",
                menu_variant=variant,
            )
        EnrolledCount.objects.create(
            meal_plan=plan, portion_type=self.kindergarten, count=10
        )

        DailyOrder.objects.create(
            user=self.client_user,
            date="2026-03-19",
            status="submitted",
            data={
                "lunch": {
                    "Škôlka": {
                        "menuCounts": {"A": 1, "B": 2, "C": 3, "V": 4},
                        "diets": {},
                    }
                },
            },
        )

        response = self.client.get(
            "/api/admin/meal-plans/gramage-dashboard/?date=2026-03-19"
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = response.json()
        self.assertEqual(
            [cg["variant"] for cg in payload["col_groups"]],
            ["A", "B", "C", "V"],
        )
        row = payload["rows"][0]
        standard_rows = {
            r["variant"]: r for r in row["sub_rows"] if r["type"] == "standard"
        }
        self.assertEqual(
            {variant: data["count"] for variant, data in standard_rows.items()},
            {"A": 1, "B": 2, "C": 3, "V": 4},
        )
        self.assertEqual(standard_rows["A"]["col_grams"][0], ["60.00", "40.00"])
        self.assertEqual(standard_rows["B"]["col_grams"][1], ["300.00"])
        self.assertEqual(standard_rows["C"]["col_grams"][2], ["600.00"])
        self.assertEqual(standard_rows["V"]["col_grams"][3], ["1000.00"])

    def test_gramage_dashboard_diet_uses_default_variant_without_diet_template(self):
        plan = DailyMealPlan.objects.create(
            date="2026-03-20",
            notes="Diet fallback",
            created_by=self.admin,
        )
        MealPlanItem.objects.create(
            meal_plan=plan,
            template=self.lunch_template,
            category="main_course",
            menu_variant="A",
        )

        DailyOrder.objects.create(
            user=self.client_user,
            date="2026-03-20",
            status="submitted",
            data={
                "lunch": {
                    "Škôlka": {
                        "menuCounts": {"A": 4},
                        "diets": {"NO MILK": 2},
                    }
                },
            },
        )

        response = self.client.get(
            "/api/admin/meal-plans/gramage-dashboard/?date=2026-03-20"
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = response.json()
        self.assertEqual(len(payload["col_groups"]), 1)
        diet_row = payload["rows"][0]["diet_summary_rows"][0]

        self.assertEqual(diet_row["name"], "NO MILK")
        self.assertEqual(diet_row["count"], 2)
        # Diets inherit the A/default 120g + 80g split with Škôlka coefficient 0.5.
        self.assertEqual(diet_row["col_grams"][0], ["120.00", "80.00"])

    def test_gramage_dashboard_prefers_explicit_diet_template(self):
        no_milk = Diet.objects.create(name="NO MILK", is_active=True)
        plan = DailyMealPlan.objects.create(
            date="2026-03-21",
            notes="Diet override",
            created_by=self.admin,
        )
        diet_template = MealTemplate.objects.create(
            category="main_course",
            name="No milk menu",
            weight_label="50g + 50g",
            base_weight_grams="100.00",
            menu_variant="A",
            diet=no_milk,
        )
        MealPlanItem.objects.create(
            meal_plan=plan,
            template=self.lunch_template,
            category="main_course",
            menu_variant="A",
        )
        MealPlanItem.objects.create(
            meal_plan=plan,
            template=diet_template,
            category="main_course",
            menu_variant="A",
            diet=no_milk,
        )

        DailyOrder.objects.create(
            user=self.client_user,
            date="2026-03-21",
            status="submitted",
            data={
                "lunch": {
                    "Škôlka": {
                        "menuCounts": {"A": 4},
                        "diets": {"NO MILK": 2},
                    }
                },
            },
        )

        response = self.client.get(
            "/api/admin/meal-plans/gramage-dashboard/?date=2026-03-21"
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = response.json()
        self.assertEqual(
            [(cg["variant"], cg["diet_name"]) for cg in payload["col_groups"]],
            [("A", None), ("A", "NO MILK")],
        )
        diet_row = payload["rows"][0]["diet_summary_rows"][0]

        self.assertEqual(diet_row["name"], "NO MILK")
        self.assertEqual(diet_row["col_grams"][0], ["0.00", "0.00"])
        self.assertEqual(diet_row["col_grams"][1], ["50.00", "50.00"])
        self.assertEqual(
            [
                (section["variant"], section["diet_name"], section["diets"])
                for section in payload["count_summary"]
            ],
            [
                ("A", None, []),
                (
                    "A",
                    "NO MILK",
                    [{"label": "Škôlka - NO MILK", "count": 2}],
                ),
            ],
        )

    def test_gramage_dashboard_exports_return_files(self):
        self._create_plan()
        DailyOrder.objects.create(
            user=self.client_user,
            date="2026-03-16",
            status="submitted",
            data={
                "breakfast": {"Škôlka": {"menuCounts": {"A": 2}, "diets": {}}},
                "lunch": {
                    "Škôlka": {
                        "menuCounts": {"A": 4},
                        "diets": {"Bezlepková": 1},
                    }
                },
                "olovrant": {"Škôlka": {"menuCounts": {"A": 1}, "diets": {}}},
            },
        )

        xlsx_response = self.client.get(
            "/api/admin/meal-plans/gramage-dashboard-xlsx/?date=2026-03-16"
        )
        self.assertEqual(xlsx_response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            xlsx_response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = openpyxl.load_workbook(io.BytesIO(xlsx_response.content))
        self.assertEqual(workbook.active["A1"].value, "Gramáž jedál — 2026-03-16")

        pdf_response = self.client.get(
            "/api/admin/meal-plans/gramage-dashboard-pdf/?date=2026-03-16"
        )
        self.assertEqual(pdf_response.status_code, status.HTTP_200_OK)
        self.assertEqual(pdf_response["Content-Type"], "application/pdf")
        self.assertTrue(pdf_response.content.startswith(b"%PDF"))

    def test_meal_plan_endpoints_require_admin(self):
        self.client.force_authenticate(user=self.client_user)

        response = self.client.get("/api/admin/meal-plans/by-date/?date=2026-03-16")

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
