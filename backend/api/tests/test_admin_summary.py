from datetime import date

from django.contrib.auth.models import User
from rest_framework import status
from rest_framework.test import APITestCase

from api.models import DailyOrder


class AdminSummaryTest(APITestCase):
    def setUp(self):
        # Create Admin
        self.admin = User.objects.create_user(
            username="admin@example.com", password="password", email="admin@example.com", is_staff=True
        )
        # Create Client
        self.client_user = User.objects.create_user(
            username="client@example.com", password="password", email="client@example.com", is_staff=False
        )

        self.today = date.today().isoformat()

        # Create complex Detailed Order 1
        DailyOrder.objects.create(
            user=self.client_user,
            date=self.today,
            status="submitted",
            data={
                "breakfast": {
                    "Dospelý (SŠ)": {"menuCounts": {"A": 2}, "diets": {"Bez lepku": 1}}
                },
                "lunch": {
                    "ZŠ 1.stupeň": {"menuCounts": {"A": 5, "B": 2}, "diets": {}},
                    "Dospelý (SŠ)": {
                        "menuCounts": {"A": 1},
                        "diets": {"Bez laktózy": 1},
                    },
                },
                "olovrant": {"Jasle": {"menuCounts": {"A": 3}, "diets": {}}},
            },
        )

        # Create second client/order
        self.client2 = User.objects.create_user(username="client2@example.com", password="password", email="client2@example.com")
        DailyOrder.objects.create(
            user=self.client2,
            date=self.today,
            status="draft",
            data={"lunch": {"ZŠ 1.stupeň": {"menuCounts": {"A": 3}, "diets": {}}}},
        )

    def test_admin_can_access_summary(self):
        self.client.force_authenticate(user=self.admin)
        response = self.client.get(f"/api/admin/summary/daily-stats/?date={self.today}")
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        data = response.json()
        self.assertEqual(data["total_orders"], 2)

        # Check Breakfast - Dospelý
        # Expecting: 2xA, 1x Gluten
        bf_stats = data["meals"]["breakfast"]["Dospelý (SŠ)"]
        self.assertEqual(bf_stats["menus"]["A"], 2)
        self.assertEqual(bf_stats["diets"]["Bez lepku"], 1)
        self.assertEqual(bf_stats["total"], 2)

        # Check Lunch - ZŠ 1.stupeň
        # Order 1: 5xA, 2xB. Order 2: 3xA. Total: 8xA, 2xB
        lunch_zs = data["meals"]["lunch"]["ZŠ 1.stupeň"]
        self.assertEqual(lunch_zs["menus"]["A"], 8)
        self.assertEqual(lunch_zs["menus"]["B"], 2)
        self.assertEqual(lunch_zs["total"], 10)

        # Check Lunch - Dospelý
        lunch_adult = data["meals"]["lunch"]["Dospelý (SŠ)"]
        self.assertEqual(lunch_adult["menus"]["A"], 1)
        self.assertEqual(lunch_adult["diets"]["Bez laktózy"], 1)

        # Check Olovrant
        olovrant_jasle = data["meals"]["olovrant"]["Jasle"]
        self.assertEqual(olovrant_jasle["menus"]["A"], 3)

    def test_client_cannot_access_summary(self):
        self.client.force_authenticate(user=self.client_user)
        response = self.client.get(f"/api/admin/summary/daily-stats/?date={self.today}")
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)


class AdminDailyReportTest(APITestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            username="admin2@example.com", password="password", email="admin2@example.com", is_staff=True
        )
        self.client_user = User.objects.create_user(
            username="anna@test.sk",
            password="password",
            first_name="Anna",
            last_name="Novák",
            email="anna@test.sk",
            is_staff=False,
        )
        self.today = date.today().isoformat()
        DailyOrder.objects.create(
            user=self.client_user,
            date=self.today,
            status="submitted",
            data={
                "breakfast": {
                    "Dospelý": {"menuCounts": {"A": 2}, "diets": {"No Milk": 1}}
                },
                "lunch": {"ZŠ": {"menuCounts": {"A": 5, "B": 1}, "diets": {}}},
                "olovrant": {},
            },
        )

    def test_daily_report_returns_rows(self):
        self.client.force_authenticate(user=self.admin)
        response = self.client.get(
            f"/api/admin/summary/daily-report/?date={self.today}"
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        data = response.json()

        self.assertEqual(data["date"], self.today)
        self.assertEqual(len(data["rows"]), 1)

        row = data["rows"][0]
        self.assertEqual(row["email"], "anna@test.sk")
        self.assertEqual(row["breakfast"]["total"], 2)
        self.assertEqual(row["lunch"]["total"], 6)
        self.assertEqual(row["olovrant"]["total"], 0)
        self.assertEqual(row["total"], 8)

        self.assertEqual(data["totals"]["grand"], 8)
        self.assertEqual(data["totals"]["breakfast"]["total"], 2)
        self.assertEqual(data["totals"]["breakfast"]["diets"]["No Milk"], 1)

    def test_daily_report_requires_date(self):
        self.client.force_authenticate(user=self.admin)
        response = self.client.get("/api/admin/summary/daily-report/")
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_daily_report_invalid_date_format(self):
        self.client.force_authenticate(user=self.admin)
        for bad_date in ["not-a-date", "2024/01/01", "01-01-2024", "<script>"]:
            response = self.client.get(
                f"/api/admin/summary/daily-report/?date={bad_date}"
            )
            self.assertEqual(
                response.status_code,
                status.HTTP_400_BAD_REQUEST,
                msg=f"Expected 400 for bad date: {bad_date!r}",
            )

    def test_daily_report_client_forbidden(self):
        self.client.force_authenticate(user=self.client_user)
        response = self.client.get(
            f"/api/admin/summary/daily-report/?date={self.today}"
        )
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_daily_report_xlsx_returns_file(self):
        self.client.force_authenticate(user=self.admin)
        response = self.client.get(
            f"/api/admin/summary/daily-report-xlsx/?date={self.today}"
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("prehlad_", response["Content-Disposition"])
        self.assertGreater(len(response.content), 100)

    def test_daily_report_xlsx_client_forbidden(self):
        self.client.force_authenticate(user=self.client_user)
        response = self.client.get(
            f"/api/admin/summary/daily-report-xlsx/?date={self.today}"
        )
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_daily_report_xlsx_invalid_date_format(self):
        self.client.force_authenticate(user=self.admin)
        for bad_date in ["not-a-date", "2024/01/01", "01-01-2024"]:
            response = self.client.get(
                f"/api/admin/summary/daily-report-xlsx/?date={bad_date}"
            )
            self.assertEqual(
                response.status_code,
                status.HTTP_400_BAD_REQUEST,
                msg=f"Expected 400 for bad date: {bad_date!r}",
            )

    def test_daily_report_flat_shape(self):
        """Flat meal shape {menuCounts, diets} must be aggregated correctly."""
        flat_user = User.objects.create_user(
            username="flatclient@example.com", password="password", email="flatclient@example.com", is_staff=False
        )
        DailyOrder.objects.create(
            user=flat_user,
            date=self.today,
            status="submitted",
            data={
                "breakfast": {"menuCounts": {"A": 3}, "diets": {"Bezlepkový": 1}},
                "lunch": {"menuCounts": {"B": 2}, "diets": {}},
                "olovrant": {},
            },
        )
        self.client.force_authenticate(user=self.admin)
        response = self.client.get(
            f"/api/admin/summary/daily-report/?date={self.today}"
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        flat_row = next(
            r for r in response.json()["rows"] if r["email"] == "flatclient@example.com"
        )
        self.assertEqual(flat_row["breakfast"]["total"], 3)
        self.assertEqual(flat_row["lunch"]["total"], 2)
        self.assertEqual(flat_row["total"], 5)

    def test_daily_report_xlsx_flat_shape(self):
        """XLSX export must include columns for flat-shape meal data."""
        flat_user = User.objects.create_user(
            username="flatclient2@example.com", password="password", email="flatclient2@example.com", is_staff=False
        )
        DailyOrder.objects.create(
            user=flat_user,
            date=self.today,
            status="submitted",
            data={
                "breakfast": {"menuCounts": {"A": 1}, "diets": {}},
                "lunch": {},
                "olovrant": {},
            },
        )
        self.client.force_authenticate(user=self.admin)
        response = self.client.get(
            f"/api/admin/summary/daily-report-xlsx/?date={self.today}"
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertGreater(len(response.content), 100)

    def test_daily_report_xlsx_has_category_columns(self):
        """XLSX must have 3-level headers: meal → category/size → menu/diet."""
        import io

        import openpyxl

        self.client.force_authenticate(user=self.admin)
        response = self.client.get(
            f"/api/admin/summary/daily-report-xlsx/?date={self.today}"
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Row 3: meal-level labels
        row3 = [ws.cell(row=3, column=c).value for c in range(1, ws.max_column + 1)]
        self.assertIn("Raňajky", row3)
        self.assertIn("Obed", row3)

        # Row 4: category-level labels (from setUp order data)
        row4 = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
        self.assertIn("Dospelý", row4)  # breakfast category in setUp
        self.assertIn("ZŠ", row4)  # lunch category in setUp

        # Row 5: menu/diet labels
        row5 = [ws.cell(row=5, column=c).value for c in range(1, ws.max_column + 1)]
        self.assertIn("Menu A", row5)
        self.assertIn("No Milk", row5)  # diet from setUp breakfast

        # Data rows start at row 6
        row6 = [ws.cell(row=6, column=c).value for c in range(1, ws.max_column + 1)]
        self.assertIn("Anna Novák", row6)

        # SPOLU row is the last row
        last_row = ws.max_row
        spolu_vals = [
            ws.cell(row=last_row, column=c).value for c in range(1, ws.max_column + 1)
        ]
        self.assertIn("SPOLU", spolu_vals)
