"""
TDD tests for JedalnicekXlsxParser.

All tests are pure-Python (no DB) because the parser is a stateless reader
that returns ParseResult dataclasses. Django is not involved.
"""

import datetime
import io
from decimal import Decimal

import openpyxl
import pytest

from api.parsers.jedalnicek_xlsx_parser import JedalnicekXlsxParser, ParsedRow

# ── workbook builders ─────────────────────────────────────────────────────────

HEADERS = [
    "Deň",
    "Chod",
    "Komponent / Jedlo",
    "Alergény",
    "Množstvo",
    "Jednotka",
    "Poznámky",
]


def _make_wb():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb


def _add_instructions(
    ws, year=2026, week=22, date_from="25.05.2026", date_to="29.05.2026"
):
    ws.cell(row=3, column=1, value="Rok")
    ws.cell(row=3, column=2, value=year)
    ws.cell(row=4, column=1, value="Týždeň")
    ws.cell(row=4, column=2, value=week)
    ws.cell(row=5, column=1, value="Dátum od")
    ws.cell(row=5, column=2, value=date_from)
    ws.cell(row=6, column=1, value="Dátum do")
    ws.cell(row=6, column=2, value=date_to)


def _add_diet_headers(ws):
    for i, h in enumerate(HEADERS, 1):
        ws.cell(row=2, column=i, value=h)


def _add_day_block(ws, start_row, day_name, meal_blocks):
    """
    meal_blocks: list of (meal_type_label, [(comp, allergens, amount, unit, notes), ...])
    Returns next free row.
    """
    row = start_row
    ws.cell(row=row, column=1, value=f"  {day_name}")
    row += 1
    for meal_label, items in meal_blocks:
        ws.cell(row=row, column=2, value=meal_label)
        row += 1
        for comp, allergens, amount, unit, notes in items:
            ws.cell(row=row, column=3, value=comp)
            ws.cell(row=row, column=4, value=allergens)
            ws.cell(row=row, column=5, value=amount)
            ws.cell(row=row, column=6, value=unit)
            ws.cell(row=row, column=7, value=notes)
            row += 1
    row += 1  # separator
    return row


def _wb_bytes(wb) -> io.BytesIO:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _minimal_valid_wb(diet_name="Klasik"):
    """Smallest valid workbook: INŠTRUKCIE + one diet sheet with one ingredient."""
    wb = _make_wb()
    ws_i = wb.create_sheet("INŠTRUKCIE")
    _add_instructions(ws_i)
    ws_d = wb.create_sheet(diet_name)
    _add_diet_headers(ws_d)
    _add_day_block(
        ws_d,
        3,
        "PONDELOK",
        [
            ("Raňajky-desiata", [("Ovsená kaša", "7", 150, "g", "")]),
        ],
    )
    return wb


# ── 1. Metadata tests ─────────────────────────────────────────────────────────


@pytest.mark.unit
class TestMetadata:
    def test_reads_year_and_week_number(self):
        wb = _make_wb()
        ws = wb.create_sheet("INŠTRUKCIE")
        _add_instructions(ws, year=2026, week=22)

        result = JedalnicekXlsxParser().parse(_wb_bytes(wb))

        assert result.week_year == 2026
        assert result.week_number == 22

    def test_computes_week_start_and_end_from_iso_week(self):
        wb = _make_wb()
        ws = wb.create_sheet("INŠTRUKCIE")
        _add_instructions(ws, year=2026, week=22)

        result = JedalnicekXlsxParser().parse(_wb_bytes(wb))

        assert result.week_start == datetime.date(2026, 5, 25)
        assert result.week_end == datetime.date(2026, 5, 29)

    def test_missing_instrukcie_sheet_is_fatal_error(self):
        wb = _make_wb()
        wb.create_sheet("Klasik")

        result = JedalnicekXlsxParser().parse(_wb_bytes(wb))

        assert not result.ok
        assert any("INŠTRUKCIE" in e for e in result.errors)

    def test_missing_year_is_fatal_error(self):
        wb = _make_wb()
        ws = wb.create_sheet("INŠTRUKCIE")
        # Only set week, not year — year cell stays empty
        ws.cell(row=4, column=1, value="Týždeň")
        ws.cell(row=4, column=2, value=22)
        ws.cell(row=5, column=1, value="Dátum od")
        ws.cell(row=5, column=2, value="25.05.2026")
        ws.cell(row=6, column=1, value="Dátum do")
        ws.cell(row=6, column=2, value="29.05.2026")

        result = JedalnicekXlsxParser().parse(_wb_bytes(wb))

        assert not result.ok
        assert any("rok" in e.lower() for e in result.errors)

    def test_missing_week_is_fatal_error(self):
        wb = _make_wb()
        ws = wb.create_sheet("INŠTRUKCIE")
        # Only set year, not week — week cell stays empty
        ws.cell(row=3, column=1, value="Rok")
        ws.cell(row=3, column=2, value=2026)
        ws.cell(row=5, column=1, value="Dátum od")
        ws.cell(row=5, column=2, value="25.05.2026")
        ws.cell(row=6, column=1, value="Dátum do")
        ws.cell(row=6, column=2, value="29.05.2026")

        result = JedalnicekXlsxParser().parse(_wb_bytes(wb))

        assert not result.ok
        assert any("týždeň" in e.lower() for e in result.errors)

    def test_year_stored_as_string_is_coerced(self):
        wb = _make_wb()
        ws = wb.create_sheet("INŠTRUKCIE")
        _add_instructions(ws, year="2026", week="22")

        result = JedalnicekXlsxParser().parse(_wb_bytes(wb))

        assert result.week_year == 2026
        assert result.week_number == 22

    def test_date_mismatch_adds_warning(self):
        wb = _make_wb()
        ws = wb.create_sheet("INŠTRUKCIE")
        _add_instructions(ws, year=2026, week=22, date_from="01.01.2025")

        result = JedalnicekXlsxParser().parse(_wb_bytes(wb))

        assert any("dátum" in w.lower() for w in result.warnings)

    def test_date_mismatch_does_not_prevent_parsing(self):
        wb = _make_wb()
        ws_i = wb.create_sheet("INŠTRUKCIE")
        _add_instructions(ws_i, year=2026, week=22, date_from="01.01.2025")
        ws_d = wb.create_sheet("Klasik")
        _add_diet_headers(ws_d)
        _add_day_block(
            ws_d,
            3,
            "PONDELOK",
            [
                ("Raňajky-desiata", [("Kaša", "", 100, "g", "")]),
            ],
        )

        result = JedalnicekXlsxParser().parse(_wb_bytes(wb))

        assert result.ok
        assert len(result.rows) == 1


# ── 2. Sheet handling ─────────────────────────────────────────────────────────


@pytest.mark.unit
class TestSheetHandling:
    def test_instrukcie_sheet_is_not_parsed_as_diet(self):
        wb = _minimal_valid_wb()

        result = JedalnicekXlsxParser().parse(_wb_bytes(wb))

        diet_sheets = {r.diet_sheet for r in result.rows}
        assert "INŠTRUKCIE" not in diet_sheets

    def test_multiple_diet_sheets_all_produce_rows(self):
        wb = _make_wb()
        ws_i = wb.create_sheet("INŠTRUKCIE")
        _add_instructions(ws_i)
        for diet in ("Klasik", "Vege"):
            ws_d = wb.create_sheet(diet)
            _add_diet_headers(ws_d)
            _add_day_block(
                ws_d,
                3,
                "PONDELOK",
                [
                    ("Raňajky-desiata", [("Kaša", "", 100, "g", "")]),
                ],
            )

        result = JedalnicekXlsxParser().parse(_wb_bytes(wb))

        diet_sheets = {r.diet_sheet for r in result.rows}
        assert "Klasik" in diet_sheets
        assert "Vege" in diet_sheets

    def test_empty_diet_sheet_produces_no_rows_and_no_error(self):
        wb = _make_wb()
        ws_i = wb.create_sheet("INŠTRUKCIE")
        _add_instructions(ws_i)
        wb.create_sheet("Klasik")

        result = JedalnicekXlsxParser().parse(_wb_bytes(wb))

        assert result.ok
        assert result.rows == []

    def test_sheet_name_preserved_as_diet_sheet_attribute(self):
        wb = _minimal_valid_wb("No Gluten")

        result = JedalnicekXlsxParser().parse(_wb_bytes(wb))

        assert all(r.diet_sheet == "No Gluten" for r in result.rows)


# ── 3. Day → date mapping ─────────────────────────────────────────────────────


@pytest.mark.unit
class TestDayDateMapping:
    def _parse_single_day(self, day_name):
        wb = _make_wb()
        ws_i = wb.create_sheet("INŠTRUKCIE")
        _add_instructions(ws_i, year=2026, week=22)
        ws_d = wb.create_sheet("Klasik")
        _add_diet_headers(ws_d)
        _add_day_block(
            ws_d,
            3,
            day_name,
            [
                ("Raňajky-desiata", [("Kaša", "", None, "g", "")]),
            ],
        )
        return JedalnicekXlsxParser().parse(_wb_bytes(wb))

    def test_pondelok_maps_to_monday(self):
        result = self._parse_single_day("PONDELOK")
        assert result.rows[0].date == datetime.date(2026, 5, 25)

    def test_utorok_maps_to_tuesday(self):
        result = self._parse_single_day("UTOROK")
        assert result.rows[0].date == datetime.date(2026, 5, 26)

    def test_streda_maps_to_wednesday(self):
        result = self._parse_single_day("STREDA")
        assert result.rows[0].date == datetime.date(2026, 5, 27)

    def test_stvrtok_maps_to_thursday(self):
        result = self._parse_single_day("ŠTVRTOK")
        assert result.rows[0].date == datetime.date(2026, 5, 28)

    def test_piatok_maps_to_friday(self):
        result = self._parse_single_day("PIATOK")
        assert result.rows[0].date == datetime.date(2026, 5, 29)

    def test_all_five_days_produce_correct_dates(self):
        wb = _make_wb()
        ws_i = wb.create_sheet("INŠTRUKCIE")
        _add_instructions(ws_i, year=2026, week=22)
        ws_d = wb.create_sheet("Klasik")
        _add_diet_headers(ws_d)
        row = 3
        for day in ("PONDELOK", "UTOROK", "STREDA", "ŠTVRTOK", "PIATOK"):
            row = _add_day_block(
                ws_d,
                row,
                day,
                [
                    ("Raňajky-desiata", [("Kaša", "", None, "g", "")]),
                ],
            )

        result = JedalnicekXlsxParser().parse(_wb_bytes(wb))

        dates = [r.date for r in result.rows]
        expected = [datetime.date(2026, 5, d) for d in (25, 26, 27, 28, 29)]
        assert dates == expected

    def test_unknown_day_name_warns_and_rows_skipped(self):
        wb = _make_wb()
        ws_i = wb.create_sheet("INŠTRUKCIE")
        _add_instructions(ws_i)
        ws_d = wb.create_sheet("Klasik")
        _add_diet_headers(ws_d)
        _add_day_block(
            ws_d,
            3,
            "SOBOTA",
            [
                ("Raňajky-desiata", [("Kaša", "", None, "g", "")]),
            ],
        )

        result = JedalnicekXlsxParser().parse(_wb_bytes(wb))

        assert any("SOBOTA" in w for w in result.warnings)
        assert result.rows == []

    def test_leading_spaces_in_day_cell_are_stripped(self):
        wb = _make_wb()
        ws_i = wb.create_sheet("INŠTRUKCIE")
        _add_instructions(ws_i, year=2026, week=22)
        ws_d = wb.create_sheet("Klasik")
        _add_diet_headers(ws_d)
        # Extra spaces — still should be recognised
        ws_d.cell(row=3, column=1, value="     PONDELOK   ")
        ws_d.cell(row=4, column=2, value="Raňajky-desiata")
        ws_d.cell(row=5, column=3, value="Kaša")
        ws_d.cell(row=5, column=6, value="g")

        result = JedalnicekXlsxParser().parse(_wb_bytes(wb))

        assert result.rows[0].date == datetime.date(2026, 5, 25)


# ── 4. Meal type → category + variant ────────────────────────────────────────


@pytest.mark.unit
class TestMealTypeMapping:
    def _parse_with_meal(self, meal_label):
        wb = _make_wb()
        ws_i = wb.create_sheet("INŠTRUKCIE")
        _add_instructions(ws_i)
        ws_d = wb.create_sheet("Klasik")
        _add_diet_headers(ws_d)
        _add_day_block(
            ws_d,
            3,
            "PONDELOK",
            [
                (meal_label, [("Jedlo", "", None, "g", "")]),
            ],
        )
        return JedalnicekXlsxParser().parse(_wb_bytes(wb))

    def test_ranajky_desiata_maps_to_breakfast(self):
        result = self._parse_with_meal("Raňajky-desiata")
        row = result.rows[0]
        assert row.category == "breakfast"
        assert row.menu_variant == ""

    def test_polievka_maps_to_lunch_variant_P(self):
        result = self._parse_with_meal("Polievka")
        row = result.rows[0]
        assert row.category == "lunch"
        assert row.menu_variant == "P"

    def test_obed_menu_a_maps_to_lunch_variant_A(self):
        result = self._parse_with_meal("Obed (Menu A)")
        row = result.rows[0]
        assert row.category == "lunch"
        assert row.menu_variant == "A"

    def test_obed_menu_b_maps_to_lunch_variant_B(self):
        result = self._parse_with_meal("Obed (Menu B)")
        row = result.rows[0]
        assert row.category == "lunch"
        assert row.menu_variant == "B"

    def test_obed_menu_c_maps_to_lunch_variant_C(self):
        result = self._parse_with_meal("Obed (Menu C)")
        row = result.rows[0]
        assert row.category == "lunch"
        assert row.menu_variant == "C"

    def test_olovrant_maps_to_snack(self):
        result = self._parse_with_meal("Olovrant")
        row = result.rows[0]
        assert row.category == "snack"
        assert row.menu_variant == ""

    def test_unknown_meal_type_warns_and_row_skipped(self):
        result = self._parse_with_meal("Neznámy chod")
        assert any("Neznámy chod" in w for w in result.warnings)
        assert result.rows == []

    def test_ingredient_without_preceding_meal_type_warns(self):
        wb = _make_wb()
        ws_i = wb.create_sheet("INŠTRUKCIE")
        _add_instructions(ws_i)
        ws_d = wb.create_sheet("Klasik")
        _add_diet_headers(ws_d)
        # Day row but no meal type row before ingredient
        ws_d.cell(row=3, column=1, value="  PONDELOK")
        ws_d.cell(
            row=4, column=3, value="Záhadná polievka"
        )  # ingredient without meal type
        ws_d.cell(row=4, column=6, value="g")

        result = JedalnicekXlsxParser().parse(_wb_bytes(wb))

        assert any("chod" in w.lower() for w in result.warnings)


# ── 5. Ingredient data extraction ────────────────────────────────────────────


@pytest.mark.unit
class TestIngredientExtraction:
    def _parse_ingredient(self, comp, allergens, amount, unit, notes):
        wb = _make_wb()
        ws_i = wb.create_sheet("INŠTRUKCIE")
        _add_instructions(ws_i)
        ws_d = wb.create_sheet("Klasik")
        _add_diet_headers(ws_d)
        _add_day_block(
            ws_d,
            3,
            "PONDELOK",
            [
                ("Raňajky-desiata", [(comp, allergens, amount, unit, notes)]),
            ],
        )
        result = JedalnicekXlsxParser().parse(_wb_bytes(wb))
        return result.rows[0] if result.rows else None

    def test_component_name_extracted(self):
        row = self._parse_ingredient("Ovsená kaša s bobuľami", "7", 150, "g", "")
        assert row.component_name == "Ovsená kaša s bobuľami"

    def test_allergens_extracted(self):
        row = self._parse_ingredient("Kaša", "1, 3, 7", 100, "g", "")
        assert row.allergens == "1, 3, 7"

    def test_allergens_empty_is_empty_string(self):
        row = self._parse_ingredient("Jablko", "", 50, "g", "")
        assert row.allergens == ""

    def test_amount_integer_extracted_as_decimal(self):
        row = self._parse_ingredient("Kaša", "", 150, "g", "")
        assert row.amount == Decimal("150")

    def test_amount_float_extracted_as_decimal(self):
        row = self._parse_ingredient("Kaša", "", 150.5, "g", "")
        assert row.amount == Decimal("150.5")

    def test_amount_none_is_none(self):
        row = self._parse_ingredient("Kaša", "", None, "g", "")
        assert row.amount is None

    def test_amount_empty_string_is_none(self):
        row = self._parse_ingredient("Kaša", "", "", "g", "")
        assert row.amount is None

    def test_amount_zero_is_decimal_zero(self):
        row = self._parse_ingredient("Kaša", "", 0, "g", "")
        assert row.amount == Decimal("0")

    def test_amount_non_numeric_string_warns_and_is_none(self):
        wb = _make_wb()
        ws_i = wb.create_sheet("INŠTRUKCIE")
        _add_instructions(ws_i)
        ws_d = wb.create_sheet("Klasik")
        _add_diet_headers(ws_d)
        ws_d.cell(row=3, column=1, value="  PONDELOK")
        ws_d.cell(row=4, column=2, value="Raňajky-desiata")
        ws_d.cell(row=5, column=3, value="Kaša")
        ws_d.cell(row=5, column=5, value="veľa")  # non-numeric
        ws_d.cell(row=5, column=6, value="g")

        result = JedalnicekXlsxParser().parse(_wb_bytes(wb))

        assert any("množstvo" in w.lower() for w in result.warnings)
        assert result.rows[0].amount is None

    def test_unit_g_extracted(self):
        row = self._parse_ingredient("Kaša", "", 100, "g", "")
        assert row.unit == "g"

    def test_unit_ml_extracted(self):
        row = self._parse_ingredient("Polievka", "", 200, "ml", "")
        assert row.unit == "ml"

    def test_unit_ks_extracted(self):
        row = self._parse_ingredient("Vajce", "", 1, "ks", "")
        assert row.unit == "ks"

    def test_unit_empty_defaults_to_g(self):
        row = self._parse_ingredient("Kaša", "", 100, None, "")
        assert row.unit == "g"

    def test_unit_empty_string_defaults_to_g(self):
        row = self._parse_ingredient("Kaša", "", 100, "", "")
        assert row.unit == "g"

    def test_notes_extracted(self):
        row = self._parse_ingredient("Kaša", "", None, "g", "Menu B, 275 g spolu")
        assert row.notes == "Menu B, 275 g spolu"

    def test_notes_none_is_empty_string(self):
        row = self._parse_ingredient("Kaša", "", None, "g", None)
        assert row.notes == ""


# ── 6. Error handling ─────────────────────────────────────────────────────────


@pytest.mark.unit
class TestErrorHandling:
    def test_non_xlsx_bytes_returns_error(self):
        result = JedalnicekXlsxParser().parse(io.BytesIO(b"not an xlsx file at all"))

        assert not result.ok
        assert any("súbor" in e.lower() or "xlsx" in e.lower() for e in result.errors)

    def test_empty_bytes_returns_error(self):
        result = JedalnicekXlsxParser().parse(io.BytesIO(b""))

        assert not result.ok

    def test_parse_result_has_no_rows_when_fatal_error(self):
        result = JedalnicekXlsxParser().parse(io.BytesIO(b"garbage"))

        assert result.rows == []

    def test_warnings_do_not_make_result_not_ok(self):
        wb = _make_wb()
        ws_i = wb.create_sheet("INŠTRUKCIE")
        _add_instructions(
            ws_i, year=2026, week=22, date_from="01.01.2025"
        )  # triggers warning
        ws_d = wb.create_sheet("Klasik")
        _add_diet_headers(ws_d)
        _add_day_block(
            ws_d,
            3,
            "PONDELOK",
            [
                ("Raňajky-desiata", [("Kaša", "", 100, "g", "")]),
            ],
        )

        result = JedalnicekXlsxParser().parse(_wb_bytes(wb))

        assert result.ok
        assert len(result.warnings) > 0


# ── 7. ParseResult structure ──────────────────────────────────────────────────


@pytest.mark.unit
class TestParseResultStructure:
    def test_result_is_ok_for_valid_file(self):
        result = JedalnicekXlsxParser().parse(_wb_bytes(_minimal_valid_wb()))
        assert result.ok

    def test_result_has_empty_errors_for_valid_file(self):
        result = JedalnicekXlsxParser().parse(_wb_bytes(_minimal_valid_wb()))
        assert result.errors == []

    def test_row_count_matches_ingredient_rows_in_workbook(self):
        wb = _make_wb()
        ws_i = wb.create_sheet("INŠTRUKCIE")
        _add_instructions(ws_i)
        ws_d = wb.create_sheet("Klasik")
        _add_diet_headers(ws_d)
        _add_day_block(
            ws_d,
            3,
            "PONDELOK",
            [
                (
                    "Raňajky-desiata",
                    [
                        ("Kaša", "7", 150, "g", ""),
                        ("Jablko", "", 50, "g", ""),
                    ],
                ),
                (
                    "Polievka",
                    [
                        ("Paradajková polievka", "", 200, "ml", ""),
                    ],
                ),
            ],
        )

        result = JedalnicekXlsxParser().parse(_wb_bytes(wb))

        assert len(result.rows) == 3

    def test_parsed_row_has_all_required_fields(self):
        result = JedalnicekXlsxParser().parse(_wb_bytes(_minimal_valid_wb()))
        row = result.rows[0]

        assert isinstance(row, ParsedRow)
        assert isinstance(row.date, datetime.date)
        assert isinstance(row.diet_sheet, str)
        assert isinstance(row.category, str)
        assert isinstance(row.menu_variant, str)
        assert isinstance(row.component_name, str)
        assert isinstance(row.allergens, str)
        assert row.amount is None or isinstance(row.amount, Decimal)
        assert isinstance(row.unit, str)
        assert isinstance(row.notes, str)


# ── 8. Integration: parse the generated template ──────────────────────────────

TEMPLATE_PATH = (
    "/home/tomas-magula/Documents/Projects/ZB/zdravy-projekt"
    "/test/jedalnicky/jedalnicky_template_W22_2026.xlsx"
)


@pytest.mark.unit
class TestIntegrationGeneratedTemplate:
    @pytest.fixture(autouse=True)
    def require_template(self):
        import os

        if not os.path.exists(TEMPLATE_PATH):
            pytest.skip(
                "Template file not generated yet — run generate_jedalnicky_template.py first"
            )

    def test_parse_succeeds_with_no_errors(self):
        with open(TEMPLATE_PATH, "rb") as f:
            result = JedalnicekXlsxParser().parse(f)
        assert result.ok, f"Errors: {result.errors}"

    def test_week_metadata_correct(self):
        with open(TEMPLATE_PATH, "rb") as f:
            result = JedalnicekXlsxParser().parse(f)
        assert result.week_year == 2026
        assert result.week_number == 22
        assert result.week_start == datetime.date(2026, 5, 25)
        assert result.week_end == datetime.date(2026, 5, 29)

    def test_all_diet_sheets_produce_rows(self):
        with open(TEMPLATE_PATH, "rb") as f:
            result = JedalnicekXlsxParser().parse(f)
        diet_sheets = {r.diet_sheet for r in result.rows}
        # At minimum Klasik and Vege should be present
        assert "Klasik" in diet_sheets
        assert "Vege" in diet_sheets

    def test_all_five_days_represented_in_klasik(self):
        with open(TEMPLATE_PATH, "rb") as f:
            result = JedalnicekXlsxParser().parse(f)
        klasik_dates = {r.date for r in result.rows if r.diet_sheet == "Klasik"}
        expected_dates = {datetime.date(2026, 5, d) for d in (25, 26, 27, 28, 29)}
        assert klasik_dates == expected_dates

    def test_all_meal_categories_present_in_klasik(self):
        with open(TEMPLATE_PATH, "rb") as f:
            result = JedalnicekXlsxParser().parse(f)
        klasik_categories = {
            r.category for r in result.rows if r.diet_sheet == "Klasik"
        }
        assert "breakfast" in klasik_categories
        assert "lunch" in klasik_categories
        assert "snack" in klasik_categories

    def test_menu_variants_abc_present_in_klasik(self):
        with open(TEMPLATE_PATH, "rb") as f:
            result = JedalnicekXlsxParser().parse(f)
        klasik_lunch_variants = {
            r.menu_variant
            for r in result.rows
            if r.diet_sheet == "Klasik" and r.category == "lunch"
        }
        # Klasik includes Menu A, B and Polievka (P)
        assert "A" in klasik_lunch_variants
        assert "B" in klasik_lunch_variants
        assert "P" in klasik_lunch_variants

    def test_amounts_are_none_because_template_is_empty(self):
        with open(TEMPLATE_PATH, "rb") as f:
            result = JedalnicekXlsxParser().parse(f)
        # All amounts must be None — nobody filled in the template
        assert all(r.amount is None for r in result.rows)

    def test_units_are_valid_values(self):
        with open(TEMPLATE_PATH, "rb") as f:
            result = JedalnicekXlsxParser().parse(f)
        valid_units = {"g", "ml", "ks"}
        for row in result.rows:
            assert (
                row.unit in valid_units
            ), f"Unexpected unit '{row.unit}' for {row.component_name}"

    def test_units_include_ml_for_soups(self):
        with open(TEMPLATE_PATH, "rb") as f:
            result = JedalnicekXlsxParser().parse(f)
        polievka_units = {
            r.unit
            for r in result.rows
            if r.category == "lunch" and r.menu_variant == "P"
        }
        assert "ml" in polievka_units

    def test_no_empty_component_names(self):
        with open(TEMPLATE_PATH, "rb") as f:
            result = JedalnicekXlsxParser().parse(f)
        empty_names = [r for r in result.rows if not r.component_name.strip()]
        assert empty_names == [], f"Found rows with empty component name: {empty_names}"

    def test_total_row_count_is_reasonable(self):
        with open(TEMPLATE_PATH, "rb") as f:
            result = JedalnicekXlsxParser().parse(f)
        # Each diet has ~40 rows per week × at least 6 diets → 240+
        assert len(result.rows) >= 200, f"Only {len(result.rows)} rows parsed"
