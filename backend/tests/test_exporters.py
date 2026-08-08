"""Tests for Report Exporters."""

import datetime
from decimal import Decimal
from io import BytesIO

import pytest
from django.contrib.auth.models import User
from openpyxl import load_workbook

from api.exporters.gramage_dashboard_xlsx_exporter import GramageDashboardXLSXExporter
from api.models import (
    Celok,
    DailyMealPlan,
    DailyOrder,
    MealPlanItem,
    MealTemplate,
    PortionType,
    Prevadzka,
)
from api.services.meal_plan_service import MealPlanService


@pytest.mark.django_db
class TestGramageDashboardExports:
    def test_dashboard_subtotals_exclude_diets_and_export_to_xlsx(self):
        user = User.objects.create_user(
            username="dashboard@example.com",
            email="dashboard@example.com",
            first_name="Dashboard",
            last_name="Client",
        )
        portion_type = PortionType.objects.create(
            name="Škôlka", coefficient=Decimal("1.0000"), sort_order=1
        )
        plan = DailyMealPlan.objects.create(
            date=datetime.date(2024, 1, 15),
            created_by=user,
        )
        breakfast = MealTemplate.objects.create(
            category="breakfast_snack",
            name="Kaša",
            weight_label="100g",
            base_weight_grams=Decimal("100.00"),
        )
        lunch_a = MealTemplate.objects.create(
            category="main_course",
            name="Menu A",
            weight_label="200g",
            base_weight_grams=Decimal("200.00"),
            menu_variant="A",
        )
        lunch_b = MealTemplate.objects.create(
            category="main_course",
            name="Menu B",
            weight_label="250g",
            base_weight_grams=Decimal("250.00"),
            menu_variant="B",
        )
        snack = MealTemplate.objects.create(
            category="afternoon_snack",
            name="Ovocie",
            weight_label="50g",
            base_weight_grams=Decimal("50.00"),
        )
        MealPlanItem.objects.create(
            meal_plan=plan,
            template=breakfast,
            category="breakfast_snack",
            menu_variant="",
        )
        MealPlanItem.objects.create(
            meal_plan=plan, template=lunch_a, category="main_course", menu_variant="A"
        )
        MealPlanItem.objects.create(
            meal_plan=plan, template=lunch_b, category="main_course", menu_variant="B"
        )
        MealPlanItem.objects.create(
            meal_plan=plan,
            template=snack,
            category="afternoon_snack",
            menu_variant="",
        )
        plan.enrolled_counts.create(portion_type=portion_type, count=1)

        celok = Celok.objects.create(nazov="Dashboard celok")
        prevadzka = Prevadzka.objects.create(celok=celok, nazov="Dashboard prevádzka")
        DailyOrder.objects.create(
            user=user,
            date=plan.date,
            prevadzka=prevadzka,
            data={
                "breakfast": {
                    "Škôlka": {"menuCounts": {"A": 4}, "diets": {"Vegan": 1}}
                },
                "lunch": {
                    "Škôlka": {
                        "menuCounts": {"A": 5, "B": 2},
                        "diets": {"Bezlepková": 2},
                    }
                },
                "olovrant": {"Škôlka": {"menuCounts": {"A": 3}, "diets": {}}},
            },
        )

        data = MealPlanService.gramage_dashboard(plan.date.isoformat())
        row = data["rows"][0]
        standard_rows = [sr for sr in row["sub_rows"] if sr["type"] == "standard"]

        assert row["standard_total_count"] == 11
        assert sorted(sr["count"] for sr in standard_rows) == [2, 3, 3, 3]
        assert {sr["label"]: sr["count"] for sr in standard_rows} == {
            "Škôlka - Raňajky-desiata": 3,
            "Škôlka - Hlavný chod Menu A": 3,
            "Škôlka - Hlavný chod Menu B": 2,
            "Škôlka - Olovrant": 3,
        }
        assert row["standard_col_grams"] == [
            ["300.00"],
            ["600.00"],
            ["500.00"],
            ["150.00"],
        ]
        assert row["diet_summary_rows"] == [
            {
                "name": "Bezlepková",
                "color": "#FDE68A",
                "base_colors": [],
                "count": 2,
                "col_grams": [["0.00"], ["400.00"], ["0.00"], ["0.00"]],
            },
            {
                "name": "Vegan",
                "color": "#FDE68A",
                "base_colors": [],
                "count": 1,
                "col_grams": [["100.00"], ["0.00"], ["0.00"], ["0.00"]],
            },
        ]
        assert data["totals"] == [["300.00"], ["600.00"], ["500.00"], ["150.00"]]

        workbook = load_workbook(BytesIO(GramageDashboardXLSXExporter(data).generate()))
        worksheet = workbook.active
        exported_values = [row for row in worksheet.iter_rows(values_only=True)]

        assert any(
            values[0] == "Súčet bez diét" and values[1] == 11
            for values in exported_values
            if values[0]
        )
        assert any(
            values[0] == "Bezlepková" and values[1] == 2
            for values in exported_values
            if values[0]
        )
        assert any(
            values[0] == "Vegan" and values[1] == 1
            for values in exported_values
            if values[0]
        )
