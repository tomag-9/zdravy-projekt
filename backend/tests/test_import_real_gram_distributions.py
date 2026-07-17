from io import StringIO

import openpyxl
import pytest
from django.core.management import call_command

from api.models import DailyMealPlan, MealCategory, MealPlanItem, MealTemplate

pytestmark = pytest.mark.django_db


def _write_workbook(path, target_date, row_values):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(
        [
            target_date,
            "Brokolicová",
            "Morčacie",
            "Ryža",
            "Miešaný šalát",
            None,
            "Koláč",
        ]
    )
    sheet.append(["KLASIK", *row_values])
    workbook.save(path)


def test_import_real_gram_distributions_creates_daily_plan(tmp_path):
    path = tmp_path / "8.7.2026_tabulka.xlsx"
    _write_workbook(path, "8.7.2026", [200, 90, 110, 25, 0, 75])

    out = StringIO()
    call_command("import_real_gram_distributions", str(path), stdout=out)

    plan = DailyMealPlan.objects.get(date="2026-07-08")
    items = {
        item.category: item
        for item in MealPlanItem.objects.filter(meal_plan=plan).select_related(
            "template"
        )
    }

    assert set(items) == {
        MealCategory.SOUP,
        MealCategory.MAIN_COURSE,
        MealCategory.AFTERNOON_SNACK,
    }
    assert items[MealCategory.SOUP].template.components == [
        {"label": "Brokolicová", "grams": "200", "unit": "ml"}
    ]
    assert items[MealCategory.MAIN_COURSE].template.components == [
        {"label": "Morčacie", "grams": "90", "unit": "g"},
        {"label": "Ryža", "grams": "110", "unit": "g"},
        {"label": "Miešaný šalát", "grams": "25", "unit": "g"},
    ]
    assert items[MealCategory.AFTERNOON_SNACK].template.components == [
        {"label": "Koláč", "grams": "75", "unit": "g"}
    ]
    assert "2026-07-08" in out.getvalue()


def test_import_real_gram_distributions_keeps_piece_exception_idempotent(tmp_path):
    path = tmp_path / "6.7.2026_tabulka.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "6.7.2026",
            "Kukuricová polievka",
            "Zeleninové rizoto",
            "Syr",
            None,
            "Grahamové pečivo",
            "Smotanová nátierka",
        ]
    )
    sheet.append(["KLASIK", 200, 200, 10, 0, 1, 25])
    workbook.save(path)

    call_command("import_real_gram_distributions", str(path), stdout=StringIO())
    call_command("import_real_gram_distributions", str(path), stdout=StringIO())

    plan = DailyMealPlan.objects.get(date="2026-07-06")
    assert MealPlanItem.objects.filter(meal_plan=plan).count() == 3
    assert MealTemplate.objects.filter(name="Real 2026-07-06 Olovrant").count() == 1

    snack = MealPlanItem.objects.get(
        meal_plan=plan,
        category=MealCategory.AFTERNOON_SNACK,
    ).template
    assert snack.components == [
        {"label": "Smotanová nátierka", "grams": "25", "unit": "g"}
    ]
    assert snack.unit_exception == {
        "component_label": "Grahamové pečivo",
        "unit": "ks",
        "counts_by_portion_type": {
            "Jasle": "1",
            "Škôlka": "1",
            # Predškolák dedí kusový pomer po `ZŠ 1.stupeň` — je s ním gramážovo
            # zhodný (250 g), oddelený je len kvôli fakturácii.
            "Predškolák": "1.5",
            "ZŠ 1.stupeň": "1.5",
            "ZŠ 2.stupeň": "1.5",
            "Dospelý (SŠ)": "2",
        },
    }


def test_import_real_gram_distributions_covers_week_28_snack_split(tmp_path):
    path = tmp_path / "7.7.2026_tabulka.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "7.7.2026",
            "Polievka",
            "Mäso",
            "Príloha",
            "Šalát",
            "Nátierka",
            "Zelenina",
        ]
    )
    sheet.append(["KLASIK", 200, 90, 110, 25, 50, 25])
    workbook.save(path)

    call_command("import_real_gram_distributions", str(path), stdout=StringIO())

    snack = MealPlanItem.objects.get(
        meal_plan__date="2026-07-07",
        category=MealCategory.AFTERNOON_SNACK,
    ).template
    assert snack.components == [
        {"label": "Nátierka", "grams": "50", "unit": "g"},
        {"label": "Zelenina", "grams": "25", "unit": "g"},
    ]
    assert snack.unit_exception is None
