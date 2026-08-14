"""Reconciling a PRODUCTION day: dashboard dump in, gramage-aligned columns out.

Production meal plans are built from generic gramage templates (`Hlavná zložka`,
`Príloha`), so the dish-name matching that works against a hand-entered dev plan
resolves almost nothing there. These tests pin the two things that make a prod
comparison meaningful: the base-gramage alignment, and the fact that a meal the
workbook has no column for (raňajky) is dropped instead of compared against 0.
"""

import json
from decimal import Decimal
from io import StringIO

import pytest
from django.core.management import call_command
from openpyxl import Workbook

from api.management.commands import reconcile_real as R

pytestmark = pytest.mark.unit

# One column per dish, mirroring Hárok1: soup, main, side, cheese, spacer, snack.
DISHES = ["Kurací vývar", "Šošovicové ragú", "Špagety", "Syr", None, "Bublanina"]
KLASIK = [200, 90, 110, 10, 0, 75]


def _sheet(facility_rows):
    """A minimal Hárok1: header, base-gramage legend, then facility blocks."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Hárok1"
    ws.cell(row=1, column=1, value="14.8.2026")
    for offset, dish in enumerate(DISHES):
        if dish:
            ws.cell(row=1, column=2 + offset, value=dish)
    legend = [
        ("KLASIK", 1.0),
        ("JASLE", 0.75),
        ("1. STUPEŇ", 1.25),
        ("2. STUPEŇ", 1.5),
        ("DOSPELÁ", 2.0),
    ]
    for index, (label, factor) in enumerate(legend):
        ws.cell(row=2 + index, column=1, value=label)
        for offset, base in enumerate(KLASIK):
            ws.cell(row=2 + index, column=2 + offset, value=base * factor)
    row = 7
    for name, address, count, grams in facility_rows:
        ws.cell(row=row, column=1, value=name)
        for offset, value in enumerate(grams):
            ws.cell(row=row, column=2 + offset, value=value)
        ws.cell(row=row + 1, column=1, value=address)
        ws.cell(row=row + 2, column=1, value=count)
        row += 3
    return wb, ws


def _col_groups(labels_by_meal):
    return [
        {
            "key": meal,
            "label": meal,
            "meal": meal,
            "components": [
                {"label": label, "base_grams": str(base), "unit": "g"}
                for label, base in components
            ],
        }
        for meal, components in labels_by_meal
    ]


PROD_COL_GROUPS = _col_groups(
    [
        ("breakfast_snack", [("Raňajky-desiata spolu", 75)]),
        ("soup", [("Hlavná zložka", 200)]),
        (
            "main_course",
            [("Hlavná zložka", 90), ("Príloha", 110), ("Syr", 10)],
        ),
        ("afternoon_snack", [("Hlavná zložka", 75)]),
    ]
)


def test_generic_prod_components_align_on_base_gramage():
    """Nameless prod components bind to the right columns via the KLASIK legend."""
    _, ws = _sheet([("Škôlka", "Ulica 1", 10, [2000, 900, 1100, 100, 0, 750])])

    columns, strategy = R._align_components(ws, PROD_COL_GROUPS)

    assert strategy == "base-gramage"
    # raňajky (75 g) must NOT steal the olovrant column despite the equal gramage:
    # the alignment keeps serving order, so it drops out with no counterpart.
    assert columns == [None, 2, 3, 4, 5, 7]


def test_breakfast_is_dropped_rather_than_compared_against_zero():
    """Hárok1 carries no raňajky column, so that meal type is not reconcilable."""
    _, ws = _sheet([("Škôlka", "Ulica 1", 10, [2000, 900, 1100, 100, 0, 750])])
    columns, _ = R._align_components(ws, PROD_COL_GROUPS)

    meal_types = set(R._column_meal_types(ws, PROD_COL_GROUPS, columns).values())

    assert meal_types == {"lunch", "snack"}


def test_dish_names_win_when_they_resolve_at_least_as_much():
    """A dev plan naming the actual dishes keeps the sharper matching key."""
    named = _col_groups(
        [
            ("soup", [("Kurací vývar", 200)]),
            (
                "main_course",
                [("Šošovicové ragú", 90), ("Špagety", 110), ("Syr", 10)],
            ),
            ("afternoon_snack", [("Bublanina", 75)]),
        ]
    )
    _, ws = _sheet([("Škôlka", "Ulica 1", 10, [2000, 900, 1100, 100, 0, 750])])

    columns, strategy = R._align_components(ws, named)

    assert strategy == "dish-name"
    assert columns == [2, 3, 4, 5, 7]


def test_gram_values_sum_the_whole_block_per_aligned_column():
    _, ws = _sheet([("Škôlka", "Ulica 1", 10, [2000, 900, 1100, 100, 0, 750])])
    columns, _ = R._align_components(ws, PROD_COL_GROUPS)

    values = R._real_gram_values(ws, [7], columns)

    assert values == [
        None,
        Decimal("2000"),
        Decimal("900"),
        Decimal("1100"),
        Decimal("100"),
        Decimal("750"),
    ]


def test_dashboard_dump_reconciles_without_touching_the_database(tmp_path):
    """`--dashboard` is the prod path: app side comes from the dumped JSON."""
    wb, _ = _sheet(
        [
            ("Škôlka", "Ulica 1", 10, [2000, 900, 1100, 100, 0, 750]),
            ("Jasle", "Ulica 2", 4, [800, 360, 440, 40, 0, 300]),
        ]
    )
    workbook_path = tmp_path / "14.8.2026_tabuľka.xlsx"
    wb.save(workbook_path)

    dump = {
        "meal_plan_id": 12,
        "col_groups": PROD_COL_GROUPS,
        "rows": [
            {
                "client": "Škôlka",
                "standard_col_grams": [
                    ["750.00"],
                    ["2000.00"],
                    ["900.00", "1100.00", "100.00"],
                    ["750.00"],
                ],
                "diet_summary_rows": [],
                "sub_rows": [
                    {"meal": "main_course", "count": 10},
                    {"meal": "afternoon_snack", "count": 10},
                    {"meal": "breakfast_snack", "count": 10},
                ],
            },
            {
                "client": "Jasle",
                # One olovrant short of reality → a Tier-1 count diff, and the
                # Tier-2 gram gap it explains must not be double-reported.
                "standard_col_grams": [
                    ["300.00"],
                    ["800.00"],
                    ["360.00", "440.00", "40.00"],
                    ["225.00"],
                ],
                "diet_summary_rows": [],
                "sub_rows": [
                    {"meal": "main_course", "count": 4},
                    {"meal": "afternoon_snack", "count": 3},
                ],
            },
        ],
    }
    dump_path = tmp_path / "prod_dash.json"
    dump_path.write_text(json.dumps(dump), encoding="utf-8")

    out = StringIO()
    call_command(
        "reconcile_real",
        date="2026-08-14",
        workbook=str(workbook_path),
        dashboard=str(dump_path),
        stdout=out,
        stderr=StringIO(),
    )
    report = json.loads(out.getvalue())

    assert report["app_source"].startswith("dashboard dump")
    assert report["column_alignment"] == "base-gramage"
    assert report["meal_types_compared"] == ["lunch", "snack"]
    assert report["meal_types_skipped"] == ["breakfast"]
    # Škôlka matches on both meals; Jasle is one olovrant short.
    assert [
        (f["facility"], f["meal_type"], f["diff"], f["status"])
        for f in report["tier1_counts"]["matched"]
    ] == [
        ("Jasle", "lunch", "0", "OK"),
        ("Jasle", "snack", "-1", "FAIL"),
        ("Škôlka", "lunch", "0", "OK"),
        ("Škôlka", "snack", "0", "OK"),
    ]
    assert report["tier2_gramage"]["diffs"] == []
